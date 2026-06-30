from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.plantilla_generator.replacements import ReplacementRule, build_rule


class DictionaryLoader:
    REQUIRED_COLUMNS = {"categoria", "valor", "reemplazo", "confianza"}

    def load(self, path: str | Path, mode: str = "markers") -> list[ReplacementRule]:
        dictionary_path = Path(path)
        suffix = dictionary_path.suffix.lower()
        if suffix == ".json":
            rows = self._load_json(dictionary_path)
        elif suffix in {".xlsx", ".xlsm"}:
            rows = self._load_workbook(dictionary_path)
        elif suffix == ".csv":
            rows = self._load_csv(dictionary_path)
        else:
            raise ValueError(f"Diccionario no soportado: {dictionary_path}")

        rules: list[ReplacementRule] = []
        for row in rows:
            rule = build_rule(
                row.get("categoria"),
                row.get("valor"),
                row.get("reemplazo"),
                row.get("confianza"),
                mode,
            )
            if rule:
                rules.append(rule)
        return self._deduplicate(rules)

    def _load_json(self, path: Path) -> list[dict[str, Any]]:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(payload, dict):
            for key in ("registros", "items", "diccionario", "campos"):
                if isinstance(payload.get(key), list):
                    return [self._normalize_row(item) for item in payload[key]]
            return [self._normalize_row(payload)]
        if isinstance(payload, list):
            return [self._normalize_row(item) for item in payload]
        raise ValueError("El diccionario JSON debe ser una lista u objeto.")

    def _load_workbook(self, path: Path) -> list[dict[str, Any]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [self._normalize_header(value) for value in rows[0]]
            result: list[dict[str, Any]] = []
            for values in rows[1:]:
                result.append(
                    {
                        headers[index]: value
                        for index, value in enumerate(values)
                        if index < len(headers) and headers[index]
                    }
                )
            return result
        finally:
            workbook.close()

    def _load_csv(self, path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [self._normalize_row(row) for row in csv.DictReader(handle)]

    def _normalize_row(self, row: Any) -> dict[str, Any]:
        if not isinstance(row, dict):
            return {}
        return {
            self._normalize_header(key): value
            for key, value in row.items()
        }

    def _normalize_header(self, value: Any) -> str:
        return str(value or "").strip().casefold().replace(" ", "_")

    def _deduplicate(self, rules: list[ReplacementRule]) -> list[ReplacementRule]:
        selected: dict[tuple[str, str], ReplacementRule] = {}
        for rule in sorted(rules, key=lambda item: len(item.value), reverse=True):
            selected.setdefault((rule.category, rule.value), rule)
        return list(selected.values())
