from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from lxml import etree as ET

from app.plantilla_generator.replacements import ReplacementRule, apply_text_replacements
from app.services.excel_fill_service import ExcelFillService


class ExcelCleaner:
    def clean(
        self,
        source: str | Path,
        destination: str | Path,
        rules: list[ReplacementRule],
    ) -> tuple[int, list[str]]:
        source_path = Path(source)
        destination_path = Path(destination)
        workbook = load_workbook(
            source_path,
            keep_vba=source_path.suffix.lower() == ".xlsm",
            data_only=False,
            read_only=True,
        )
        total = 0
        values: list[str] = []
        updates: dict[str, dict[str, str]] = {}
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if not isinstance(cell.value, str):
                            continue
                        new_text, count, replaced = apply_text_replacements(
                            cell.value,
                            rules,
                        )
                        if count:
                            updates.setdefault(sheet.title, {})[cell.coordinate] = new_text
                            total += count
                            values.extend(replaced)
        finally:
            workbook.close()

        self._write_updated_package(source_path, destination_path, updates)
        return total, sorted(set(values))

    def _write_updated_package(
        self,
        source: Path,
        destination: Path,
        updates: dict[str, dict[str, str]],
    ) -> None:
        destination.parent.mkdir(parents=True, exist_ok=True)
        helper = ExcelFillService()
        replacements: dict[str, bytes] = {}
        with ZipFile(source) as archive:
            sheet_paths = helper._sheet_paths(archive)
            for sheet_name, cells in updates.items():
                sheet_path = sheet_paths.get(sheet_name)
                if not sheet_path:
                    continue
                sheet_xml = ET.fromstring(archive.read(sheet_path))
                for coordinate, value in cells.items():
                    helper._set_cell_value(
                        sheet_xml,
                        coordinate,
                        value,
                        "string",
                    )
                replacements[sheet_path] = ET.tostring(
                    sheet_xml,
                    encoding="utf-8",
                    xml_declaration=True,
                    standalone=True,
                )
        helper._write_package(source, destination, replacements)
