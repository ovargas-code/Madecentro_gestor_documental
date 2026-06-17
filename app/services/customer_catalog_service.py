from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook


CUSTOMER_KEYS = {
    "cliente_razon_social",
    "cliente_nit",
    "cliente_anio_vinculacion",
    "cliente_ano_vinculacion",
}


class CustomerCatalogService:
    REQUIRED_COLUMNS = {
        "cliente_razon_social",
        "cliente_nit",
        "cliente_anio_vinculacion",
    }

    def read_excel(self, path: str | Path) -> list[dict[str, str]]:
        source = Path(path)
        if not source.is_file():
            raise FileNotFoundError(f"No existe el catálogo de clientes: {source}")
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                raise ValueError("El catálogo de clientes está vacío.")
            normalized_headers = [
                self._normalize_header(value)
                for value in headers
            ]
            missing = self.REQUIRED_COLUMNS - set(normalized_headers)
            if missing:
                raise ValueError(
                    "Faltan columnas requeridas: " + ", ".join(sorted(missing))
                )
            positions = {
                header: normalized_headers.index(header)
                for header in self.REQUIRED_COLUMNS
            }
            customers: list[dict[str, str]] = []
            for row_number, row in enumerate(rows, start=2):
                reason = self._value(row, positions["cliente_razon_social"])
                nit = self._value(row, positions["cliente_nit"])
                year = self._value(row, positions["cliente_anio_vinculacion"])
                if not reason and not nit:
                    continue
                if not reason or not nit or not year:
                    raise ValueError(
                        f"La fila {row_number} tiene datos incompletos."
                    )
                customers.append(
                    {
                        "cliente_razon_social": reason,
                        "cliente_nit": nit,
                        "cliente_anio_vinculacion": year,
                        "cliente_ano_vinculacion": year,
                        "nit_normalizado": self.normalize_nit(nit),
                    }
                )
            if not customers:
                raise ValueError("El catálogo no contiene clientes válidos.")
            return customers
        finally:
            workbook.close()

    def upsert_excel(
        self,
        path: str | Path,
        customer: dict[str, str],
    ) -> bool:
        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)
        headers = [
            "cliente_razon_social",
            "cliente_nit",
            "cliente_anio_vinculacion",
        ]
        if target.is_file():
            workbook = load_workbook(target)
        else:
            workbook = Workbook()
            workbook.active.append(headers)

        try:
            sheet = workbook.active
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not first_row or not any(first_row):
                sheet.append(headers)
                first_row = tuple(headers)
            normalized_headers = [self._normalize_header(value) for value in first_row]
            missing = self.REQUIRED_COLUMNS - set(normalized_headers)
            if missing:
                raise ValueError(
                    "Faltan columnas requeridas: " + ", ".join(sorted(missing))
                )
            positions = {
                header: normalized_headers.index(header) + 1
                for header in self.REQUIRED_COLUMNS
            }
            normalized_nit = customer["nit_normalizado"]
            row_to_update: int | None = None
            nit_column = positions["cliente_nit"]
            for row_number in range(2, sheet.max_row + 1):
                existing_nit = self._value(
                    tuple(sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)),
                    nit_column - 1,
                )
                if self.normalize_nit(existing_nit) == normalized_nit:
                    row_to_update = row_number
                    break

            row_number = row_to_update or sheet.max_row + 1
            sheet.cell(row_number, positions["cliente_razon_social"]).value = customer[
                "cliente_razon_social"
            ]
            sheet.cell(row_number, positions["cliente_nit"]).value = customer[
                "cliente_nit"
            ]
            sheet.cell(row_number, positions["cliente_anio_vinculacion"]).value = customer[
                "cliente_anio_vinculacion"
            ]
            workbook.save(target)
            return row_to_update is not None
        finally:
            workbook.close()

    def normalize_nit(self, value: object) -> str:
        return re.sub(r"[^0-9]", "", str(value or ""))

    def _normalize_header(self, value: object) -> str:
        normalized = unicodedata.normalize("NFKD", str(value or "").casefold())
        ascii_value = "".join(
            char
            for char in normalized
            if not unicodedata.combining(char)
        )
        return re.sub(r"[^a-z0-9]+", "_", ascii_value).strip("_")

    def _value(self, row: tuple[object, ...], index: int) -> str:
        if index >= len(row) or row[index] is None:
            return ""
        value = row[index]
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()
