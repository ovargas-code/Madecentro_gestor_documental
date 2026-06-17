from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import fitz
from lxml import etree
from openpyxl import load_workbook


WORD_NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
}


class DocumentIdentityService:
    """Builds stable structural identities for supported form formats."""

    def fingerprint(
        self,
        path: str | Path,
        payload: dict[str, Any] | None = None,
    ) -> str:
        source = Path(path)
        suffix = source.suffix.lower()
        if suffix == ".pdf":
            structure = self._pdf_structure(source)
        elif suffix == ".xlsx":
            structure = self._excel_structure(source, payload or {})
        elif suffix == ".docx":
            structure = self._word_structure(source)
        else:
            raise ValueError("Formato no soportado para identificar la plantilla.")
        encoded = json.dumps(
            structure,
            ensure_ascii=True,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()

    def matches(
        self,
        path: str | Path,
        payload: dict[str, Any],
    ) -> bool:
        expected = str(payload.get("template_fingerprint") or "")
        if not expected:
            return self.compatibility_score(path, payload) > 0
        return self.fingerprint(path, payload) == expected

    def compatibility_score(
        self,
        path: str | Path,
        payload: dict[str, Any],
    ) -> int:
        source = Path(path)
        suffix = source.suffix.lower()
        payload_format = str(payload.get("format") or "").lower().lstrip(".")
        if payload_format and payload_format != suffix.lstrip("."):
            return 0
        if suffix == ".pdf":
            actual = {
                item["name"]
                for item in self._pdf_structure(source)["fields"]
            }
            expected = set((payload.get("mapping") or {}).keys())
            return len(actual & expected) if expected.issubset(actual) else 0
        if suffix == ".xlsx":
            workbook = load_workbook(source, read_only=False, data_only=False)
            try:
                score = 0
                for field in payload.get("cells", []):
                    sheet = str(field.get("sheet") or "")
                    cell = str(field.get("cell") or "")
                    if sheet in workbook.sheetnames and cell:
                        workbook[sheet][cell]
                        score += 1
                return score
            finally:
                workbook.close()
        if suffix == ".docx":
            actual = {
                item["field_id"]
                for item in self._word_structure(source)["slots"]
            }
            expected = {
                str(field.get("field_id") or "")
                for field in payload.get("fields", [])
            }
            return len(actual & expected) if expected.issubset(actual) else 0
        return 0

    def _pdf_structure(self, path: Path) -> dict[str, Any]:
        fields: list[dict[str, Any]] = []
        with fitz.open(path) as document:
            for page_number, page in enumerate(document):
                for widget in page.widgets() or []:
                    if not widget.field_name:
                        continue
                    fields.append(
                        {
                            "name": widget.field_name,
                            "type": int(widget.field_type),
                            "page": page_number,
                            "rect": [round(value, 1) for value in widget.rect],
                        }
                    )
            return {
                "format": "pdf",
                "pages": document.page_count,
                "fields": sorted(fields, key=lambda item: (item["page"], item["name"])),
            }

    def _excel_structure(
        self,
        path: Path,
        payload: dict[str, Any],
    ) -> dict[str, Any]:
        mutable = {
            (str(field.get("sheet") or ""), str(field.get("cell") or ""))
            for field in payload.get("cells", [])
        }
        workbook = load_workbook(path, read_only=False, data_only=False)
        try:
            sheets: list[dict[str, Any]] = []
            for sheet in workbook.worksheets:
                stable_cells: list[list[Any]] = []
                for row in sheet.iter_rows():
                    for cell in row:
                        if (sheet.title, cell.coordinate) in mutable:
                            continue
                        if cell.value in (None, "") and not cell.style_id:
                            continue
                        value = cell.value
                        stable_value = (
                            f"={value}"
                            if isinstance(value, str) and value.startswith("=")
                            else str(value or "")
                        )
                        stable_cells.append(
                            [cell.coordinate, stable_value, int(cell.style_id)]
                        )
                sheets.append(
                    {
                        "name": sheet.title,
                        "rows": sheet.max_row,
                        "columns": sheet.max_column,
                        "merged": sorted(str(item) for item in sheet.merged_cells.ranges),
                        "cells": stable_cells,
                    }
                )
            return {"format": "xlsx", "sheets": sheets}
        finally:
            workbook.close()

    def _word_structure(self, path: Path) -> dict[str, Any]:
        with ZipFile(path) as archive:
            root = etree.fromstring(archive.read("word/document.xml"))
        slots: list[dict[str, str]] = []
        for index, control in enumerate(
            root.xpath(".//w:sdt", namespaces=WORD_NS),
            start=1,
        ):
            names = control.xpath(
                "./w:sdtPr/w:tag/@w:val | ./w:sdtPr/w:alias/@w:val",
                namespaces=WORD_NS,
            )
            label = str(names[0]) if names else f"Control {index}"
            slots.append(
                {
                    "field_id": f"sdt:{label}" if names else f"sdt-index:{index}",
                    "kind": "content_control",
                    "label": label,
                }
            )
        for table_index, table in enumerate(
            root.xpath(".//w:tbl", namespaces=WORD_NS),
            start=1,
        ):
            rows = table.xpath("./w:tr", namespaces=WORD_NS)
            for row_index, row in enumerate(rows, start=1):
                cells = row.xpath("./w:tc", namespaces=WORD_NS)
                for column_index, _cell in enumerate(cells, start=1):
                    label = ""
                    if column_index > 1:
                        label = self._word_text(cells[column_index - 2])
                    slots.append(
                        {
                            "field_id": (
                                f"table:{table_index}:{row_index}:{column_index}"
                            ),
                            "kind": "table_cell",
                            "label": label,
                        }
                    )
        return {"format": "docx", "slots": slots}

    def _word_text(self, element: etree._Element) -> str:
        return " ".join(
            value.strip()
            for value in element.xpath(".//w:t/text()", namespaces=WORD_NS)
            if value.strip()
        ).strip()
