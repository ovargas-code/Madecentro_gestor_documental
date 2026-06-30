from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from lxml import etree
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
IMAGE_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
)
DRAWING_PATH = "xl/drawings/drawing1.xml"
DRAWING_RELS_PATH = "xl/drawings/_rels/drawing1.xml.rels"
SIGNATURE_MEDIA_PATH = "xl/media/madecentro_signature.png"


class ExcelSignatureService:
    def apply_signature(
        self,
        workbook_path: str | Path,
        signature_path: str | Path,
        config: dict[str, object] | None = None,
    ) -> Path:
        workbook_path = Path(workbook_path).resolve()
        signature_path = Path(signature_path).resolve()
        if not workbook_path.is_file():
            raise FileNotFoundError(f"No existe el Excel: {workbook_path}")
        if not signature_path.is_file():
            raise FileNotFoundError(f"No existe la firma: {signature_path}")

        config = config or {}
        signature = self._prepare_signature(signature_path)
        config = self._resolve_layout(workbook_path, signature, config)
        temp_path = workbook_path.with_name(
            f".{workbook_path.stem}.signature.tmp.xlsx"
        )
        try:
            with ZipFile(workbook_path) as source:
                if (
                    DRAWING_PATH not in source.namelist()
                    or DRAWING_RELS_PATH not in source.namelist()
                ):
                    return self._apply_with_openpyxl(
                        workbook_path,
                        signature,
                        config,
                    )
                drawing = etree.fromstring(source.read(DRAWING_PATH))
                relationships = etree.fromstring(
                    source.read(DRAWING_RELS_PATH)
                )
                relation_id = self._upsert_relationship(relationships)
                self._remove_existing_signature(drawing)
                drawing.append(self._signature_anchor(relation_id, config))
                replacements = {
                    DRAWING_PATH: etree.tostring(
                        drawing,
                        encoding="UTF-8",
                        xml_declaration=True,
                        standalone=True,
                    ),
                    DRAWING_RELS_PATH: etree.tostring(
                        relationships,
                        encoding="UTF-8",
                        xml_declaration=True,
                        standalone=True,
                    ),
                    SIGNATURE_MEDIA_PATH: signature,
                }
                self._write_package(source, temp_path, replacements)
            temp_path.replace(workbook_path)
        finally:
            temp_path.unlink(missing_ok=True)
        return workbook_path

    def _prepare_signature(self, path: Path) -> bytes:
        with Image.open(path) as source:
            image = source.convert("RGBA")
            pixels = []
            for red, green, blue, _ in image.getdata():
                darkness = 255 - min(red, green, blue)
                alpha = 0 if darkness < 10 else min(255, darkness * 3)
                pixels.append((0, 0, 0, alpha))
            image.putdata(pixels)
            bounding_box = image.getchannel("A").getbbox()
            if bounding_box:
                image = image.crop(bounding_box)
            buffer = BytesIO()
            image.save(buffer, format="PNG", optimize=True)
            return buffer.getvalue()

    def _upsert_relationship(self, relationships: etree._Element) -> str:
        for relation in relationships:
            if relation.get("Target") == "../media/madecentro_signature.png":
                return str(relation.get("Id"))
        used_ids = {
            int(value[3:])
            for relation in relationships
            if (value := str(relation.get("Id") or "")).startswith("rId")
            and value[3:].isdigit()
        }
        relation_id = f"rId{max(used_ids, default=0) + 1}"
        etree.SubElement(
            relationships,
            f"{{{PACKAGE_REL_NS}}}Relationship",
            {
                "Id": relation_id,
                "Type": IMAGE_REL_TYPE,
                "Target": "../media/madecentro_signature.png",
            },
        )
        return relation_id

    def _remove_existing_signature(self, drawing: etree._Element) -> None:
        for anchor in drawing.xpath(
            "./*[xdr:pic/xdr:nvPicPr/xdr:cNvPr"
            "[@name='Firma Madecentro']]",
            namespaces={"xdr": XDR_NS},
        ):
            drawing.remove(anchor)

    def _signature_anchor(
        self,
        relation_id: str,
        config: dict[str, object] | None = None,
    ) -> etree._Element:
        config = config or {}
        cell = str(config.get("cell") or "G114")
        column_letters, row_number = coordinate_from_string(cell)
        column = column_index_from_string(column_letters) - 1
        row = row_number - 1
        col_offset = int(config.get("col_offset_emu", 0))
        row_offset = int(config.get("row_offset_emu", 0))
        anchor = etree.Element(
            f"{{{XDR_NS}}}oneCellAnchor",
        )
        self._marker(
            anchor,
            "from",
            column=column,
            row=row,
            column_offset=col_offset,
            row_offset=row_offset,
        )
        etree.SubElement(
            anchor,
            f"{{{XDR_NS}}}ext",
            cx=str(int(config.get("width_pixels", 240)) * 9525),
            cy=str(int(config.get("height_pixels", 80)) * 9525),
        )
        picture = etree.SubElement(anchor, f"{{{XDR_NS}}}pic")
        non_visual = etree.SubElement(picture, f"{{{XDR_NS}}}nvPicPr")
        etree.SubElement(
            non_visual,
            f"{{{XDR_NS}}}cNvPr",
            id="50000",
            name="Firma Madecentro",
        )
        locks = etree.SubElement(non_visual, f"{{{XDR_NS}}}cNvPicPr")
        etree.SubElement(
            locks,
            f"{{{A_NS}}}picLocks",
            noChangeAspect="1",
        )
        fill = etree.SubElement(picture, f"{{{XDR_NS}}}blipFill")
        etree.SubElement(
            fill,
            f"{{{A_NS}}}blip",
            {f"{{{REL_NS}}}embed": relation_id},
        )
        stretch = etree.SubElement(fill, f"{{{A_NS}}}stretch")
        etree.SubElement(stretch, f"{{{A_NS}}}fillRect")
        shape = etree.SubElement(picture, f"{{{XDR_NS}}}spPr")
        geometry = etree.SubElement(
            shape,
            f"{{{A_NS}}}prstGeom",
            prst="rect",
        )
        etree.SubElement(geometry, f"{{{A_NS}}}avLst")
        etree.SubElement(anchor, f"{{{XDR_NS}}}clientData")
        return anchor

    def _apply_with_openpyxl(
        self,
        workbook_path: Path,
        signature: bytes,
        config: dict[str, object],
    ) -> Path:
        workbook = load_workbook(workbook_path)
        image_buffer = BytesIO(signature)
        try:
            sheet_name = str(config.get("sheet") or "")
            sheet = (
                workbook[sheet_name]
                if sheet_name in workbook.sheetnames
                else workbook[workbook.sheetnames[0]]
            )
            image = ExcelImage(image_buffer)
            image.width = int(config.get("width_pixels", 240))
            image.height = int(config.get("height_pixels", 80))
            sheet.add_image(image, str(config.get("cell") or "A1"))
            workbook.save(workbook_path)
        finally:
            workbook.close()
            image_buffer.close()
        return workbook_path

    def _resolve_layout(
        self,
        workbook_path: Path,
        signature: bytes,
        config: dict[str, object],
    ) -> dict[str, object]:
        if not config.get("cell"):
            detected = self._detect_signature_area(workbook_path)
            if detected:
                config = {**config, **detected}
            else:
                config = {**config, "cell": "A1"}
        if not config.get("cell"):
            return config
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
        try:
            sheet_name = str(config.get("sheet") or "")
            sheet = (
                workbook[sheet_name]
                if sheet_name in workbook.sheetnames
                else workbook[workbook.sheetnames[0]]
            )
            start = str(config["cell"])
            end = str(config.get("end_cell") or start)
            start_letters, start_row = coordinate_from_string(start)
            end_letters, end_row = coordinate_from_string(end)
            start_column = column_index_from_string(start_letters)
            end_column = column_index_from_string(end_letters)
            box_width = sum(
                self._column_width_pixels(sheet, column)
                for column in range(start_column, end_column + 1)
            )
            box_height = sum(
                self._row_height_pixels(sheet, row)
                for row in range(start_row, end_row + 1)
            )
            column_widths = [
                self._column_width_pixels(sheet, column)
                for column in range(start_column, end_column + 1)
            ]
            row_heights = [
                self._row_height_pixels(sheet, row)
                for row in range(start_row, end_row + 1)
            ]
        finally:
            workbook.close()
        padding = int(config.get("padding_pixels", 3))
        available_width = max(1, box_width - padding * 2)
        available_height = max(1, box_height - padding * 2)
        with Image.open(BytesIO(signature)) as image:
            image_width, image_height = image.size
        scale = min(
            available_width / max(image_width, 1),
            available_height / max(image_height, 1),
        )
        rendered_width = max(1, round(image_width * scale))
        rendered_height = max(1, round(image_height * scale))
        horizontal_align = str(config.get("horizontal_align") or "center").lower()
        horizontal_offset = int(config.get("horizontal_offset_pixels", 0))
        if horizontal_align == "left":
            left = padding + max(0, horizontal_offset)
        elif horizontal_align == "right":
            left = padding + max(0, available_width - rendered_width)
        else:
            left = padding + max(0, (available_width - rendered_width) // 2)
        top = padding + max(0, (available_height - rendered_height) // 2)
        marker_column = start_column
        marker_left = left
        for width in column_widths:
            if marker_left < width:
                break
            marker_left -= width
            marker_column += 1
        marker_row = start_row
        marker_top = top
        for height in row_heights:
            if marker_top < height:
                break
            marker_top -= height
            marker_row += 1
        pixels_to_emu = 9525
        resolved = dict(config)
        resolved.update(
            {
                "cell": self._coordinate(marker_column, marker_row),
                "width_pixels": rendered_width,
                "height_pixels": rendered_height,
                "col_offset_emu": marker_left * pixels_to_emu,
                "row_offset_emu": marker_top * pixels_to_emu,
            }
        )
        return resolved

    def _detect_signature_area(self, workbook_path: Path) -> dict[str, object] | None:
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
        try:
            for sheet in workbook.worksheets:
                candidates: list[tuple[int, dict[str, object]]] = []
                for row in sheet.iter_rows():
                    for cell in row:
                        normalized = self._normalize_value(cell.value)
                        if not normalized:
                            continue
                        if self._is_signature_line(normalized):
                            merged = self._merged_range_for_cell(sheet, cell.coordinate)
                            candidates.append(
                                (
                                    self._signature_line_score(sheet, cell.row, cell.column),
                                    self._signature_line_target(sheet, cell, merged),
                                )
                            )
                        elif self._has_signature_word(normalized):
                            target = self._signature_target_near_label(sheet, cell)
                            if target:
                                score = (
                                    205
                                    if normalized == "firma"
                                    else 210
                                    if "representante legal" in normalized
                                    else 170
                                )
                                candidates.append((score, target))
                if candidates:
                    return max(candidates, key=lambda item: item[0])[1]
        finally:
            workbook.close()
        return None

    def _normalize_value(self, value: object) -> str:
        import unicodedata

        normalized = unicodedata.normalize("NFKD", str(value or "").casefold())
        normalized = "".join(
            char
            for char in normalized
            if not unicodedata.combining(char)
        )
        return " ".join(normalized.split())

    def _has_signature_word(self, normalized: str) -> bool:
        import re

        return re.search(r"\bfirma\b", normalized) is not None

    def _is_signature_line(self, value: str) -> bool:
        compact = value.replace(" ", "")
        return len(compact) >= 8 and set(compact) <= {"_"}

    def _merged_range_for_cell(self, sheet: object, coordinate: str) -> object | None:
        for merged in sheet.merged_cells.ranges:
            if coordinate in merged:
                return merged
        return None

    def _signature_line_score(
        self,
        sheet: object,
        row_number: int,
        column_number: int,
    ) -> int:
        score = 10
        for row in range(max(1, row_number - 8), min(sheet.max_row, row_number + 4) + 1):
            for cell in sheet[row]:
                normalized = self._normalize_value(cell.value)
                if not normalized:
                    continue
                distance = abs(row - row_number) + min(
                    abs(cell.column - column_number),
                    8,
                )
                if normalized == "firma":
                    score = max(score, 220 - distance * 5)
                elif self._has_signature_word(normalized) and "representante legal" in normalized:
                    score = max(score, 210 - distance * 5)
                elif normalized.startswith("12. firma"):
                    score = max(score, 190 - distance * 4)
                elif "representante legal" in normalized and distance <= 6:
                    score = max(score, 120 - distance * 4)
        return score

    def _signature_line_target(
        self,
        sheet: object,
        cell: object,
        merged: object | None,
    ) -> dict[str, object]:
        min_row = merged.min_row if merged else cell.row
        max_row = merged.max_row if merged else cell.row
        min_col = merged.min_col if merged else cell.column
        max_col = merged.max_col if merged else cell.column
        has_label_below = False
        for row in range(max_row + 1, min(sheet.max_row, max_row + 3) + 1):
            for candidate in sheet[row]:
                normalized = self._normalize_value(candidate.value)
                if normalized == "firma" or "representante legal" in normalized:
                    has_label_below = True
                    break
            if has_label_below:
                break
        start_row = min_row
        end_row = max_row
        if has_label_below:
            blank_rows: list[int] = []
            for row in range(min_row - 1, max(0, min_row - 5), -1):
                if not self._excel_range_is_blank(sheet, row, min_col, max_col):
                    break
                blank_rows.append(row)
            if blank_rows:
                start_row = min(blank_rows)
                end_row = min_row - 1
        return self._signature_target_config(sheet, start_row, end_row, min_col, max_col)

    def _signature_target_near_label(
        self,
        sheet: object,
        cell: object,
    ) -> dict[str, object] | None:
        merged = self._merged_range_for_cell(sheet, cell.coordinate)
        min_row = merged.min_row if merged else cell.row
        max_row = merged.max_row if merged else cell.row
        min_col = merged.min_col if merged else max(1, cell.column - 2)
        max_col = merged.max_col if merged else min(sheet.max_column, cell.column + 4)
        raw_text = str(cell.value or "")
        normalized_text = self._normalize_value(raw_text)
        if merged and (
            merged.max_row - merged.min_row >= 3
            or raw_text.count("\n") >= 2
        ):
            signature_max_col = (
                min_col + max(2, (max_col - min_col) // 2)
                if "huella" in normalized_text
                else max_col
            )
            start_row = min(max_row, min_row + max(1, (max_row - min_row) // 4))
            end_row = min(max_row, start_row + max(2, (max_row - min_row) // 3))
            return self._signature_target_config(
                sheet,
                start_row,
                end_row,
                min_col,
                signature_max_col,
                {
                    "horizontal_align": "left",
                    "horizontal_offset_pixels": 16,
                }
            )
        if "huella" in normalized_text and self._has_signature_word(normalized_text):
            signature_min_col = min(sheet.max_column, min_col + 1)
            signature_max_col = max(
                signature_min_col,
                max_col - max(2, (max_col - min_col + 1) // 4),
            )
            below = self._blank_signature_block(
                sheet,
                start=min(sheet.max_row, max_row + 1),
                stop=min(sheet.max_row, max_row + 6),
                step=1,
                min_col=signature_min_col,
                max_col=signature_max_col,
            )
            if below:
                return self._signature_target_config(
                    sheet,
                    min(below),
                    max(below),
                    signature_min_col,
                    signature_max_col,
                )
        above_merged = self._merged_signature_area_above_label(
            sheet,
            min_row,
            min_col,
            max_col,
        )
        if above_merged:
            return self._signature_target_config(
                sheet,
                above_merged.min_row,
                above_merged.max_row,
                above_merged.min_col,
                above_merged.max_col,
            )
        if max_col - min_col < 2:
            max_col = min(sheet.max_column, min_col + 3)

        above = self._blank_signature_block(
            sheet,
            start=max(1, min_row - 1),
            stop=max(1, min_row - 5),
            step=-1,
            min_col=min_col,
            max_col=max_col,
        )
        if above:
            return self._signature_target_config(
                sheet,
                min(above),
                max(above),
                min_col,
                max_col,
            )
        below = self._blank_signature_block(
            sheet,
            start=min(sheet.max_row, max_row + 1),
            stop=min(sheet.max_row, max_row + 5),
            step=1,
            min_col=min_col,
            max_col=max_col,
        )
        if below:
            return self._signature_target_config(
                sheet,
                min(below),
                max(below),
                min_col,
                max_col,
            )
        return self._signature_target_config(
            sheet,
            max(1, min_row - 1),
            max(1, min_row - 1),
            min_col,
            max_col,
        )

    def _merged_signature_area_above_label(
        self,
        sheet: object,
        label_row: int,
        min_col: int,
        max_col: int,
    ) -> object | None:
        best: object | None = None
        best_score = -1
        for merged in sheet.merged_cells.ranges:
            if merged.max_row >= label_row or label_row - merged.max_row > 5:
                continue
            overlap = min(max_col, merged.max_col) - max(min_col, merged.min_col) + 1
            if overlap <= 0:
                continue
            row_span = merged.max_row - merged.min_row + 1
            col_span = merged.max_col - merged.min_col + 1
            if row_span < 2 or col_span < 3:
                continue
            score = overlap * 10 + row_span * 3 - (label_row - merged.max_row)
            if score > best_score:
                best = merged
                best_score = score
        return best

    def _blank_signature_block(
        self,
        sheet: object,
        start: int,
        stop: int,
        step: int,
        min_col: int,
        max_col: int,
    ) -> list[int]:
        rows: list[int] = []
        for row in range(start, stop + step, step):
            if row < 1 or row > sheet.max_row:
                continue
            if not self._excel_range_is_blank(sheet, row, min_col, max_col):
                if rows:
                    break
                continue
            rows.append(row)
            if len(rows) >= 3:
                break
        return rows

    def _excel_range_is_blank(
        self,
        sheet: object,
        row: int,
        min_col: int,
        max_col: int,
    ) -> bool:
        for column in range(min_col, max_col + 1):
            if sheet.cell(row, column).value not in (None, ""):
                return False
        for merged in sheet.merged_cells.ranges:
            if (
                merged.min_row <= row <= merged.max_row
                and merged.max_col >= min_col
                and merged.min_col <= max_col
            ):
                anchor = sheet.cell(merged.min_row, merged.min_col)
                if anchor.value not in (None, ""):
                    return False
        return True

    def _signature_target_config(
        self,
        sheet: object,
        start_row: int,
        end_row: int,
        min_col: int,
        max_col: int,
        extra: dict[str, object] | None = None,
    ) -> dict[str, object]:
        config: dict[str, object] = {
            "sheet": sheet.title,
            "cell": sheet.cell(start_row, min_col).coordinate,
            "end_cell": sheet.cell(end_row, max_col).coordinate,
            "fit": "contain",
            "padding_pixels": 3,
        }
        if extra:
            config.update(extra)
        return config

    def _coordinate(self, column: int, row: int) -> str:
        from openpyxl.utils import get_column_letter

        return f"{get_column_letter(column)}{row}"

    def _column_width_pixels(self, sheet: object, column: int) -> int:
        from openpyxl.utils import get_column_letter

        width = sheet.column_dimensions[get_column_letter(column)].width
        width = 13.0 if width is None else float(width)
        return max(1, round(width * 7 + 5))

    def _row_height_pixels(self, sheet: object, row: int) -> int:
        height = sheet.row_dimensions[row].height
        height = 15.0 if height is None else float(height)
        return max(1, round(height * 96 / 72))

    def _marker(
        self,
        anchor: etree._Element,
        name: str,
        column: int,
        row: int,
        column_offset: int = 0,
        row_offset: int = 0,
    ) -> None:
        marker = etree.SubElement(anchor, f"{{{XDR_NS}}}{name}")
        etree.SubElement(marker, f"{{{XDR_NS}}}col").text = str(column)
        etree.SubElement(marker, f"{{{XDR_NS}}}colOff").text = str(column_offset)
        etree.SubElement(marker, f"{{{XDR_NS}}}row").text = str(row)
        etree.SubElement(marker, f"{{{XDR_NS}}}rowOff").text = str(row_offset)

    def _write_package(
        self,
        source: ZipFile,
        target_path: Path,
        replacements: dict[str, bytes],
    ) -> None:
        written: set[str] = set()
        with ZipFile(
            target_path,
            "w",
            compression=ZIP_DEFLATED,
        ) as target:
            for item in source.infolist():
                data = replacements.get(item.filename, source.read(item.filename))
                target.writestr(self._copy_zip_info(item), data)
                written.add(item.filename)
            for filename, data in replacements.items():
                if filename not in written:
                    target.writestr(filename, data)

    def _copy_zip_info(self, item: ZipInfo) -> ZipInfo:
        copied = ZipInfo(item.filename, item.date_time)
        copied.compress_type = item.compress_type
        copied.comment = item.comment
        copied.extra = item.extra
        copied.internal_attr = item.internal_attr
        copied.external_attr = item.external_attr
        copied.create_system = item.create_system
        return copied
