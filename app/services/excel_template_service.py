from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
VML_NS = "urn:schemas-microsoft-com:vml"
EXCEL_NS = "urn:schemas-microsoft-com:office:excel"
XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"

MASTER_KEY_BY_CELL = {
    "FORMULARIO!N9": "ciudad",
    "FORMULARIO!U9": "departamento",
    "FORMULARIO!B16": "razon_social",
    "FORMULARIO!Q16": "nit",
    "FORMULARIO!C18": "direccion",
    "FORMULARIO!P18": "ciudad",
    "FORMULARIO!S18": "telefono",
    "FORMULARIO!W18": "celular",
    "FORMULARIO!H21": "correo",
    "FORMULARIO!C28": "actividad_principal",
    "FORMULARIO!B35": "representante_legal",
    "FORMULARIO!V35": "representante_id",
    "FORMULARIO!S40": "representante_telefono",
    "FORMULARIO!B70": "banco_1",
    "FORMULARIO!V70": "ciudad",
    "FORMULARIO!K71": "cuenta_1",
    "FORMULARIO!B72": "banco_2",
    "FORMULARIO!K73": "cuenta_2",
    "FORMULARIO!B59": "junta_1_id",
    "FORMULARIO!C59": "junta_1_tipo_id",
    "FORMULARIO!E59": "junta_1_nombre",
    "FORMULARIO!B60": "junta_2_id",
    "FORMULARIO!C60": "junta_2_tipo_id",
    "FORMULARIO!E60": "junta_2_nombre",
    "FORMULARIO!B61": "junta_3_id",
    "FORMULARIO!C61": "junta_3_tipo_id",
    "FORMULARIO!E61": "junta_3_nombre",
    "FORMULARIO!B62": "junta_4_id",
    "FORMULARIO!C62": "junta_4_tipo_id",
    "FORMULARIO!E62": "junta_4_nombre",
    "FORMULARIO!B63": "junta_5_id",
    "FORMULARIO!N59": "junta_6_id",
    "FORMULARIO!Q59": "junta_6_tipo_id",
    "FORMULARIO!S59": "junta_6_nombre",
    "FORMULARIO!N60": "junta_7_id",
    "FORMULARIO!Q60": "junta_7_tipo_id",
    "FORMULARIO!S60": "junta_7_nombre",
    "FORMULARIO!N61": "junta_8_id",
    "FORMULARIO!Q61": "junta_8_tipo_id",
    "FORMULARIO!S61": "junta_8_nombre",
    "FORMULARIO!N62": "junta_9_id",
    "FORMULARIO!Q62": "junta_9_tipo_id",
    "FORMULARIO!S62": "junta_9_nombre",
}

VALUE_FORMAT_BY_CELL = {
    "FORMULARIO!B70": "ENTIDAD:    {value}",
    "FORMULARIO!V70": "CIUDAD:    {value}",
    "FORMULARIO!K71": "NÚMERO DE CUENTA:    {value}",
    "FORMULARIO!B72": "ENTIDAD   {value}",
    "FORMULARIO!K73": "NÚMERO DE CUENTA     {value}",
}


class ExcelTemplateService:
    def compare_workbooks(
        self,
        empty_path: str | Path,
        completed_path: str | Path,
    ) -> dict[str, Any]:
        empty_path = Path(empty_path)
        completed_path = Path(completed_path)
        self._validate_paths(empty_path, completed_path)

        empty_book = load_workbook(empty_path, data_only=False, read_only=False)
        completed_book = load_workbook(
            completed_path,
            data_only=False,
            read_only=False,
        )
        try:
            if empty_book.sheetnames != completed_book.sheetnames:
                raise ValueError("Los libros no tienen las mismas hojas.")

            cells: list[dict[str, Any]] = []
            for sheet_name in empty_book.sheetnames:
                empty_sheet = empty_book[sheet_name]
                completed_sheet = completed_book[sheet_name]
                max_row = max(empty_sheet.max_row, completed_sheet.max_row)
                max_column = max(
                    empty_sheet.max_column,
                    completed_sheet.max_column,
                )
                for row in range(1, max_row + 1):
                    for column in range(1, max_column + 1):
                        empty_cell = empty_sheet.cell(row, column)
                        completed_cell = completed_sheet.cell(row, column)
                        if empty_cell.value == completed_cell.value:
                            continue
                        field_id = f"{sheet_name}!{empty_cell.coordinate}"
                        label = self._infer_label(
                            empty_sheet,
                            row,
                            column,
                        )
                        cells.append(
                            {
                                "field_id": field_id,
                                "sheet": sheet_name,
                                "cell": empty_cell.coordinate,
                                "label": label,
                                "empty_value": empty_cell.value,
                                "sample_value": completed_cell.value,
                                "value_type": self._value_type(
                                    completed_cell.value
                                ),
                                "master_key": MASTER_KEY_BY_CELL.get(
                                    field_id,
                                    "",
                                ),
                                "value_format": VALUE_FORMAT_BY_CELL.get(
                                    field_id,
                                ),
                                "auto_value": self._automatic_date_value(
                                    empty_cell.coordinate,
                                    label,
                                ),
                                "preserve_reference": self._is_selection_mark(
                                    empty_cell.value,
                                    completed_cell.value,
                                ),
                            }
                        )
        finally:
            empty_book.close()
            completed_book.close()

        controls = self._compare_controls(empty_path, completed_path)
        self._add_control_context(empty_path, controls)
        payload = {
            "name": "mapeo_formulario_excel",
            "template_file": empty_path.name,
            "reference_file": completed_path.name,
            "cells": cells,
            "controls": controls,
        }
        drawing_marks = self._drawing_text_marks(completed_path)
        if drawing_marks:
            payload["drawing_marks"] = drawing_marks
        return payload

    def _validate_paths(self, empty_path: Path, completed_path: Path) -> None:
        for path in (empty_path, completed_path):
            if not path.is_file():
                raise FileNotFoundError(f"No existe el archivo Excel: {path}")
            if path.suffix.lower() != ".xlsx":
                raise ValueError(f"El archivo no es XLSX: {path}")

    def _value_type(self, value: object) -> str:
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)):
            return "number"
        return "string"

    def _automatic_date_value(
        self,
        coordinate: str,
        label: str,
    ) -> str | None:
        match = re.fullmatch(r"[A-Z]+([1-9][0-9]*)", coordinate)
        if not match or int(match.group(1)) > 15:
            return None
        normalized = unicodedata.normalize("NFKD", label)
        normalized = "".join(
            char
            for char in normalized.casefold()
            if not unicodedata.combining(char)
        )
        normalized = " ".join(normalized.split())
        if "fecha" in normalized and any(
            token in normalized
            for token in (
                "solicitud",
                "solicitur",
                "creacion",
                "diligenciamiento",
            )
        ):
            return "current_date"
        return {
            "dia": "current_day",
            "mes": "current_month_name",
            "ano": "current_year",
        }.get(normalized)

    def _is_selection_mark(
        self,
        empty_value: object,
        completed_value: object,
    ) -> bool:
        return (
            empty_value in (None, "")
            and str(completed_value or "").strip().casefold() in {"x", "✓"}
        )

    def _infer_label(self, sheet: Any, row: int, column: int) -> str:
        for candidate_column in range(column - 1, 0, -1):
            value = sheet.cell(row, candidate_column).value
            if value not in (None, ""):
                return str(value).strip()
        for candidate_row in range(row - 1, max(row - 4, 0), -1):
            value = sheet.cell(candidate_row, column).value
            if value not in (None, ""):
                return str(value).strip()
        return ""

    def _compare_controls(
        self,
        empty_path: Path,
        completed_path: Path,
    ) -> list[dict[str, Any]]:
        empty_controls = self._read_controls(empty_path)
        completed_controls = self._read_controls(completed_path)
        controls: list[dict[str, Any]] = []
        for control_id, empty_control in empty_controls.items():
            completed_control = completed_controls.get(control_id)
            if not completed_control:
                continue
            controls.append(
                {
                    **empty_control,
                    "field_id": f"control:{control_id}",
                    "kind": "checkbox",
                    "label": " / ".join(empty_control.get("nearby_text", []))
                    or f"Checkbox {control_id}",
                    "sample_checked": completed_control["checked"],
                    "master_key": "",
                }
            )
        return controls

    def _add_control_context(
        self,
        workbook_path: Path,
        controls: list[dict[str, Any]],
    ) -> None:
        workbook = load_workbook(
            workbook_path,
            data_only=False,
            read_only=False,
        )
        try:
            for control in controls:
                sheet_name = str(control.get("sheet") or "")
                if sheet_name not in workbook.sheetnames:
                    sheet_name = workbook.sheetnames[0]
                    control["sheet"] = sheet_name
                sheet = workbook[sheet_name]
                anchor_values = [
                    int(value.strip())
                    for value in str(control["anchor"]).split(",")
                ]
                column = anchor_values[0] + 1
                row = anchor_values[2] + 1
                control["anchor_cell"] = f"{get_column_letter(column)}{row}"
                candidates: list[tuple[int, str]] = []
                for candidate_row in range(max(1, row - 2), row + 3):
                    for candidate_column in range(
                        max(1, column - 6),
                        column + 7,
                    ):
                        value = sheet.cell(
                            candidate_row,
                            candidate_column,
                        ).value
                        if value in (None, ""):
                            continue
                        distance = abs(candidate_row - row) + abs(
                            candidate_column - column
                        )
                        candidates.append((distance, str(value).strip()))
                control["nearby_text"] = [
                    value
                    for _, value in sorted(candidates, key=lambda item: item[0])[
                        :5
                    ]
                ]
                control["label"] = " / ".join(control["nearby_text"]) or str(
                    control.get("field_id") or ""
                )
        finally:
            workbook.close()

    def _read_controls(self, path: Path) -> dict[str, dict[str, Any]]:
        with ZipFile(path) as archive:
            required_parts = {
                "xl/worksheets/_rels/sheet1.xml.rels",
                "xl/drawings/vmlDrawing1.vml",
            }
            if not required_parts.issubset(set(archive.namelist())):
                return {}
            sheet_xml = ET.fromstring(
                archive.read("xl/worksheets/sheet1.xml")
            )
            rels_xml = ET.fromstring(
                archive.read("xl/worksheets/_rels/sheet1.xml.rels")
            )
            relationships = {
                relation.attrib["Id"]: Path(relation.attrib["Target"]).name
                for relation in rels_xml.findall(f"{{{REL_NS}}}Relationship")
            }
            controls_by_shape = {
                int(control.attrib["shapeId"]): relationships[
                    control.attrib[f"{{{OFFICE_REL_NS}}}id"]
                ]
                for control in sheet_xml.findall(
                    f".//{{{SHEET_NS}}}control"
                )
            }
            vml_xml = ET.fromstring(
                archive.read("xl/drawings/vmlDrawing1.vml")
            )

        controls: dict[str, dict[str, Any]] = {}
        for shape in vml_xml.findall(f".//{{{VML_NS}}}shape"):
            client_data = shape.find(f"{{{EXCEL_NS}}}ClientData")
            if (
                client_data is None
                or client_data.attrib.get("ObjectType") != "Checkbox"
            ):
                continue
            shape_id = int(shape.attrib["id"].rsplit("s", 1)[-1])
            prop_file = controls_by_shape.get(shape_id)
            if not prop_file:
                continue
            anchor_element = client_data.find(f"{{{EXCEL_NS}}}Anchor")
            anchor = (
                anchor_element.text.strip()
                if anchor_element is not None and anchor_element.text
                else ""
            )
            controls[prop_file] = {
                "control_id": prop_file,
                "sheet": "FORMULARIO",
                "shape_id": shape_id,
                "anchor": anchor,
                "checked": client_data.find(
                    f"{{{EXCEL_NS}}}Checked"
                )
                is not None,
            }
        return controls

    def _drawing_text_marks(self, path: Path) -> list[dict[str, str]]:
        marks: list[dict[str, str]] = []
        with ZipFile(path) as archive:
            for name in archive.namelist():
                if not (
                    name.startswith("xl/drawings/drawing")
                    and name.endswith(".xml")
                ):
                    continue
                drawing = ET.fromstring(archive.read(name))
                for anchor in list(drawing):
                    texts = [
                        str(element.text or "").strip()
                        for element in anchor.iter(f"{{{A_NS}}}t")
                    ]
                    if any(text.casefold() == "x" for text in texts):
                        marks.append(
                            {
                                "path": name,
                                "text": "X",
                                "xml": ET.tostring(
                                    anchor,
                                    encoding="unicode",
                                ),
                            }
                        )
        return marks
