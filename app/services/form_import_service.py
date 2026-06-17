from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import fitz
from lxml import etree
from openpyxl import load_workbook

from app.core.settings import MAPPINGS_DIR
from app.models.import_models import ImportedValue, ImportChange
from app.services.document_identity_service import DocumentIdentityService
from app.services.word_template_service import WordTemplateService


WORD_NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
}

WORD_LABEL_ALIASES = {
    "actividad economica": "actividad_principal",
    "actividad principal": "actividad_principal",
    "banco 1": "banco_1",
    "banco 2": "banco_2",
    "celular": "celular",
    "ciudad": "ciudad",
    "correo": "correo",
    "correo electronico": "correo",
    "cuenta 1": "cuenta_1",
    "cuenta 2": "cuenta_2",
    "departamento": "departamento",
    "direccion": "direccion",
    "nit": "nit",
    "pais": "pais",
    "razon social": "razon_social",
    "razon o denominacion social": "razon_social",
    "representante legal": "representante_legal",
    "nombre representante legal": "representante_legal",
    "identificacion representante legal": "representante_id",
    "documento representante legal": "representante_id",
    "cargo representante legal": "representante_cargo",
    "telefono representante legal": "representante_telefono",
    "celular representante legal": "representante_celular",
    "correo representante legal": "representante_email",
    "telefono": "telefono",
    "titular cuenta 1": "titular_cuenta_1",
    "titular cuenta 2": "titular_cuenta_2",
}


class FormImportService:
    def __init__(self, mappings_dir: Path = MAPPINGS_DIR) -> None:
        self.mappings_dir = Path(mappings_dir)
        self.document_identity = DocumentIdentityService()
        self.word_templates = WordTemplateService()
        self.last_resolved_payload: dict[str, Any] | None = None

    def extract(self, path: str | Path) -> list[ImportedValue]:
        source = Path(path)
        if not source.is_file():
            raise FileNotFoundError(f"No existe el formulario: {source}")
        suffix = source.suffix.lower()
        payload = self._resolve_payload(source)
        self.last_resolved_payload = payload
        if suffix == ".pdf":
            values = self._extract_pdf(source, payload)
        elif suffix == ".xlsx":
            values = self._extract_excel(source, payload)
        elif suffix == ".docx":
            values = self._extract_word(source, payload)
        else:
            raise ValueError("Formato no soportado. Use PDF, XLSX o DOCX.")
        if not values:
            raise ValueError(
                "No se encontraron valores asociados a datos maestros en el formulario."
            )
        return self._deduplicate(values)

    def compare(
        self,
        extracted: list[ImportedValue],
        current_rows: list[dict[str, Any]],
    ) -> list[ImportChange]:
        current = {str(row["clave"]): row for row in current_rows}
        changes: list[ImportChange] = []
        for item in extracted:
            row = current.get(item.master_key)
            current_value = str(row["valor"]) if row else ""
            if self._same_value(current_value, item.value):
                continue
            changes.append(
                ImportChange(
                    master_key=item.master_key,
                    current_value=current_value,
                    new_value=item.value,
                    category=(
                        str(row["categoria"])
                        if row
                        else self._category_for_key(item.master_key)
                    ),
                    source_field=item.source_field,
                )
            )
        return changes

    def _extract_pdf(
        self,
        path: Path,
        payload: dict[str, Any] | None = None,
    ) -> list[ImportedValue]:
        field_values: dict[str, str] = {}
        with fitz.open(path) as document:
            for page in document:
                for widget in page.widgets() or []:
                    field_name = widget.field_name or ""
                    value = widget.field_value
                    if not field_name or value in (None, "", "Off"):
                        continue
                    field_values[field_name] = str(value).strip()

        mapping = (
            {
                str(field): str(key)
                for field, key in (payload.get("mapping") or {}).items()
            }
            if payload
            else self._best_pdf_mapping(set(field_values))
        )
        values: list[ImportedValue] = []
        for field_name, master_key in mapping.items():
            value = field_values.get(field_name, "").strip()
            if master_key and value:
                values.append(
                    ImportedValue(master_key, value, f"PDF:{field_name}")
                )
        return values

    def _best_pdf_mapping(self, field_names: set[str]) -> dict[str, str]:
        candidates: list[tuple[int, dict[str, str]]] = []
        for path in self.mappings_dir.glob("*.json"):
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
                raw_mapping = payload.get("mapping") if isinstance(payload, dict) else None
                if not isinstance(raw_mapping, dict):
                    continue
                mapping = {
                    str(field): str(key)
                    for field, key in raw_mapping.items()
                    if isinstance(field, str) and isinstance(key, str)
                }
                score = sum(
                    1
                    for field, key in mapping.items()
                    if field in field_names and key
                )
                candidates.append((score, mapping))
            except (OSError, ValueError, TypeError):
                continue
        if not candidates or max(score for score, _ in candidates) == 0:
            raise ValueError("No hay un mapeo PDF compatible con el formulario.")
        return max(candidates, key=lambda item: item[0])[1]

    def _extract_excel(
        self,
        path: Path,
        payload: dict[str, Any] | None = None,
    ) -> list[ImportedValue]:
        if payload is None:
            mapping_path = self.mappings_dir / "mapeo_formulario_excel.json"
            if not mapping_path.is_file():
                raise FileNotFoundError(f"No existe el mapeo Excel: {mapping_path}")
            payload = json.loads(mapping_path.read_text(encoding="utf-8"))
        workbook = load_workbook(path, data_only=False, read_only=False)
        values: list[ImportedValue] = []
        try:
            for field in payload.get("cells", []):
                master_key = str(field.get("master_key") or "")
                sheet_name = str(field.get("sheet") or "")
                coordinate = str(field.get("cell") or "")
                if not master_key or sheet_name not in workbook.sheetnames:
                    continue
                value = workbook[sheet_name][coordinate].value
                if value in (None, ""):
                    continue
                normalized = self._remove_value_format(
                    str(value),
                    field.get("value_format"),
                )
                if normalized:
                    values.append(
                        ImportedValue(
                            master_key,
                            normalized,
                            f"Excel:{sheet_name}!{coordinate}",
                        )
                    )
        finally:
            workbook.close()
        return self._remove_incomplete_board_members(values)

    def _extract_word(
        self,
        path: Path,
        payload: dict[str, Any] | None = None,
    ) -> list[ImportedValue]:
        with ZipFile(path) as archive:
            if "word/document.xml" not in archive.namelist():
                raise ValueError("El DOCX no contiene un documento Word valido.")
            root = etree.fromstring(archive.read("word/document.xml"))

        if payload and isinstance(payload.get("fields"), list):
            slots = self.word_templates._slots(root)
            values: list[ImportedValue] = []
            for field in payload["fields"]:
                master_key = str(field.get("master_key") or "")
                field_id = str(field.get("field_id") or "")
                slot = slots.get(field_id)
                value = self._word_text(slot["element"]) if slot else ""
                if master_key and value:
                    values.append(
                        ImportedValue(master_key, value, f"Word:{field_id}")
                    )
            return values

        values = []
        values.extend(self._word_content_controls(root))
        values.extend(self._word_bookmarks(root))
        values.extend(self._word_tables(root))
        return values

    def _resolve_payload(self, path: Path) -> dict[str, Any] | None:
        candidates: list[tuple[int, bool, Path, dict[str, Any]]] = []
        for mapping_path in self.mappings_dir.glob("*.json"):
            try:
                payload = json.loads(mapping_path.read_text(encoding="utf-8"))
                if not isinstance(payload, dict):
                    continue
                score = self.document_identity.compatibility_score(path, payload)
                if score <= 0:
                    continue
                exact = self.document_identity.matches(path, payload)
                candidates.append((score, exact, mapping_path, payload))
            except (OSError, ValueError, TypeError, KeyError):
                continue
        if not candidates:
            return None
        exact_candidates = [item for item in candidates if item[1]]
        pool = exact_candidates or candidates
        best_score = max(item[0] for item in pool)
        best = [item for item in pool if item[0] == best_score]
        if len(best) > 1:
            names = ", ".join(sorted(item[2].name for item in best))
            raise ValueError(
                "El formulario coincide con varios mapeos. "
                f"Revise las plantillas: {names}"
            )
        return best[0][3]

    def _word_content_controls(
        self,
        root: etree._Element,
    ) -> list[ImportedValue]:
        values: list[ImportedValue] = []
        for control in root.xpath(".//w:sdt", namespaces=WORD_NS):
            tags = control.xpath(
                "./w:sdtPr/w:tag/@w:val | ./w:sdtPr/w:alias/@w:val",
                namespaces=WORD_NS,
            )
            master_key = self._master_key_from_labels(tags)
            value = self._word_text(control)
            if master_key and value:
                values.append(
                    ImportedValue(master_key, value, f"Word:control:{tags[0]}")
                )
        return values

    def _word_bookmarks(self, root: etree._Element) -> list[ImportedValue]:
        values: list[ImportedValue] = []
        for bookmark in root.xpath(".//w:bookmarkStart", namespaces=WORD_NS):
            name = bookmark.get(f"{{{WORD_NS['w']}}}name", "")
            master_key = self._master_key_from_labels([name])
            if not master_key:
                continue
            texts: list[str] = []
            current = bookmark.getnext()
            while current is not None:
                if current.tag == f"{{{WORD_NS['w']}}}bookmarkEnd":
                    break
                texts.extend(current.xpath(".//w:t/text()", namespaces=WORD_NS))
                current = current.getnext()
            value = " ".join(texts).strip()
            if value:
                values.append(
                    ImportedValue(master_key, value, f"Word:bookmark:{name}")
                )
        return values

    def _word_tables(self, root: etree._Element) -> list[ImportedValue]:
        values: list[ImportedValue] = []
        for row in root.xpath(".//w:tr", namespaces=WORD_NS):
            cells = row.xpath("./w:tc", namespaces=WORD_NS)
            if len(cells) < 2:
                continue
            label = self._word_text(cells[0])
            master_key = self._master_key_from_labels([label])
            value = self._word_text(cells[1])
            if master_key and value:
                values.append(
                    ImportedValue(master_key, value, f"Word:tabla:{label}")
                )
        return values

    def _word_text(self, element: etree._Element) -> str:
        return " ".join(
            part.strip()
            for part in element.xpath(".//w:t/text()", namespaces=WORD_NS)
            if part.strip()
        ).strip()

    def _master_key_from_labels(self, labels: list[str]) -> str:
        master_keys = {
            self._normalize_label(key): key
            for key in set(WORD_LABEL_ALIASES.values())
        }
        for label in labels:
            normalized = self._normalize_label(label)
            if normalized in WORD_LABEL_ALIASES:
                return WORD_LABEL_ALIASES[normalized]
            if normalized in master_keys:
                return master_keys[normalized]
        return ""

    def _remove_value_format(
        self,
        value: str,
        value_format: object,
    ) -> str:
        if not isinstance(value_format, str) or "{value}" not in value_format:
            return value.strip()
        prefix, suffix = value_format.split("{value}", 1)
        result = value
        if prefix and result.startswith(prefix):
            result = result[len(prefix):]
        if suffix and result.endswith(suffix):
            result = result[: -len(suffix)]
        return result.strip()

    def _deduplicate(
        self,
        values: list[ImportedValue],
    ) -> list[ImportedValue]:
        selected: dict[str, ImportedValue] = {}
        for item in values:
            if not item.value.strip():
                continue
            previous = selected.get(item.master_key)
            if previous and not self._same_value(previous.value, item.value):
                raise ValueError(
                    f"El formulario contiene valores distintos para "
                    f"'{item.master_key}': '{previous.value}' y '{item.value}'."
                )
            selected[item.master_key] = item
        return list(selected.values())

    def _remove_incomplete_board_members(
        self,
        values: list[ImportedValue],
    ) -> list[ImportedValue]:
        complete_members = {
            match.group(1)
            for item in values
            if (match := re.fullmatch(r"junta_(\d+)_nombre", item.master_key))
            and item.value.strip()
        }
        return [
            item
            for item in values
            if not (
                (match := re.fullmatch(r"junta_(\d+)_.+", item.master_key))
                and match.group(1) not in complete_members
            )
        ]

    def _same_value(self, left: str, right: str) -> bool:
        return " ".join(left.split()).casefold() == " ".join(right.split()).casefold()

    def _category_for_key(self, master_key: str) -> str:
        if master_key.startswith("junta_"):
            return "junta_directiva"
        return "importado"

    def _normalize_label(self, value: str) -> str:
        value = unicodedata.normalize("NFKD", value)
        value = "".join(char for char in value if not unicodedata.combining(char))
        value = value.casefold().replace("_", " ")
        value = re.sub(r"[^a-z0-9 ]+", " ", value)
        return " ".join(value.split())
