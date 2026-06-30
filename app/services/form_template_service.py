from __future__ import annotations

from collections import Counter
from pathlib import Path
import re
from typing import Any

import fitz
from openpyxl import load_workbook

from app.core.settings import signature_path
from app.models.import_models import ImportedValue
from app.services.ai_mapping_service import AiMappingService
from app.services.customer_catalog_service import CUSTOMER_KEYS
from app.services.document_identity_service import DocumentIdentityService
from app.services.excel_fill_service import ExcelFillService
from app.services.excel_signature_service import ExcelSignatureService
from app.services.excel_template_service import ExcelTemplateService
from app.services.pdf_field_service import PdfFieldService
from app.services.pdf_fill_service import PdfFillService
from app.services.pdf_signature_service import PdfSignatureService
from app.services.word_template_service import WordTemplateService
from app.services.word_signature_service import WordSignatureService


class FormTemplateService:
    def __init__(self, use_openai_mapping: bool = False) -> None:
        self.ai_mapping = (
            AiMappingService()
            if use_openai_mapping
            else AiMappingService(api_key="")
        )
        self.document_identity = DocumentIdentityService()
        self.pdf_fields = PdfFieldService()
        self.pdf_fill = PdfFillService()
        self.pdf_signature = PdfSignatureService()
        self.excel_templates = ExcelTemplateService()
        self.excel_fill = ExcelFillService()
        self.excel_signature = ExcelSignatureService()
        self.word_templates = WordTemplateService()
        self.word_signature = WordSignatureService()

    def learn(
        self,
        empty_path: str | Path,
        completed_path: str | Path,
        master_data: dict[str, str],
    ) -> dict[str, Any]:
        empty = Path(empty_path)
        completed = Path(completed_path)
        if empty.suffix.lower() != completed.suffix.lower():
            raise ValueError("La plantilla vacia y la diligenciada deben tener el mismo formato.")
        suffix = empty.suffix.lower()
        if suffix == ".pdf":
            payload = self._learn_pdf(empty, completed)
        elif suffix == ".xlsx":
            payload = self.excel_templates.compare_workbooks(empty, completed)
            payload["format"] = "xlsx"
        elif suffix == ".docx":
            payload = self.word_templates.compare_documents(empty, completed)
        else:
            raise ValueError("Formato no soportado. Use PDF, XLSX o DOCX.")
        mapping_data = {
            **{key: "" for key in CUSTOMER_KEYS},
            "dia_expedicion": "",
            "mes_expedicion": "",
            "ano_expedicion": "",
            **master_data,
        }
        self.suggest_mapping(payload, mapping_data)
        payload["schema_version"] = 2
        payload["template_fingerprint"] = self.document_identity.fingerprint(
            empty,
            payload,
        )
        payload.setdefault("signature", {"enabled": True, "mode": "auto"})
        return payload

    def learn_pdf_acroform(
        self,
        pdf_path: str | Path,
        master_data: dict[str, str],
    ) -> dict[str, Any]:
        source = Path(pdf_path)
        if source.suffix.lower() != ".pdf":
            raise ValueError("La plantilla debe ser un archivo PDF.")
        fields = self.pdf_fields.list_fields(source)
        if not fields:
            raise ValueError("El PDF no contiene campos AcroForm editables.")
        unique_fields: dict[str, Any] = {}
        for field in fields:
            unique_fields.setdefault(field.field_name, field)
        mapping = {
            field_name: ""
            for field_name in unique_fields
        }
        payload: dict[str, Any] = {
            "name": f"mapeo_{source.stem}",
            "format": "pdf",
            "schema_version": 2,
            "template_file": source.name,
            "mapping": mapping,
            "samples": {
                field_name: str(field.value or "")
                for field_name, field in unique_fields.items()
            },
            "field_metadata": [
                {
                    "field_id": field_name,
                    "field_type": field.field_type,
                    "page": field.page,
                    "options": field.options,
                }
                for field_name, field in unique_fields.items()
            ],
            "signature": {"enabled": True, "mode": "auto"},
        }
        mapping_data = {
            **{key: "" for key in CUSTOMER_KEYS},
            "dia_expedicion": "",
            "mes_expedicion": "",
            "ano_expedicion": "",
            **master_data,
        }
        self.suggest_mapping(payload, mapping_data)
        payload["template_fingerprint"] = self.document_identity.fingerprint(
            source,
            payload,
        )
        return payload

    def suggest_mapping(
        self,
        payload: dict[str, Any],
        master_data: dict[str, str],
    ) -> None:
        fields = self.fields(payload)
        master_keys = list(master_data)
        labels = [
            str(field.get("label") or field.get("field_id") or "")
            for field in fields
        ]
        suggestions = self.ai_mapping.suggest_mapping(labels, master_keys)
        label_counts = Counter(
            self.ai_mapping._canonical_field_name(label)
            for label in labels
            if label
        )
        has_separate_tax_id_check_digit = "dv" in label_counts
        values_to_keys: dict[str, list[str]] = {}
        for key, value in master_data.items():
            normalized = self._normalize_value(value)
            if normalized:
                values_to_keys.setdefault(normalized, []).append(key)
        for field, label in zip(fields, labels):
            if field.get("master_key"):
                continue
            sample = self._normalize_value(field.get("sample_value"))
            matches = (
                values_to_keys.get(sample, [])
                if self._is_distinctive_sample(sample)
                else []
            )
            canonical_label = self.ai_mapping._canonical_field_name(label)
            label_suggestion = (
                suggestions.get(label, "")
                if label_counts.get(canonical_label, 0) == 1
                else ""
            )
            if label_suggestion:
                field["master_key"] = label_suggestion
            elif len(matches) == 1:
                field["master_key"] = matches[0]
            if (
                not field.get("master_key")
                and not field.get("preserve_reference")
            ):
                field["master_key"] = self._excel_fallback_key(
                    field,
                    str(field.get("label") or ""),
                )
            self._set_value_transform(
                field,
                label,
                has_separate_tax_id_check_digit,
            )
        if isinstance(payload.get("mapping"), dict):
            payload["mapping"] = {
                str(field.get("field_id") or ""): str(
                    field.get("master_key") or ""
                )
                for field in fields
            }

    def fields(self, payload: dict[str, Any]) -> list[dict[str, Any]]:
        if isinstance(payload.get("fields"), list):
            return payload["fields"]
        if isinstance(payload.get("cells"), list):
            return [*payload["cells"], *payload.get("controls", [])]
        if isinstance(payload.get("mapping"), dict):
            samples = payload.get("samples", {})
            return [
                {
                    "field_id": field_name,
                    "label": field_name,
                    "sample_value": samples.get(field_name, ""),
                    "master_key": master_key,
                }
                for field_name, master_key in payload["mapping"].items()
            ]
        return []

    def apply_field_mapping(
        self,
        payload: dict[str, Any],
        mapping: dict[str, str],
    ) -> dict[str, Any]:
        if isinstance(payload.get("mapping"), dict):
            payload["mapping"] = dict(mapping)
            return payload
        for field in self.fields(payload):
            field_id = str(field.get("field_id") or "")
            if field_id in mapping:
                field["master_key"] = mapping[field_id]
        return payload

    def mapping(self, payload: dict[str, Any]) -> dict[str, str]:
        if isinstance(payload.get("mapping"), dict):
            return {
                str(field): str(key)
                for field, key in payload["mapping"].items()
            }
        return {
            str(field.get("field_id") or ""): str(field.get("master_key") or "")
            for field in self.fields(payload)
        }

    def fill(
        self,
        template_path: str | Path,
        output_path: str | Path,
        payload: dict[str, Any],
        master_data: dict[str, str],
    ) -> Path:
        suffix = Path(template_path).suffix.lower()
        signature = signature_path()
        signature_config = payload.get("signature", {})
        signature_enabled = bool(
            signature
            and (
                not isinstance(signature_config, dict)
                or signature_config.get("enabled", True)
            )
        )
        if suffix == ".pdf":
            result = self.pdf_fill.fill_pdf(
                template_path,
                output_path,
                self.mapping(payload),
                master_data,
            )
            if signature_enabled:
                self.pdf_signature.apply_signature(
                    result,
                    signature,
                    signature_config if isinstance(signature_config, dict) else {},
                )
            return result
        if suffix == ".xlsx":
            result = self.excel_fill.fill_workbook(
                template_path,
                output_path,
                payload,
                master_data,
            )
            if signature_enabled:
                excel_config = (
                    signature_config if isinstance(signature_config, dict) else {}
                )
                if "cell" not in excel_config:
                    detected = self._excel_signature_cell(Path(template_path))
                    if detected:
                        excel_config = {**excel_config, **detected}
                self.excel_signature.apply_signature(
                    result,
                    signature,
                    excel_config,
                )
            return result
        if suffix == ".docx":
            result = self.word_templates.fill_document(
                template_path,
                output_path,
                payload,
                master_data,
            )
            if signature_enabled:
                self.word_signature.apply_signature(
                    result,
                    signature,
                    signature_config if isinstance(signature_config, dict) else {},
                )
            return result
        raise ValueError("Formato de plantilla no soportado.")

    def sample_values(self, payload: dict[str, Any]) -> list[ImportedValue]:
        values: list[ImportedValue] = []
        if isinstance(payload.get("mapping"), dict):
            samples = payload.get("samples", {})
            for field_id, master_key in payload["mapping"].items():
                value = str(samples.get(field_id, "") or "").strip()
                if master_key and value:
                    values.append(
                        ImportedValue(str(master_key), value, f"Plantilla:{field_id}")
                    )
            return self._deduplicate_samples(values)
        for field in self.fields(payload):
            master_key = str(field.get("master_key") or "")
            value = str(field.get("sample_value") or "").strip()
            if master_key and value:
                values.append(
                    ImportedValue(
                        master_key,
                        value,
                        f"Plantilla:{field.get('field_id', '')}",
                    )
                )
        return self._deduplicate_samples(values)

    def _learn_pdf(self, empty: Path, completed: Path) -> dict[str, Any]:
        empty_fields = self.pdf_fields.list_fields(empty)
        if not empty_fields:
            raise ValueError("El PDF vacio no contiene campos AcroForm editables.")
        completed_values: dict[str, str] = {}
        with fitz.open(completed) as document:
            for page in document:
                for widget in page.widgets() or []:
                    if widget.field_name:
                        completed_values[widget.field_name] = str(
                            widget.field_value or ""
                        )
        missing = sorted(
            field.field_name
            for field in empty_fields
            if field.field_name not in completed_values
        )
        if missing:
            raise ValueError(
                "El PDF diligenciado no corresponde a la plantilla vacia."
            )
        mapping = {field.field_name: "" for field in empty_fields}
        return {
            "name": "mapeo_formulario_pdf",
            "format": "pdf",
            "mapping": mapping,
            "samples": completed_values,
        }

    def _normalize_value(self, value: object) -> str:
        return " ".join(str(value or "").split()).casefold()

    def _is_distinctive_sample(self, sample: str) -> bool:
        compact = "".join(char for char in sample if char.isalnum())
        if compact.isdigit():
            return len(compact) >= 6
        return len(compact) >= 4

    def _excel_fallback_key(
        self,
        field: dict[str, Any],
        label: str,
    ) -> str:
        cell = str(field.get("cell") or "").casefold()
        sheet = str(field.get("sheet") or "")
        sample = field.get("sample_value")
        if not cell or not sheet or sample in (None, ""):
            return ""
        canonical_label = self.ai_mapping._canonical_field_name(label)
        base = canonical_label or "campo"
        return f"{base}_{cell}"

    def _set_value_transform(
        self,
        field: dict[str, Any],
        label: str,
        has_separate_tax_id_check_digit: bool,
    ) -> None:
        if field.get("master_key") != "nit":
            return
        canonical_label = self.ai_mapping._canonical_field_name(label)
        if canonical_label == "dv":
            field["value_transform"] = "tax_id_check_digit"
        elif (
            has_separate_tax_id_check_digit
            and canonical_label in {"nit", "nit_o_no_de_identificacion"}
        ):
            field["value_transform"] = "tax_id_number"

    def _has_excel_signature_area(self, template_path: Path) -> bool:
        return self._excel_signature_cell(template_path) is not None

    def _excel_signature_cell(
        self,
        template_path: Path,
    ) -> dict[str, Any] | None:
        workbook = load_workbook(
            template_path,
            read_only=False,
            data_only=False,
        )
        try:
            for sheet in workbook.worksheets:
                candidates: list[tuple[int, dict[str, Any]]] = []
                for row in sheet.iter_rows():
                    for cell in row:
                        normalized = self._normalize_value(cell.value)
                        if not normalized:
                            continue
                        if self._is_signature_line(normalized):
                            merged = self._merged_range_for_cell(sheet, cell.coordinate)
                            target = self._signature_line_target(
                                sheet,
                                cell,
                                merged,
                            )
                            candidates.append(
                                (
                                    self._signature_line_score(
                                        sheet,
                                        cell.row,
                                        cell.column,
                                    ),
                                    target,
                                )
                            )
                        elif self._has_signature_word(normalized):
                            target = self._signature_target_near_label(
                                sheet,
                                cell,
                                normalized,
                            )
                            if target:
                                score = (
                                    205
                                    if normalized == "firma"
                                    else 210
                                    if "representante legal" in normalized
                                    else 170
                                )
                                candidates.append((score, target))
                if candidates:
                    return max(candidates, key=lambda item: item[0])[1]
        finally:
            workbook.close()
        return None

    def _signature_line_score(
        self,
        sheet: Any,
        row_number: int,
        column_number: int,
    ) -> int:
        score = 10
        for row in range(max(1, row_number - 8), min(sheet.max_row, row_number + 4) + 1):
            for cell in sheet[row]:
                normalized = self._normalize_value(cell.value)
                if not normalized:
                    continue
                distance = abs(row - row_number) + min(
                    abs(cell.column - column_number),
                    8,
                )
                if normalized == "firma":
                    score = max(score, 220 - distance * 5)
                elif "firma del representante legal" in normalized:
                    score = max(score, 210 - distance * 5)
                elif normalized.startswith("12. firma"):
                    score = max(score, 190 - distance * 4)
                elif "representante legal" in normalized and distance <= 6:
                    score = max(score, 120 - distance * 4)
        return score

    def _signature_line_target(
        self,
        sheet: Any,
        cell: Any,
        merged: Any | None,
    ) -> dict[str, Any]:
        min_row = merged.min_row if merged else cell.row
        max_row = merged.max_row if merged else cell.row
        min_col = merged.min_col if merged else cell.column
        max_col = merged.max_col if merged else cell.column
        has_label_below = False
        for row in range(max_row + 1, min(sheet.max_row, max_row + 3) + 1):
            for candidate in sheet[row]:
                normalized = self._normalize_value(candidate.value)
                if normalized == "firma" or "representante legal" in normalized:
                    has_label_below = True
                    break
            if has_label_below:
                break
        start_row = min_row
        end_row = max_row
        if has_label_below:
            blank_rows: list[int] = []
            for row in range(min_row - 1, max(0, min_row - 5), -1):
                if not self._excel_range_is_blank(
                    sheet,
                    row,
                    min_col,
                    max_col,
                ):
                    break
                blank_rows.append(row)
            if blank_rows:
                start_row = min(blank_rows)
                end_row = min_row - 1
        return {
            "sheet": sheet.title,
            "cell": sheet.cell(start_row, min_col).coordinate,
            "end_cell": sheet.cell(end_row, max_col).coordinate,
            "fit": "contain",
            "padding_pixels": 3,
        }

    def _signature_target_near_label(
        self,
        sheet: Any,
        cell: Any,
        normalized_label: str,
    ) -> dict[str, Any] | None:
        merged = self._merged_range_for_cell(sheet, cell.coordinate)
        min_row = merged.min_row if merged else cell.row
        max_row = merged.max_row if merged else cell.row
        min_col = merged.min_col if merged else max(1, cell.column - 2)
        max_col = merged.max_col if merged else min(sheet.max_column, cell.column + 4)
        raw_text = str(cell.value or "")
        if merged and (
            merged.max_row - merged.min_row >= 3
            or raw_text.count("\n") >= 2
        ):
            signature_max_col = (
                min_col + max(2, (max_col - min_col) // 2)
                if "huella" in self._normalize_value(raw_text)
                else max_col
            )
            start_row = min(max_row, min_row + max(1, (max_row - min_row) // 4))
            end_row = min(max_row, start_row + max(2, (max_row - min_row) // 3))
            return self._signature_target_config(
                sheet,
                start_row,
                end_row,
                min_col,
                signature_max_col,
                {
                    "horizontal_align": "left",
                    "horizontal_offset_pixels": 16,
                },
            )
        if "huella" in normalized_label and self._has_signature_word(normalized_label):
            signature_min_col = min(sheet.max_column, min_col + 1)
            signature_max_col = max(
                signature_min_col,
                max_col - max(2, (max_col - min_col + 1) // 4),
            )
            below = self._blank_signature_block(
                sheet,
                start=min(sheet.max_row, max_row + 1),
                stop=min(sheet.max_row, max_row + 6),
                step=1,
                min_col=signature_min_col,
                max_col=signature_max_col,
            )
            if below:
                return self._signature_target_config(
                    sheet,
                    min(below),
                    max(below),
                    signature_min_col,
                    signature_max_col,
                )
        above_merged = self._merged_signature_area_above_label(
            sheet,
            min_row,
            min_col,
            max_col,
        )
        if above_merged:
            return self._signature_target_config(
                sheet,
                above_merged.min_row,
                above_merged.max_row,
                above_merged.min_col,
                above_merged.max_col,
            )
        if max_col - min_col < 2:
            max_col = min(sheet.max_column, min_col + 3)

        above = self._blank_signature_block(
            sheet,
            start=max(1, min_row - 1),
            stop=max(1, min_row - 5),
            step=-1,
            min_col=min_col,
            max_col=max_col,
        )
        if above:
            start_row, end_row = min(above), max(above)
            return self._signature_target_config(
                sheet,
                start_row,
                end_row,
                min_col,
                max_col,
            )

        below = self._blank_signature_block(
            sheet,
            start=min(sheet.max_row, max_row + 1),
            stop=min(sheet.max_row, max_row + 5),
            step=1,
            min_col=min_col,
            max_col=max_col,
        )
        if below:
            start_row, end_row = min(below), max(below)
            return self._signature_target_config(
                sheet,
                start_row,
                end_row,
                min_col,
                max_col,
            )

        if "representante legal" in normalized_label:
            target = self._signature_target_above(sheet, min_row)
            if target:
                return target
        return self._signature_target_config(
            sheet,
            max(1, min_row - 1),
            max(1, min_row - 1),
            min_col,
            max_col,
        )

    def _merged_signature_area_above_label(
        self,
        sheet: Any,
        label_row: int,
        min_col: int,
        max_col: int,
    ) -> Any | None:
        best: Any | None = None
        best_score = -1
        for merged in sheet.merged_cells.ranges:
            if merged.max_row >= label_row or label_row - merged.max_row > 5:
                continue
            overlap = min(max_col, merged.max_col) - max(min_col, merged.min_col) + 1
            if overlap <= 0:
                continue
            row_span = merged.max_row - merged.min_row + 1
            col_span = merged.max_col - merged.min_col + 1
            if row_span < 2 or col_span < 3:
                continue
            score = overlap * 10 + row_span * 3 - (label_row - merged.max_row)
            if score > best_score:
                best = merged
                best_score = score
        return best

    def _blank_signature_block(
        self,
        sheet: Any,
        start: int,
        stop: int,
        step: int,
        min_col: int,
        max_col: int,
    ) -> list[int]:
        rows: list[int] = []
        end = stop + step
        for row in range(start, end, step):
            if row < 1 or row > sheet.max_row:
                continue
            if not self._excel_range_is_blank(sheet, row, min_col, max_col):
                if rows:
                    break
                continue
            rows.append(row)
            if len(rows) >= 3:
                break
        return rows

    def _signature_target_config(
        self,
        sheet: Any,
        start_row: int,
        end_row: int,
        min_col: int,
        max_col: int,
        extra: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        config: dict[str, Any] = {
            "sheet": sheet.title,
            "cell": sheet.cell(start_row, min_col).coordinate,
            "end_cell": sheet.cell(end_row, max_col).coordinate,
            "fit": "contain",
            "padding_pixels": 3,
        }
        if extra:
            config.update(extra)
        return config

    def _excel_range_is_blank(
        self,
        sheet: Any,
        row: int,
        min_col: int,
        max_col: int,
    ) -> bool:
        for column in range(min_col, max_col + 1):
            if sheet.cell(row, column).value not in (None, ""):
                return False
        for merged in sheet.merged_cells.ranges:
            if (
                merged.min_row <= row <= merged.max_row
                and merged.max_col >= min_col
                and merged.min_col <= max_col
            ):
                anchor = sheet.cell(merged.min_row, merged.min_col)
                if anchor.value not in (None, ""):
                    return False
        return True

    def _is_signature_line(self, value: str) -> bool:
        compact = value.replace(" ", "")
        return len(compact) >= 8 and set(compact) <= {"_"}

    def _has_signature_word(self, normalized: str) -> bool:
        return re.search(r"\bfirma\b", normalized) is not None

    def _merged_range_for_cell(self, sheet: Any, coordinate: str) -> Any | None:
        for merged in sheet.merged_cells.ranges:
            if coordinate in merged:
                return merged
        return None

    def _signature_target_below(
        self,
        sheet: Any,
        row_number: int,
    ) -> dict[str, Any] | None:
        for row in range(row_number + 1, min(sheet.max_row, row_number + 4) + 1):
            for cell in sheet[row]:
                normalized = self._normalize_value(cell.value)
                if self._is_signature_line(normalized):
                    merged = self._merged_range_for_cell(sheet, cell.coordinate)
                    return {
                        "sheet": sheet.title,
                        "cell": (
                            sheet.cell(merged.min_row, merged.min_col).coordinate
                            if merged
                            else cell.coordinate
                        ),
                        "end_cell": (
                            sheet.cell(merged.max_row, merged.max_col).coordinate
                            if merged
                            else cell.coordinate
                        ),
                        "fit": "contain",
                        "padding_pixels": 3,
                    }
        return None

    def _signature_target_above(
        self,
        sheet: Any,
        row_number: int,
    ) -> dict[str, Any] | None:
        for row in range(row_number - 1, max(1, row_number - 5) - 1, -1):
            for cell in sheet[row]:
                normalized = self._normalize_value(cell.value)
                if self._is_signature_line(normalized):
                    merged = self._merged_range_for_cell(sheet, cell.coordinate)
                    return {
                        "sheet": sheet.title,
                        "cell": (
                            sheet.cell(merged.min_row, merged.min_col).coordinate
                            if merged
                            else cell.coordinate
                        ),
                        "end_cell": (
                            sheet.cell(merged.max_row, merged.max_col).coordinate
                            if merged
                            else cell.coordinate
                        ),
                        "fit": "contain",
                        "padding_pixels": 3,
                    }
        return None

    def _deduplicate_samples(
        self,
        values: list[ImportedValue],
    ) -> list[ImportedValue]:
        selected: dict[str, ImportedValue] = {}
        conflicting: set[str] = set()
        for item in values:
            previous = selected.get(item.master_key)
            if previous and self._normalize_value(previous.value) != self._normalize_value(
                item.value
            ):
                conflicting.add(item.master_key)
                continue
            selected[item.master_key] = item
        return [
            item
            for key, item in selected.items()
            if key not in conflicting
        ]
