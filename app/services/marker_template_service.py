from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.plantilla_generator.replacements import canonical_category


MARKER_RE = re.compile(r"<<\s*([A-Z0-9_]+)\s*>>")


class MarkerTemplateService:
    """Build mapping payloads from generated marker-based templates."""

    MASTER_KEY_CANDIDATES = {
        "razon_social": ("razon_social",),
        "nit": ("nit",),
        "representante_legal": ("representante_legal",),
        "cedula_representante": ("representante_id", "cedula_representante"),
        "direccion": ("direccion",),
        "telefono": ("telefono",),
        "correo": ("correo",),
        "pais": ("pais",),
        "banco": ("banco_1", "banco"),
        "cuenta_bancaria": ("cuenta_1", "cuenta_bancaria"),
    }

    def build_excel_payload(
        self,
        source_path: str | Path,
        template_output_path: str | Path,
        available_master_keys: list[str],
    ) -> dict[str, Any]:
        source = Path(source_path)
        target = Path(template_output_path)
        if source.suffix.lower() != ".xlsx":
            raise ValueError("Por ahora solo se pueden registrar plantillas generadas XLSX.")

        workbook = load_workbook(source, data_only=False, read_only=False)
        fields: list[dict[str, Any]] = []
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if not isinstance(cell.value, str):
                            continue
                        markers = MARKER_RE.findall(cell.value)
                        if not markers:
                            continue
                        marker = markers[0]
                        master_key = self._master_key_for_marker(
                            marker,
                            available_master_keys,
                        )
                        if not master_key:
                            continue
                        original_text = cell.value
                        clean_text = MARKER_RE.sub("", original_text).strip()
                        value_format = self._value_format(original_text, marker)
                        field: dict[str, Any] = {
                            "field_id": f"{sheet.title}!{cell.coordinate}",
                            "sheet": sheet.title,
                            "cell": cell.coordinate,
                            "label": self._label_for_cell(sheet, cell.row, cell.column),
                            "empty_value": clean_text,
                            "sample_value": f"<<{marker}>>",
                            "value_type": "string",
                            "master_key": master_key,
                        }
                        if value_format:
                            field["value_format"] = value_format
                        fields.append(field)
                        cell.value = clean_text

            if not fields:
                raise ValueError("No se encontraron marcadores reconocidos en el Excel.")
            target.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(target)
        finally:
            workbook.close()

        return {
            "name": f"mapeo_{source.stem}",
            "format": "xlsx",
            "schema_version": 2,
            "template_file": target.name,
            "reference_file": source.name,
            "cells": fields,
            "controls": [],
            "signature": {"enabled": True, "mode": "auto"},
        }

    def _master_key_for_marker(
        self,
        marker: str,
        available_master_keys: list[str],
    ) -> str:
        available = {str(key).strip() for key in available_master_keys if str(key).strip()}
        canonical = canonical_category(marker)
        candidates = self.MASTER_KEY_CANDIDATES.get(canonical, (canonical,))
        for candidate in candidates:
            if candidate in available:
                return candidate
        return ""

    def _value_format(self, original_text: str, marker: str) -> str:
        exact_marker = f"<<{marker}>>"
        if original_text.strip() == exact_marker:
            return ""
        return original_text.replace(exact_marker, "{value}").strip()

    def _label_for_cell(self, sheet: Any, row: int, column: int) -> str:
        for candidate_column in range(column - 1, 0, -1):
            value = sheet.cell(row, candidate_column).value
            if value not in (None, ""):
                return str(value).strip()
        for candidate_row in range(row - 1, max(row - 4, 0), -1):
            value = sheet.cell(candidate_row, column).value
            if value not in (None, ""):
                return str(value).strip()
        return ""
