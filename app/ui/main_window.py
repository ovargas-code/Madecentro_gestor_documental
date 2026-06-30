from __future__ import annotations

import csv
import os
import re
import shutil
import sqlite3
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from PySide6.QtCore import Qt
from PySide6.QtGui import QIcon, QPixmap
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app.core.settings import (
    BASE_DIR,
    CUSTOMER_CATALOG_PATH,
    EXCEL_TEMPLATES_DIR,
    INPUT_DIR,
    LOGO_DIR,
    MASTER_DIR,
    MASTER_DATA_EXPORT_PATH,
    OUTPUT_DIR,
    PDF_TEMPLATES_DIR,
    SIGNATURE_DIR,
    WORD_TEMPLATES_DIR,
    ensure_directories,
    logo_path,
    signature_path,
)
from app.database.database_service import DatabaseService
from app.models.schemas import MasterData, TemplateRecord
from app.services.ai_mapping_service import AiMappingService
from app.services.customer_catalog_service import (
    CUSTOMER_KEYS,
    CustomerCatalogService,
)
from app.services.form_import_service import FormImportService
from app.services.form_template_service import FormTemplateService
from app.services.mapping_service import MappingService
from app.services.marker_template_service import MarkerTemplateService
from app.services.pdf_field_service import PdfFieldService
from app.services.pdf_fill_service import PdfFillService
from app.plantilla_generator import TemplateGenerator
from app.plantilla_generator.report_writer import GenerationReportRow
from app.ui.import_preview_dialog import ImportPreviewDialog
from app.ui.pdf_learning_dialog import PdfLearningDialog
from app.ui.theme import APP_STYLESHEET


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        ensure_directories()
        self.db = DatabaseService()
        self.db.initialize()
        self.pdf_fields = PdfFieldService()
        self.mapping_service = MappingService()
        self.pdf_fill = PdfFillService()
        self.ai_mapping = AiMappingService()
        self.customer_catalog = CustomerCatalogService()
        self.form_import = FormImportService()
        self.form_templates = FormTemplateService()
        self.template_generator = TemplateGenerator()
        self.marker_templates = MarkerTemplateService()

        self.current_pdf: Path | None = None
        self.current_template_id: int | None = None
        self.current_reference_path: Path | None = None
        self.current_fields: list[str] = []
        self.current_mapping_path: Path | None = None
        self.current_mapping_payload: dict[str, object] | None = None
        self.selected_master_id: int | None = None
        self.selected_certificate_customer: dict[str, object] | None = None
        self._load_customer_catalog_if_available()

        self.setWindowTitle("Madecentro | Gestión inteligente de formularios")
        self.setMinimumSize(1100, 720)
        self.resize(1360, 860)
        self.setStyleSheet(APP_STYLESHEET)

        central = QWidget()
        central_layout = QVBoxLayout(central)
        central_layout.setContentsMargins(0, 0, 0, 0)
        central_layout.setSpacing(0)
        central_layout.addWidget(self._build_brand_header())

        self.tabs = QTabWidget()
        central_layout.addWidget(self.tabs, 1)
        self.setCentralWidget(central)

        self._build_templates_tab()
        self._build_mapping_tab()
        self._build_template_generator_tab()
        self._build_master_tab()
        self._build_certificate_tab()
        self._build_instructions_tab()
        self.refresh_all()
        self._sync_master_data_file()

    def _build_brand_header(self) -> QFrame:
        header = QFrame()
        header.setObjectName("brandHeader")
        header.setFixedHeight(88)
        layout = QHBoxLayout(header)
        layout.setContentsMargins(28, 15, 28, 15)
        layout.setSpacing(16)

        self.logo_label = QLabel()
        self.logo_label.setFixedSize(185, 56)
        self.logo_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self._refresh_logo()

        divider = QFrame()
        divider.setObjectName("brandDivider")
        divider.setFixedSize(3, 42)

        text_layout = QVBoxLayout()
        text_layout.setSpacing(2)
        title = QLabel("Gestión inteligente de formularios")
        title.setObjectName("brandTitle")
        subtitle = QLabel(
            "Automatización documental y administración de datos maestros"
        )
        subtitle.setObjectName("brandSubtitle")
        text_layout.addWidget(title)
        text_layout.addWidget(subtitle)

        layout.addWidget(self.logo_label)
        layout.addWidget(divider)
        layout.addLayout(text_layout)
        layout.addStretch()
        return header

    def _refresh_logo(self) -> None:
        path = logo_path()
        if path:
            pixmap = QPixmap(str(path))
            self.setWindowIcon(QIcon(str(path)))
            self.logo_label.setPixmap(
                pixmap.scaled(
                    self.logo_label.size(),
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation,
                )
            )
            self.logo_label.setText("")
            return
        self.logo_label.setPixmap(QPixmap())
        self.logo_label.setText("MADECENTRO")
        self.logo_label.setStyleSheet(
            "color: #F08419; font-size: 18px; font-weight: 800;"
        )

    def _page_intro(self, title: str, subtitle: str) -> QVBoxLayout:
        layout = QVBoxLayout()
        layout.setSpacing(3)
        title_label = QLabel(title)
        title_label.setObjectName("pageTitle")
        subtitle_label = QLabel(subtitle)
        subtitle_label.setObjectName("pageSubtitle")
        subtitle_label.setWordWrap(True)
        layout.addWidget(title_label)
        layout.addWidget(subtitle_label)
        return layout

    def _card(self) -> tuple[QFrame, QVBoxLayout]:
        frame = QFrame()
        frame.setObjectName("card")
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(18, 16, 18, 16)
        layout.setSpacing(12)
        return frame, layout

    def _section_title(self, text: str) -> QLabel:
        label = QLabel(text)
        label.setObjectName("sectionTitle")
        return label

    def _prepare_table(self, table: QTableWidget) -> None:
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setShowGrid(False)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setHighlightSections(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        table.horizontalHeader().setStretchLastSection(True)

    def _build_master_tab(self) -> None:
        tab = QWidget()
        self.master_tab = tab
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(24, 22, 24, 24)
        layout.setSpacing(16)
        layout.addLayout(
            self._page_intro(
                "Datos maestros",
                "Centraliza la información que se reutiliza al diligenciar cada formulario.",
            )
        )

        edit_card, edit_layout = self._card()
        edit_layout.addWidget(self._section_title("Crear o editar un dato"))
        form = QGridLayout()
        form.setHorizontalSpacing(12)
        form.setVerticalSpacing(8)
        save_btn = QPushButton("Guardar fila")
        clear_btn = QPushButton("Limpiar")
        edit_btn = QPushButton("Editar fila seleccionada")
        delete_btn = QPushButton("Eliminar clave seleccionada")
        save_btn.setProperty("role", "primary")
        delete_btn.setProperty("role", "danger")
        self.master_key_input = QLineEdit()
        self.master_key_input.setPlaceholderText("Ej. nit")
        self.master_value_input = QLineEdit()
        self.master_value_input.setPlaceholderText("Valor del dato maestro")
        self.master_category_input = QLineEdit("general")
        form.addWidget(QLabel("Clave"), 0, 0)
        form.addWidget(QLabel("Valor"), 0, 1)
        form.addWidget(QLabel("Categoría"), 0, 2)
        form.addWidget(self.master_key_input, 1, 0)
        form.addWidget(self.master_value_input, 1, 1)
        form.addWidget(self.master_category_input, 1, 2)
        form.setColumnStretch(1, 2)
        edit_layout.addLayout(form)
        edit_actions = QHBoxLayout()
        edit_actions.addStretch()
        edit_actions.addWidget(edit_btn)
        edit_actions.addWidget(clear_btn)
        edit_actions.addWidget(save_btn)
        edit_layout.addLayout(edit_actions)
        layout.addWidget(edit_card)

        self.master_table = QTableWidget(0, 4)
        self.master_table.setHorizontalHeaderLabels(["ID", "Clave", "Valor", "Categoría"])
        self._prepare_table(self.master_table)
        table_card, table_layout = self._card()
        table_layout.addWidget(self._section_title("Información registrada"))
        filters = QHBoxLayout()
        self.master_search_input = QLineEdit()
        self.master_search_input.setPlaceholderText("Buscar por clave, valor o categoría")
        self.master_category_filter = QComboBox()
        self.master_category_filter.addItem("Todas las categorías", "")
        self.master_count_label = QLabel()
        self.master_count_label.setObjectName("mutedText")
        filters.addWidget(self.master_search_input, 2)
        filters.addWidget(self.master_category_filter, 1)
        filters.addWidget(self.master_count_label)
        table_layout.addLayout(filters)
        table_layout.addWidget(self.master_table)
        self.master_table.horizontalHeader().setSectionResizeMode(
            0,
            QHeaderView.ResizeToContents,
        )
        self.master_table.horizontalHeader().setSectionResizeMode(
            1,
            QHeaderView.ResizeToContents,
        )
        self.master_table.horizontalHeader().setSectionResizeMode(
            2,
            QHeaderView.Stretch,
        )
        self.master_table.horizontalHeader().setSectionResizeMode(
            3,
            QHeaderView.ResizeToContents,
        )
        layout.addWidget(table_card, 1)

        self.master_advanced_toggle = QPushButton("Mostrar avanzado")
        self.master_advanced_toggle.setCheckable(True)
        layout.addWidget(self.master_advanced_toggle)

        self.master_advanced_card, import_layout = self._card()
        import_layout.addWidget(self._section_title("Avanzado"))
        import_actions = QHBoxLayout()
        import_btn = QPushButton("Cargar CSV/XLSX maestros")
        import_form_btn = QPushButton("Revisar datos desde formulario")
        export_xlsx_btn = QPushButton("Exportar maestros XLSX")
        history_btn = QPushButton("Historial de importaciones")
        import_form_btn.setProperty("role", "primary")
        import_actions.addWidget(import_btn)
        import_actions.addWidget(import_form_btn)
        import_actions.addWidget(export_xlsx_btn)
        import_actions.addWidget(history_btn)
        import_actions.addWidget(delete_btn)
        import_actions.addStretch()
        import_layout.addLayout(import_actions)
        sync_note = QLabel(
            "madecentro.db es la base principal; datos_maestros.csv se actualiza "
            "automaticamente cuando guardas o apruebas cambios."
        )
        sync_note.setObjectName("mutedText")
        sync_note.setWordWrap(True)
        import_layout.addWidget(sync_note)
        self.master_advanced_card.setVisible(False)
        layout.addWidget(self.master_advanced_card)

        import_btn.clicked.connect(self.import_master_data)
        import_form_btn.clicked.connect(self.import_completed_form)
        export_xlsx_btn.clicked.connect(self.export_master_data_xlsx)
        history_btn.clicked.connect(self.show_import_history)
        save_btn.clicked.connect(self.save_master_row)
        clear_btn.clicked.connect(self.clear_master_selection)
        edit_btn.clicked.connect(self.edit_selected_master_row)
        delete_btn.clicked.connect(self.delete_selected_master_row)
        self.master_search_input.textChanged.connect(self.apply_master_filters)
        self.master_category_filter.currentTextChanged.connect(self.apply_master_filters)
        self.master_advanced_toggle.toggled.connect(self.toggle_master_advanced)
        self.master_table.cellClicked.connect(self.load_master_row_from_selection)
        self.master_table.itemSelectionChanged.connect(self.load_selected_master_row)
        self.tabs.addTab(tab, "Datos maestros")

    def _build_templates_tab(self) -> None:
        tab = QWidget()
        self.templates_tab = tab
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(24, 22, 24, 24)
        layout.setSpacing(16)
        layout.addLayout(
            self._page_intro(
                "Formularios",
                "Selecciona una plantilla, registra una nueva si hace falta y crea el formulario diligenciado.",
            )
        )

        self.template_table = QTableWidget(0, 5)
        self.template_table.setHorizontalHeaderLabels(
            ["ID", "Nombre", "Formato", "Plantilla vacia", "Referencia"]
        )
        self.template_search_input = QLineEdit()
        self.template_search_input.setPlaceholderText(
            "Buscar por nombre, formato o archivo"
        )
        template_search_btn = QPushButton("Buscar plantilla")
        template_clear_btn = QPushButton("Limpiar")
        self.template_count_label = QLabel()
        self.template_count_label.setObjectName("mutedText")
        self.template_list = QListWidget()
        self.template_list.setObjectName("templateList")
        self.template_list.setSpacing(4)
        self.field_table = QTableWidget(0, 4)
        self.field_table.setHorizontalHeaderLabels(
            ["Campo", "Tipo", "Ubicacion", "Valor de referencia"]
        )
        self._prepare_table(self.template_table)
        self._prepare_table(self.field_table)

        main_area = QHBoxLayout()
        main_area.setSpacing(16)

        templates_card, templates_layout = self._card()
        templates_layout.addWidget(self._section_title("1. Elegir plantilla"))
        search_actions = QHBoxLayout()
        search_actions.addWidget(self.template_search_input, 1)
        search_actions.addWidget(template_search_btn)
        search_actions.addWidget(template_clear_btn)
        templates_layout.addLayout(search_actions)
        templates_layout.addWidget(self.template_count_label)
        self.template_list.setMinimumHeight(340)
        templates_layout.addWidget(self.template_list, 1)
        main_area.addWidget(templates_card, 3)

        side_layout = QVBoxLayout()
        side_layout.setSpacing(16)

        action_card, action_layout = self._card()
        action_layout.addWidget(self._section_title("2. Crear formulario"))
        self.selected_pdf_label = QLabel("Plantilla: sin seleccionar")
        self.selected_mapping_label = QLabel("Mapeo: sin guardar")
        self.logo_status_label = QLabel()
        self.signature_status_label = QLabel()
        self.selected_pdf_label.setObjectName("summaryText")
        self.selected_mapping_label.setObjectName("summaryText")
        self.logo_status_label.setObjectName("summaryText")
        self.signature_status_label.setObjectName("summaryText")
        fill_actions = QHBoxLayout()
        fill_btn = QPushButton("Crear formulario diligenciado")
        open_btn = QPushButton("Abrir carpeta de salida")
        signature_btn = QPushButton("Elegir firma")
        fill_btn.setProperty("role", "primary")
        fill_actions.addWidget(fill_btn)
        fill_actions.addWidget(open_btn)
        fill_actions.addStretch()
        fill_actions.addWidget(signature_btn)
        action_layout.addLayout(fill_actions)
        side_layout.addWidget(action_card)

        register_card, register_layout = self._card()
        register_layout.addWidget(self._section_title("Agregar plantilla"))
        actions = QGridLayout()
        learn_pdf_btn = QPushButton("Registrar PDF editable")
        load_btn = QPushButton("Registrar plantilla con ejemplo")
        list_btn = QPushButton("Ver campos")
        delete_btn = QPushButton("Eliminar plantilla seleccionada")
        load_btn.setProperty("role", "primary")
        delete_btn.setProperty("role", "danger")
        self.template_name_input = QLineEdit()
        self.template_name_input.setPlaceholderText("Ej. Formulario de vinculación")
        actions.addWidget(QLabel("Nombre"), 0, 0, 1, 2)
        actions.addWidget(self.template_name_input, 1, 0, 1, 2)
        actions.addWidget(load_btn, 2, 0)
        actions.addWidget(learn_pdf_btn, 2, 1)
        actions.setColumnStretch(0, 1)
        actions.setColumnStretch(1, 1)
        register_layout.addLayout(actions)
        side_layout.addWidget(register_card)

        fields_card, fields_layout = self._card()
        fields_layout.addWidget(self._section_title("Avanzado"))
        detail_actions = QHBoxLayout()
        detail_actions.addWidget(list_btn)
        detail_actions.addWidget(delete_btn)
        detail_actions.addStretch()
        fields_layout.addLayout(detail_actions)
        self.field_table.setMaximumHeight(240)
        fields_layout.addWidget(self.field_table)
        side_layout.addWidget(fields_card, 1)

        main_area.addLayout(side_layout, 2)
        layout.addLayout(main_area, 1)

        learn_pdf_btn.clicked.connect(self.learn_pdf_template)
        load_btn.clicked.connect(self.load_template)
        list_btn.clicked.connect(self.list_pdf_fields)
        delete_btn.clicked.connect(self.delete_selected_template)
        template_search_btn.clicked.connect(self.apply_template_search)
        template_clear_btn.clicked.connect(self.clear_template_search)
        self.template_search_input.returnPressed.connect(self.apply_template_search)
        self.template_search_input.textChanged.connect(self.apply_template_search)
        self.template_list.itemClicked.connect(self.select_template_from_list)
        self.template_table.cellClicked.connect(self.select_template_from_table)
        fill_btn.clicked.connect(self.fill_pdf)
        open_btn.clicked.connect(self.open_output_folder)
        signature_btn.clicked.connect(self.choose_signature_file)
        self._refresh_asset_status()
        self.tabs.addTab(tab, "Formularios")

    def _build_mapping_tab(self) -> None:
        tab = QWidget()
        self.mapping_tab = tab
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(24, 22, 24, 24)
        layout.setSpacing(16)
        layout.addLayout(
            self._page_intro(
                "Mapeo",
                "Relaciona cada campo del formulario con su dato maestro correspondiente.",
            )
        )

        actions_card, actions_layout = self._card()
        actions_layout.addWidget(self._section_title("Acciones de mapeo"))
        actions = QHBoxLayout()
        self.create_mapping_btn = QPushButton("Crear mapeo vacio")
        self.suggest_mapping_btn = QPushButton("Sugerir con diccionario")
        self.complete_ai_mapping_btn = QPushButton("Completar faltantes con IA")
        load_btn = QPushButton("Cargar mapeo")
        save_btn = QPushButton("Guardar mapeo")
        save_btn.setProperty("role", "primary")
        self.mapping_name_input = QLineEdit("mapeo_madecentro")
        actions.addWidget(self.create_mapping_btn)
        actions.addWidget(self.suggest_mapping_btn)
        actions.addWidget(self.complete_ai_mapping_btn)
        actions.addWidget(load_btn)
        actions.addWidget(QLabel("Nombre"))
        actions.addWidget(self.mapping_name_input)
        actions.addWidget(save_btn)
        actions_layout.addLayout(actions)
        layout.addWidget(actions_card)

        self.mapping_table = QTableWidget(0, 3)
        self.mapping_table.setHorizontalHeaderLabels(
            ["Campo formulario", "Campo maestro", "Valor actual"]
        )
        self._prepare_table(self.mapping_table)
        mapping_card, mapping_layout = self._card()
        mapping_layout.addWidget(self._section_title("Relaciones del formulario"))
        mapping_layout.addWidget(self.mapping_table)
        layout.addWidget(mapping_card, 1)

        self.create_mapping_btn.clicked.connect(self.create_empty_mapping)
        self.suggest_mapping_btn.clicked.connect(self.suggest_mapping)
        self.complete_ai_mapping_btn.clicked.connect(self.complete_mapping_with_ai)
        load_btn.clicked.connect(self.load_mapping)
        save_btn.clicked.connect(self.save_mapping)
        self._set_mapping_actions_enabled(False)
        self.tabs.addTab(tab, "Mapeo")

    def _build_template_generator_tab(self) -> None:
        tab = QWidget()
        self.template_generator_tab = tab
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(24, 22, 24, 24)
        layout.setSpacing(16)
        layout.addLayout(
            self._page_intro(
                "Generador de plantillas",
                "Crea plantillas reutilizables desde formularios diligenciados y conserva una copia de referencia.",
            )
        )

        config_card, config_layout = self._card()
        self.generator_input_input = QLineEdit(str(INPUT_DIR))
        self.generator_dictionary_input = QLineEdit(
            str(self._default_template_generator_dictionary())
        )
        self.generator_output_input = QLineEdit(str(BASE_DIR / "plantillas_generadas"))
        self.generator_mode_combo = QComboBox()
        self.generator_mode_combo.addItem("Marcadores", "markers")
        self.generator_mode_combo.addItem("Borrar datos", "blank")

        input_btn = QPushButton("Elegir")
        dictionary_btn = QPushButton("Elegir")
        output_btn = QPushButton("Elegir")

        primary_label = QLabel("Flujo principal")
        primary_label.setObjectName("sectionTitle")
        config_layout.addWidget(primary_label)

        actions = QHBoxLayout()
        add_files_btn = QPushButton("Agregar documentos")
        learn_completed_btn = QPushButton("Crear plantilla desde diligenciado")
        open_output_btn = QPushButton("Abrir salida")
        learn_completed_btn.setProperty("role", "primary")
        actions.addWidget(add_files_btn)
        actions.addWidget(learn_completed_btn)
        actions.addWidget(open_output_btn)
        actions.addStretch()
        config_layout.addLayout(actions)

        self.generator_status_label = QLabel(
            "Flujo recomendado: agregue un formulario diligenciado y cree la plantilla desde ese archivo."
        )
        self.generator_status_label.setObjectName("mutedText")
        self.generator_status_label.setWordWrap(True)
        config_layout.addWidget(self.generator_status_label)

        self.generator_advanced_toggle = QPushButton("Mostrar avanzado")
        self.generator_advanced_toggle.setCheckable(True)
        config_layout.addWidget(self.generator_advanced_toggle)

        self.generator_advanced_card, advanced_layout = self._card()
        advanced_layout.addWidget(self._section_title("Avanzado / lote"))
        form = QGridLayout()
        form.setHorizontalSpacing(10)
        form.setVerticalSpacing(9)
        form.addWidget(QLabel("Carpeta de entrada"), 0, 0)
        form.addWidget(self.generator_input_input, 0, 1)
        form.addWidget(input_btn, 0, 2)
        form.addWidget(QLabel("Carpeta de salida"), 1, 0)
        form.addWidget(self.generator_output_input, 1, 1)
        form.addWidget(output_btn, 1, 2)
        form.addWidget(QLabel("Diccionario"), 2, 0)
        form.addWidget(self.generator_dictionary_input, 2, 1)
        form.addWidget(dictionary_btn, 2, 2)
        form.addWidget(QLabel("Modo lote"), 3, 0)
        form.addWidget(self.generator_mode_combo, 3, 1)
        form.setColumnStretch(1, 1)
        advanced_layout.addLayout(form)

        advanced_actions = QHBoxLayout()
        run_btn = QPushButton("Limpiar documentos en lote")
        register_btn = QPushButton("Registrar archivo limpio como plantilla")
        advanced_actions.addWidget(run_btn)
        advanced_actions.addWidget(register_btn)
        advanced_actions.addStretch()
        advanced_layout.addLayout(advanced_actions)
        self.generator_advanced_card.setVisible(False)
        layout.addWidget(config_card)
        layout.addWidget(self.generator_advanced_card)

        self.generator_report_table = QTableWidget(0, 6)
        self.generator_report_table.setHorizontalHeaderLabels(
            [
                "Archivo",
                "Tipo",
                "Estado",
                "Reemplazos",
                "Valores reemplazados",
                "Errores",
            ]
        )
        self._prepare_table(self.generator_report_table)
        self.generator_report_card, report_layout = self._card()
        report_layout.addWidget(self._section_title("Reporte del ultimo proceso"))
        report_layout.addWidget(self.generator_report_table)
        self.generator_report_card.setVisible(False)
        layout.addWidget(self.generator_report_card, 1)

        input_btn.clicked.connect(self.choose_template_generator_input)
        dictionary_btn.clicked.connect(self.choose_template_generator_dictionary)
        output_btn.clicked.connect(self.choose_template_generator_output)
        add_files_btn.clicked.connect(self.add_template_generator_documents)
        run_btn.clicked.connect(self.generate_templates_from_ui)
        learn_completed_btn.clicked.connect(self.create_template_from_completed_ui)
        register_btn.clicked.connect(self.register_generated_template_from_ui)
        open_output_btn.clicked.connect(self.open_template_generator_output)
        self.generator_advanced_toggle.toggled.connect(
            self.toggle_template_generator_advanced
        )
        self.tabs.addTab(tab, "Generador")

    def _build_certificate_tab(self) -> None:
        tab = QWidget()
        self.certificate_tab = tab
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(24, 22, 24, 24)
        layout.setSpacing(16)
        layout.addLayout(
            self._page_intro(
                "Crear certificado",
                "Busca un cliente por razón social o NIT y genera su certificado automáticamente.",
            )
        )

        search_card, search_layout = self._card()
        search_layout.addWidget(self._section_title("Buscar cliente"))
        search_actions = QHBoxLayout()
        self.certificate_search_input = QLineEdit()
        self.certificate_search_input.setPlaceholderText(
            "Ej. Maderas del Norte o 900456789"
        )
        search_btn = QPushButton("Buscar")
        new_customer_btn = QPushButton("Nuevo cliente certificado")
        import_customers_btn = QPushButton("Importar clientes certificados")
        search_btn.setProperty("role", "primary")
        search_actions.addWidget(self.certificate_search_input, 1)
        search_actions.addWidget(search_btn)
        search_actions.addWidget(new_customer_btn)
        search_actions.addWidget(import_customers_btn)
        search_layout.addLayout(search_actions)
        self.certificate_catalog_status = QLabel()
        self.certificate_catalog_status.setObjectName("mutedText")
        search_layout.addWidget(self.certificate_catalog_status)
        layout.addWidget(search_card)

        self.certificate_results_table = QTableWidget(0, 4)
        self.certificate_results_table.setHorizontalHeaderLabels(
            ["ID", "Razón social", "NIT", "Año de vinculación"]
        )
        self._prepare_table(self.certificate_results_table)
        results_card, results_layout = self._card()
        results_layout.addWidget(self._section_title("Clientes encontrados"))
        results_layout.addWidget(self.certificate_results_table)
        layout.addWidget(results_card, 1)

        action_card, action_layout = self._card()
        action_layout.addWidget(self._section_title("Generar certificado"))
        self.selected_certificate_label = QLabel("Cliente: sin seleccionar")
        self.selected_certificate_label.setObjectName("statusMissing")
        action_layout.addWidget(self.selected_certificate_label)
        certificate_actions = QHBoxLayout()
        create_btn = QPushButton("Crear certificado")
        open_output_btn = QPushButton("Abrir carpeta de salida")
        create_btn.setProperty("role", "primary")
        certificate_actions.addWidget(create_btn)
        certificate_actions.addWidget(open_output_btn)
        certificate_actions.addStretch()
        action_layout.addLayout(certificate_actions)
        layout.addWidget(action_card)

        search_btn.clicked.connect(self.search_certificate_customers)
        new_customer_btn.clicked.connect(self.create_certificate_customer)
        import_customers_btn.clicked.connect(self.import_certificate_customers)
        self.certificate_search_input.returnPressed.connect(
            self.search_certificate_customers
        )
        self.certificate_results_table.cellClicked.connect(
            self.select_certificate_customer
        )
        create_btn.clicked.connect(self.create_certificate)
        open_output_btn.clicked.connect(self.open_output_folder)
        self.tabs.addTab(tab, "Crear certificado")

    def _build_instructions_tab(self) -> None:
        tab = QWidget()
        outer_layout = QVBoxLayout(tab)
        outer_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(24, 22, 24, 28)
        layout.setSpacing(16)
        layout.addLayout(
            self._page_intro(
                "Instructivo de uso",
                "Guía rápida para crear formularios y mantener la información actualizada.",
            )
        )

        intro_card, intro_layout = self._card()
        intro_layout.addWidget(self._section_title("Uso diario"))
        intro = QLabel(
            "Para el trabajo normal use Formularios para generar documentos y Generador "
            "solo cuando necesite registrar una plantilla nueva desde un archivo diligenciado."
        )
        intro.setWordWrap(True)
        intro_layout.addWidget(intro)
        layout.addWidget(intro_card)

        steps = [
            (
                "1. Crear un formulario",
                [
                    "Abra Formularios y busque la plantilla.",
                    "Seleccione la plantilla correcta.",
                    "Revise que el mapeo cargado y la firma activa sean los esperados.",
                    "Pulse Crear formulario diligenciado.",
                    "Abra la salida y revise el documento final antes de enviarlo.",
                ],
                "forms",
            ),
            (
                "2. Registrar una plantilla nueva",
                [
                    "Use Generador cuando el formulario no exista en Formularios.",
                    "Pulse Agregar documentos y seleccione un formulario diligenciado.",
                    "Pulse Crear plantilla desde diligenciado; la app crea una copia diligenciada y una vacía.",
                    "Revise el nombre sugerido, luego ajuste el mapeo si faltan campos.",
                    "En Formularios, Agregar plantilla es el método manual: úselo solo si ya tiene el formulario vacío y el mismo formulario diligenciado.",
                    "Registrar plantilla con ejemplo compara esos dos archivos y crea el mapeo inicial.",
                ],
                "generator",
            ),
            (
                "3. Revisar el mapeo",
                [
                    "Use Mapeo cuando una plantilla no complete bien sus campos.",
                    "Sugerir con diccionario usa memoria local y no depende de internet.",
                    "Completar faltantes con IA solo intenta llenar campos que sigan vacíos.",
                    "Revise la columna Valor actual para confirmar qué dato se insertará.",
                    "Guarde el mapeo cuando las relaciones sean correctas.",
                ],
                "mapping",
            ),
            (
                "4. Mantener datos maestros",
                [
                    "Use Datos maestros para buscar, editar o guardar claves oficiales de Madecentro.",
                    "El buscador filtra por clave, valor o categoría; Limpiar deja el formulario listo para un dato nuevo.",
                    "Mostrar avanzado contiene cargas masivas, revisión desde formulario, exportación, historial y eliminación.",
                    "No elimine una clave si la confirmación indica que se usa en plantillas.",
                ],
                "masters",
            ),
            (
                "5. Crear certificados",
                [
                    "Use Crear certificado para buscar clientes certificados por razón social o NIT.",
                    "Importar clientes certificados está en esta pestaña porque alimenta solo este flujo.",
                    "Seleccione el cliente correcto y pulse Crear certificado.",
                ],
                "certificate",
            ),
            (
                "6. Avanzado y control final",
                [
                    "En Formularios use Elegir firma si necesita cambiar la firma activa.",
                    "En Generador, Mostrar avanzado permite limpiar documentos en lote y registrar archivos limpios.",
                    "datos_maestros.csv se sincroniza automáticamente al guardar o aprobar cambios.",
                    "Conserve siempre los originales y revise fecha, NIT, razón social, bancos, casillas y firma.",
                ],
                "output",
            ),
        ]
        for title, items, action_key in steps:
            card, card_layout = self._card()
            card_layout.addWidget(self._section_title(title))
            for item in items:
                label = QLabel(f"• {item}")
                label.setWordWrap(True)
                label.setObjectName("instructionStep")
                card_layout.addWidget(label)
            actions = QHBoxLayout()
            actions.addStretch()
            if action_key == "forms":
                output_button = QPushButton("Abrir salida")
                output_button.clicked.connect(
                    lambda: self._open_directory(OUTPUT_DIR)
                )
                button = QPushButton("Ir a Formularios")
                button.setProperty("role", "primary")
                button.clicked.connect(lambda: self.tabs.setCurrentWidget(self.templates_tab))
                actions.addWidget(output_button)
                actions.addWidget(button)
            elif action_key == "input":
                input_button = QPushButton("Abrir entrada")
                input_button.clicked.connect(lambda: self._open_directory(INPUT_DIR))
                button = QPushButton("Ir a Formularios")
                button.setProperty("role", "primary")
                button.clicked.connect(lambda: self.tabs.setCurrentWidget(self.templates_tab))
                actions.addWidget(input_button)
                actions.addWidget(button)
            elif action_key == "mapping":
                button = QPushButton("Ir a Mapeo")
                button.setProperty("role", "primary")
                button.clicked.connect(lambda: self.tabs.setCurrentWidget(self.mapping_tab))
                actions.addWidget(button)
            elif action_key == "generator":
                button = QPushButton("Ir a Generador")
                button.setProperty("role", "primary")
                button.clicked.connect(
                    lambda: self.tabs.setCurrentWidget(self.template_generator_tab)
                )
                actions.addWidget(button)
            elif action_key == "masters":
                button = QPushButton("Ir a Datos maestros")
                button.setProperty("role", "primary")
                button.clicked.connect(lambda: self.tabs.setCurrentWidget(self.master_tab))
                actions.addWidget(button)
            elif action_key == "certificate":
                button = QPushButton("Ir a Crear certificado")
                button.setProperty("role", "primary")
                button.clicked.connect(lambda: self.tabs.setCurrentWidget(self.certificate_tab))
                actions.addWidget(button)
            elif action_key == "output":
                button = QPushButton("Abrir salida")
                button.setProperty("role", "primary")
                button.clicked.connect(
                    lambda: self._open_directory(OUTPUT_DIR)
                )
                actions.addWidget(button)
            card_layout.addLayout(actions)
            layout.addWidget(card)

        warning = QLabel(
            "Importante: conserve siempre los archivos originales. No edite manualmente "
            "las plantillas guardadas dentro de la carpeta plantillas."
        )
        warning.setObjectName("instructionWarning")
        warning.setWordWrap(True)
        layout.addWidget(warning)
        layout.addStretch()

        scroll.setWidget(content)
        outer_layout.addWidget(scroll)
        self.tabs.addTab(tab, "Instructivo")

    def refresh_all(self) -> None:
        self.refresh_master_table()
        self.refresh_template_table()
        self.refresh_certificate_status()
        self.refresh_selected_template()
        self.refresh_mapping_combos()
        self.refresh_selected_mapping()
        self._refresh_logo()
        self._refresh_asset_status()

    def _refresh_asset_status(self) -> None:
        if not hasattr(self, "logo_status_label"):
            return
        logo = logo_path()
        signature = signature_path()
        self.logo_status_label.setText(
            f"Logo: {logo.name}" if logo else "Logo: archivo pendiente"
        )
        self.logo_status_label.setObjectName("summaryText")
        self.signature_status_label.setText(
            f"Firma: {signature.name}"
            if signature
            else "Firma: archivo pendiente"
        )
        self.signature_status_label.setObjectName("summaryText")
        self.logo_status_label.style().unpolish(self.logo_status_label)
        self.logo_status_label.style().polish(self.logo_status_label)
        self.signature_status_label.style().unpolish(self.signature_status_label)
        self.signature_status_label.style().polish(self.signature_status_label)

    def refresh_master_table(self) -> None:
        rows = self.db.list_master_data()
        self.master_rows = rows
        self._refresh_master_category_filter(rows)
        self.apply_master_filters()

    def _refresh_master_category_filter(self, rows: list[dict[str, object]]) -> None:
        current = self.master_category_filter.currentData()
        categories = sorted(
            {
                str(row.get("categoria") or "").strip()
                for row in rows
                if str(row.get("categoria") or "").strip()
            }
        )
        self.master_category_filter.blockSignals(True)
        self.master_category_filter.clear()
        self.master_category_filter.addItem("Todas las categorías", "")
        for category in categories:
            self.master_category_filter.addItem(category, category)
        index = self.master_category_filter.findData(current)
        self.master_category_filter.setCurrentIndex(index if index >= 0 else 0)
        self.master_category_filter.blockSignals(False)

    def apply_master_filters(self, *_args: object) -> None:
        rows = getattr(self, "master_rows", [])
        query = self.master_search_input.text().strip().casefold()
        category = str(self.master_category_filter.currentData() or "")
        filtered_rows = []
        for row in rows:
            row_category = str(row.get("categoria") or "")
            searchable = " ".join(
                str(row.get(key) or "")
                for key in ("clave", "valor", "categoria")
            ).casefold()
            if category and row_category != category:
                continue
            if query and query not in searchable:
                continue
            filtered_rows.append(row)

        self.master_table.setRowCount(len(filtered_rows))
        for row_index, row in enumerate(filtered_rows):
            for col, key in enumerate(["id", "clave", "valor", "categoria"]):
                item = QTableWidgetItem(str(row[key]))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.master_table.setItem(row_index, col, item)
        total = len(rows)
        shown = len(filtered_rows)
        self.master_count_label.setText(f"{shown} de {total}")

    def toggle_master_advanced(self, checked: bool) -> None:
        self.master_advanced_card.setVisible(checked)
        self.master_advanced_toggle.setText(
            "Ocultar avanzado" if checked else "Mostrar avanzado"
        )

    def refresh_template_table(self) -> None:
        rows = self._registered_template_rows()
        self.template_table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            values = [
                row["id"],
                self._template_display_name(row),
                row["formato"].upper(),
                row["ruta_pdf"],
                row["ruta_referencia"],
            ]
            for col, value in enumerate(values):
                self.template_table.setItem(
                    row_index,
                    col,
                    QTableWidgetItem(str(value)),
                )
        self.apply_template_search()

    def _registered_template_rows(self) -> list[dict[str, object]]:
        certificate_ids = self._certificate_template_ids()
        return [
            row
            for row in self.db.list_templates()
            if int(row["id"]) not in certificate_ids
            and self._template_file_exists(row)
        ]

    def _template_file_exists(self, row: dict[str, object]) -> bool:
        path_text = str(row.get("ruta_pdf") or "").strip()
        return bool(path_text) and Path(path_text).is_file()

    def apply_template_search(self, _text: str | None = None) -> None:
        if not hasattr(self, "template_list"):
            return
        query = self.template_search_input.text().strip().casefold()
        rows = self._registered_template_rows()
        if query:
            rows = [
                row
                for row in rows
                if self._template_matches_query(row, query)
            ]
        self._populate_template_list(rows)

    def clear_template_search(self) -> None:
        self.template_search_input.clear()
        self.apply_template_search()

    def focus_template_search(self) -> None:
        if hasattr(self, "templates_tab"):
            self.tabs.setCurrentWidget(self.templates_tab)
        if hasattr(self, "template_search_input"):
            self.template_search_input.setFocus()
            self.template_search_input.selectAll()

    def _template_matches_query(
        self,
        row: dict[str, object],
        query: str,
    ) -> bool:
        searchable = " ".join(
            [
                self._template_display_name(row),
                str(row.get("nombre", "")),
                str(row.get("formato", "")),
                Path(str(row.get("ruta_pdf", ""))).name,
                Path(str(row.get("ruta_referencia", ""))).name,
            ]
        ).casefold()
        return query in searchable

    def _template_display_name(self, row: dict[str, object]) -> str:
        name = str(row.get("nombre") or "").strip()
        if not name:
            name = Path(str(row.get("ruta_pdf") or "")).stem
        return self._clean_template_display_name(name)

    def _clean_template_display_name(self, value: str) -> str:
        cleaned = value.strip()
        path = Path(cleaned)
        if path.suffix.lower() in {".pdf", ".xlsx", ".docx"}:
            cleaned = path.stem
        cleaned = re.sub(
            r"(?i)(?:[\s_-]+)(vacio|vacío|plantilla|formato)$",
            "",
            cleaned,
        ).strip()
        return cleaned or value.strip() or "Plantilla"

    def _populate_template_list(self, rows: list[dict[str, object]]) -> None:
        self.template_list.clear()
        self.template_count_label.setText(
            f"Plantillas visibles: {len(rows)}"
        )
        for row in rows:
            template_id = int(row["id"])
            name = self._template_display_name(row)
            fmt = str(row["formato"]).upper()
            item = QListWidgetItem(f"{name}   ·   {fmt}")
            item.setData(Qt.UserRole, template_id)
            item.setToolTip(str(row["ruta_pdf"]))
            self.template_list.addItem(item)
            if self.current_template_id == template_id:
                item.setSelected(True)
                self.template_list.scrollToItem(item)

    def _sync_template_selection(self) -> None:
        if not self.current_template_id:
            return
        if hasattr(self, "template_list"):
            for row in range(self.template_list.count()):
                item = self.template_list.item(row)
                selected = int(item.data(Qt.UserRole)) == self.current_template_id
                item.setSelected(selected)
                if selected:
                    self.template_list.scrollToItem(item)
        if hasattr(self, "template_table"):
            for row in range(self.template_table.rowCount()):
                id_item = self.template_table.item(row, 0)
                if id_item and int(id_item.text()) == self.current_template_id:
                    self.template_table.selectRow(row)
                    self.template_table.scrollToItem(id_item)
                    break

    def refresh_selected_template(self) -> None:
        certificate_ids = self._certificate_template_ids()
        templates = [
            item
            for item in self.db.list_templates()
            if int(item["id"]) not in certificate_ids
            and self._template_file_exists(item)
        ]
        if not templates:
            self._clear_template_state()
            return
        self._set_current_template(templates[0])

    def refresh_certificate_status(self) -> None:
        if not hasattr(self, "certificate_catalog_status"):
            return
        customer_count = self.db.count_certificate_customers()
        context = self._certificate_template_context()
        template_text = (
            f"Plantilla: {context[0]['nombre']}"
            if context
            else "Plantilla de certificado: pendiente"
        )
        self.certificate_catalog_status.setText(
            f"Clientes disponibles: {customer_count} | {template_text}"
        )

    def refresh_mapping_combos(self) -> None:
        master_keys = self._available_mapping_keys()
        mapped_keys = (
            list(self.form_templates.mapping(self.current_mapping_payload).values())
            if self.current_mapping_payload
            else []
        )
        available_keys = sorted(
            {key for key in master_keys + mapped_keys if key}
        )
        for row in range(self.mapping_table.rowCount()):
            combo = self.mapping_table.cellWidget(row, 1)
            if isinstance(combo, QComboBox):
                current = combo.currentText()
                combo.clear()
                combo.addItem("")
                combo.addItems(available_keys)
                combo.setCurrentText(current)
                self._set_mapping_value_preview(row, current)

    def refresh_selected_mapping(self) -> None:
        mappings = self.db.list_mappings(self.current_template_id)
        if not mappings:
            self.current_mapping_path = None
            self.current_mapping_payload = None
            self.selected_mapping_label.setText("Mapeo: sin guardar")
            return
        latest_path = next(
            (
                Path(str(mapping["ruta_json"]))
                for mapping in mappings
                if Path(str(mapping["ruta_json"])).is_file()
            ),
            Path(str(mappings[0]["ruta_json"])),
        )
        if latest_path.exists():
            try:
                self.current_mapping_payload = self.mapping_service.load_payload(
                    latest_path
                )
            except (OSError, ValueError):
                self.current_mapping_payload = None
            self.current_mapping_path = latest_path
            self.selected_mapping_label.setText(f"Mapeo: {latest_path}")

    def import_master_data(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(self, "Importar datos maestros", "data/maestros", "Datos (*.csv *.xlsx)")
        if not file_path:
            return
        try:
            items = self._read_master_file(Path(file_path))
            if not items:
                raise ValueError("El archivo no contiene filas validas.")
            count = self.db.bulk_upsert_master_data(items)
            self._sync_master_data_file()
        except (OSError, ValueError, ValidationError) as exc:
            self._show_error("No se pudieron importar los datos", exc)
            return
        self.refresh_all()
        QMessageBox.information(
            self,
            "Maestros actualizados",
            f"Datos cargados o actualizados: {count}\n"
            f"CSV sincronizado: {MASTER_DATA_EXPORT_PATH}",
        )

    def choose_template_generator_input(self) -> None:
        directory = QFileDialog.getExistingDirectory(
            self,
            "Elegir carpeta de documentos diligenciados",
            self.generator_input_input.text().strip() or str(BASE_DIR),
        )
        if directory:
            self.generator_input_input.setText(directory)

    def choose_template_generator_dictionary(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Elegir diccionario Madecentro",
            str(MASTER_DIR),
            "Diccionario (*.xlsx *.xlsm *.json *.csv)",
        )
        if file_path:
            self.generator_dictionary_input.setText(file_path)

    def choose_template_generator_output(self) -> None:
        directory = QFileDialog.getExistingDirectory(
            self,
            "Elegir carpeta de salida",
            self.generator_output_input.text().strip() or str(BASE_DIR),
        )
        if directory:
            self.generator_output_input.setText(directory)

    def toggle_template_generator_advanced(self, checked: bool) -> None:
        self.generator_advanced_card.setVisible(checked)
        self.generator_report_card.setVisible(
            checked and self.generator_report_table.rowCount() > 0
        )
        self.generator_advanced_toggle.setText(
            "Ocultar avanzado" if checked else "Mostrar avanzado"
        )

    def add_template_generator_documents(self) -> None:
        input_dir = Path(self.generator_input_input.text().strip())
        input_dir.mkdir(parents=True, exist_ok=True)
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Agregar documentos al generador",
            str(BASE_DIR),
            "Documentos (*.pdf *.docx *.xlsx *.xlsm)",
        )
        if not file_paths:
            return

        copied = 0
        skipped = 0
        try:
            for file_path in file_paths:
                source = Path(file_path)
                if source.suffix.lower() not in {".pdf", ".docx", ".xlsx", ".xlsm"}:
                    skipped += 1
                    continue
                target = self._unique_template_generator_input_path(
                    input_dir / source.name
                )
                if source.resolve() == target.resolve():
                    skipped += 1
                    continue
                shutil.copy2(source, target)
                copied += 1
        except OSError as exc:
            self._show_error("No se pudieron agregar los documentos", exc)
            return

        self.generator_input_input.setText(str(input_dir))
        self.generator_status_label.setText(
            f"Documentos agregados a {input_dir}: {copied} | Omitidos: {skipped}"
        )

    def _unique_template_generator_input_path(self, target: Path) -> Path:
        if not target.exists():
            return target
        index = 2
        while True:
            candidate = target.with_name(f"{target.stem}_{index}{target.suffix}")
            if not candidate.exists():
                return candidate
            index += 1

    def generate_templates_from_ui(self) -> None:
        input_dir = Path(self.generator_input_input.text().strip())
        dictionary_path = Path(self.generator_dictionary_input.text().strip())
        output_dir = Path(self.generator_output_input.text().strip())
        mode = str(self.generator_mode_combo.currentData() or "markers")
        if not input_dir.is_dir():
            QMessageBox.warning(
                self,
                "Entrada no valida",
                "Seleccione una carpeta de entrada existente.",
            )
            return
        if not dictionary_path.is_file():
            QMessageBox.warning(
                self,
                "Diccionario no valido",
                "Seleccione un diccionario .xlsx, .xlsm, .json o .csv.",
            )
            return
        try:
            rows = self.template_generator.run(
                input_dir=input_dir,
                dictionary_path=dictionary_path,
                output_dir=output_dir,
                mode=mode,
            )
        except (OSError, ValueError) as exc:
            self._show_error("No se pudieron generar las plantillas", exc)
            return

        self._apply_signature_to_generated_files(rows)
        self.template_generator.report_writer.write(output_dir, rows)
        self._populate_template_generator_report(rows)
        self.generator_advanced_toggle.setChecked(True)
        self.generator_report_card.setVisible(True)
        ok = sum(1 for row in rows if row.estado == "ok")
        errors = sum(1 for row in rows if row.estado == "error")
        report_path = output_dir / "reporte_generacion_plantillas.xlsx"
        self.generator_status_label.setText(
            f"Procesados: {len(rows)} | OK: {ok} | Errores: {errors} | Reporte: {report_path}"
        )
        QMessageBox.information(
            self,
            "Generacion finalizada",
            f"Procesados: {len(rows)}\nOK: {ok}\nErrores: {errors}\n\nReporte:\n{report_path}",
        )

    def _apply_signature_to_generated_files(
        self,
        rows: list[GenerationReportRow],
    ) -> None:
        signature = signature_path()
        if not signature:
            return
        for row in rows:
            if row.estado != "ok":
                continue
            output_path = Path(row.ruta_salida or "")
            if not output_path.is_file():
                continue
            try:
                suffix = output_path.suffix.lower()
                if suffix == ".xlsx":
                    self.form_templates.excel_signature.apply_signature(
                        output_path,
                        signature,
                    )
                elif suffix == ".pdf":
                    self.form_templates.pdf_signature.apply_signature(
                        output_path,
                        signature,
                    )
                elif suffix == ".docx":
                    self.form_templates.word_signature.apply_signature(
                        output_path,
                        signature,
                    )
            except (OSError, ValueError, RuntimeError) as exc:
                row.estado = "error"
                row.errores = (
                    f"{row.errores}; Firma: {exc}"
                    if row.errores
                    else f"Firma: {exc}"
                )

    def open_template_generator_output(self) -> None:
        output_dir = Path(self.generator_output_input.text().strip())
        self._open_directory(output_dir)

    def create_template_from_completed_ui(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar formulario diligenciado",
            self.generator_input_input.text().strip() or str(BASE_DIR),
            "Excel diligenciado (*.xlsx)",
        )
        if not file_path:
            return
        completed_source = Path(file_path)
        if completed_source.suffix.lower() != ".xlsx":
            QMessageBox.warning(
                self,
                "Formato no soportado",
                "Por ahora este flujo automatico soporta XLSX.",
            )
            return
        dictionary_path = Path(self.generator_dictionary_input.text().strip())
        if not dictionary_path.is_file():
            QMessageBox.warning(
                self,
                "Diccionario no valido",
                "Seleccione un diccionario antes de crear la plantilla.",
            )
            return

        suggested_name = self._suggest_template_name_from_completed(
            completed_source
        )
        name, accepted = QInputDialog.getText(
            self,
            "Nombre de plantilla",
            "Nombre del proveedor/cliente o de la plantilla:",
            text=suggested_name,
        )
        if not accepted:
            return
        name = name.strip() or completed_source.stem

        safe_template_name = self._safe_output_name(name)
        work_dir = self._template_storage_dir(
            Path(self.generator_output_input.text().strip()),
            name,
        )
        empty_path = work_dir / f"{safe_template_name}_vacio.xlsx"
        completed_path = work_dir / f"{safe_template_name}_diligenciado.xlsx"
        template_id: int | None = None
        try:
            work_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(completed_source, completed_path)
            self.template_generator.excel_cleaner.clean(
                completed_source,
                empty_path,
                self.template_generator.dictionary_loader.load(
                    dictionary_path,
                    mode="blank",
                ),
            )
            template_id, mapping_path, payload, target = self._register_template_pair(
                name=name,
                empty_source=empty_path,
                completed_source=completed_path,
            )
        except (OSError, ValueError, ValidationError, sqlite3.DatabaseError) as exc:
            if template_id is not None:
                try:
                    self.db.delete_template(template_id)
                except sqlite3.DatabaseError:
                    pass
            self._show_error("No se pudo crear la plantilla desde el diligenciado", exc)
            return

        self._set_current_template(
            {
                "id": template_id,
                "nombre": name,
                "ruta_pdf": str(target.resolve()),
                "formato": "xlsx",
                "ruta_referencia": str(completed_path.resolve()),
            }
        )
        self.current_mapping_path = mapping_path
        self.current_mapping_payload = payload
        self.refresh_template_table()
        self._populate_mapping_table(self.form_templates.mapping(payload))
        self._show_current_fields()
        self.tabs.setCurrentWidget(self.mapping_tab)
        QMessageBox.information(
            self,
            "Plantilla creada",
            f"Carpeta de trabajo:\n{work_dir}\n\n"
            f"Campos detectados: {len(self.form_templates.fields(payload))}\n"
            "Revisa el mapeo y ajusta los campos que falten.",
        )

    def _suggest_template_name_from_completed(self, path: Path) -> str:
        base_name = self._clean_party_name(path.stem)
        document_type = self._detect_template_document_type(path)
        if not document_type:
            return base_name
        normalized_base = base_name.casefold()
        normalized_type = document_type.casefold()
        if normalized_base.startswith(normalized_type):
            return base_name
        return f"{document_type}-{base_name}"

    def _detect_template_document_type(self, path: Path) -> str:
        text = self._xlsx_preview_text(path)
        if not text:
            return ""
        if (
            "formato creacion y o actualizacion de terceros" in text
            or "creacion y o actualizacion de terceros" in text
            or "creacion actualizacion de terceros" in text
        ):
            return "Creación Terceros"
        if (
            "creacion proveedor" in text
            or "actualizacion proveedor" in text
            or "creacion y o actualizacion de proveedor" in text
        ):
            return "Creación Proveedor"
        if "siplat" in text or "siplaft" in text:
            return "Siplat"
        if (
            "sagrilaft" in text
            or "sarlaft" in text
            or "riesgo de lavado de activos" in text
            or "conocimiento contraparte" in text
        ):
            return "Sagrilaft"
        return ""

    def _xlsx_preview_text(self, path: Path) -> str:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            snippets: list[str] = []
            for sheet in workbook.worksheets[:3]:
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, 90),
                    values_only=True,
                ):
                    for value in row:
                        if isinstance(value, str) and value.strip():
                            snippets.append(value)
                            if len(snippets) >= 160:
                                return self._normalize_template_text(" ".join(snippets))
            return self._normalize_template_text(" ".join(snippets))
        finally:
            workbook.close()

    def _normalize_template_text(self, value: str) -> str:
        import unicodedata

        normalized = unicodedata.normalize("NFKD", value.casefold())
        normalized = "".join(
            char
            for char in normalized
            if not unicodedata.combining(char)
        )
        return re.sub(r"[^a-z0-9]+", " ", normalized)

    def _clean_party_name(self, value: str) -> str:
        cleaned = re.sub(
            r"(?i)(?:[\s_-]+)(vacio|vacío|diligenciado|plantilla|formato)$",
            "",
            value,
        )
        cleaned = re.sub(
            r"(?i)^(creacion|creación)\s*(proveedor|terceros)?[\s_-]+",
            "",
            cleaned,
        )
        cleaned = re.sub(
            r"(?i)^(sagrilaft|sarlaft|siplat|siplaft)[\s_-]+",
            "",
            cleaned,
        )
        cleaned = re.sub(r"[_-]+", " ", cleaned).strip()
        return cleaned or value.strip() or "Plantilla"

    def register_generated_template_from_ui(self) -> None:
        output_dir = Path(self.generator_output_input.text().strip())
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar plantilla generada",
            str(output_dir),
            "Excel generado (*.xlsx)",
        )
        if not file_path:
            return
        source = Path(file_path)
        name, accepted = QInputDialog.getText(
            self,
            "Nombre de plantilla",
            "Nombre para registrar:",
            text=source.stem,
        )
        if not accepted:
            return
        name = name.strip() or source.stem
        template_id: int | None = None
        version_dir = self._template_storage_dir(EXCEL_TEMPLATES_DIR, name)
        target = version_dir / source.name
        mapping_name = f"mapeo_{name}"
        try:
            payload = self.marker_templates.build_excel_payload(
                source,
                target,
                self._available_mapping_keys(),
            )
            template_id = self.db.add_template(
                TemplateRecord(
                    nombre=name,
                    ruta_pdf=str(target.resolve()),
                    descripcion="Plantilla generada desde marcadores",
                    formato="xlsx",
                    ruta_referencia=str(source.resolve()),
                )
            )
            payload["template_id"] = template_id
            payload["template_fingerprint"] = self.form_templates.document_identity.fingerprint(
                target,
                payload,
            )
            mapping_path = self.mapping_service.save_payload(
                payload,
                mapping_name,
                f"plantilla_{template_id}_{mapping_name}",
            )
            self.db.add_mapping_record(
                mapping_name,
                str(mapping_path.resolve()),
                template_id,
            )
            version = self.db.add_template_version(
                template_id,
                str(target.resolve()),
                str(source.resolve()),
                str(mapping_path.resolve()),
                str(payload.get("template_fingerprint") or ""),
            )
            payload["template_version"] = version
            mapping_path = self.mapping_service.save_payload(
                payload,
                mapping_name,
                f"plantilla_{template_id}_{mapping_name}",
            )
        except (OSError, ValueError, ValidationError, sqlite3.DatabaseError) as exc:
            if template_id is not None:
                try:
                    self.db.delete_template(template_id)
                except sqlite3.DatabaseError:
                    pass
            self._show_error("No se pudo registrar la plantilla generada", exc)
            return

        self._set_current_template(
            {
                "id": template_id,
                "nombre": name,
                "ruta_pdf": str(target.resolve()),
                "formato": "xlsx",
                "ruta_referencia": str(source.resolve()),
            }
        )
        self.current_mapping_path = mapping_path
        self.current_mapping_payload = payload
        self.refresh_template_table()
        self._populate_mapping_table(self.form_templates.mapping(payload))
        self._show_current_fields()
        self.tabs.setCurrentWidget(self.templates_tab)
        QMessageBox.information(
            self,
            "Plantilla generada registrada",
            f"Plantilla: {name}\n"
            f"Campos con marcador: {len(payload.get('cells', []))}\n\n"
            "Ya puedes diligenciarla desde Formularios o revisar el mapeo.",
        )

    def _register_template_pair(
        self,
        name: str,
        empty_source: Path,
        completed_source: Path,
    ) -> tuple[int, Path, dict[str, object], Path]:
        target_dir = {
            ".pdf": PDF_TEMPLATES_DIR,
            ".xlsx": EXCEL_TEMPLATES_DIR,
            ".docx": WORD_TEMPLATES_DIR,
        }.get(empty_source.suffix.lower())
        if target_dir is None:
            raise ValueError("Use un formulario PDF, XLSX o DOCX.")
        version_dir = self._template_storage_dir(target_dir, name)
        target = version_dir / empty_source.name
        reference = version_dir / (
            f"{completed_source.stem}_referencia{completed_source.suffix.lower()}"
        )
        template_id: int | None = None
        try:
            payload = self.form_templates.learn(
                empty_source,
                completed_source,
                self.db.get_master_data(),
            )
            version_dir.mkdir(parents=True, exist_ok=True)
            if empty_source.resolve() != target.resolve():
                shutil.copy2(empty_source, target)
            if completed_source.resolve() != reference.resolve():
                shutil.copy2(completed_source, reference)
            template_id = self.db.add_template(
                TemplateRecord(
                    nombre=name,
                    ruta_pdf=str(target.resolve()),
                    descripcion=f"Formulario {empty_source.suffix.upper()}",
                    formato=empty_source.suffix.lower().lstrip("."),
                    ruta_referencia=str(reference.resolve()),
                )
            )
            mapping_name = f"mapeo_{name}"
            payload["template_id"] = template_id
            mapping_path = self.mapping_service.save_payload(
                payload,
                mapping_name,
                f"plantilla_{template_id}_{mapping_name}",
            )
            self.db.add_mapping_record(
                mapping_name,
                str(mapping_path.resolve()),
                template_id,
            )
            version = self.db.add_template_version(
                template_id,
                str(target.resolve()),
                str(reference.resolve()),
                str(mapping_path.resolve()),
                str(payload.get("template_fingerprint") or ""),
            )
            payload["template_version"] = version
            mapping_path = self.mapping_service.save_payload(
                payload,
                mapping_name,
                f"plantilla_{template_id}_{mapping_name}",
            )
            return template_id, mapping_path, payload, target
        except Exception:
            if template_id is not None:
                self.db.delete_template(template_id)
            raise

    def _populate_template_generator_report(
        self,
        rows: list[GenerationReportRow],
    ) -> None:
        self.generator_report_table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            values = [
                row.archivo,
                row.tipo,
                row.estado,
                str(row.cantidad_reemplazos),
                row.valores_reemplazados,
                row.errores,
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.generator_report_table.setItem(row_index, col, item)

    def _default_template_generator_dictionary(self) -> Path:
        candidates = [
            MASTER_DIR / "diccionario_madecentro.json",
            MASTER_DIR / "diccionario_madecentro.xlsx",
            BASE_DIR / "diccionario_madecentro.xlsx",
            BASE_DIR / "diccionario_madecentro.json",
        ]
        for candidate in candidates:
            if candidate.is_file():
                return candidate
        return candidates[0]

    def sync_master_data_file_clicked(self) -> None:
        try:
            self._sync_master_data_file()
        except OSError as exc:
            self._show_error("No se pudo sincronizar el CSV", exc)
            return
        QMessageBox.information(
            self,
            "CSV sincronizado",
            f"datos_maestros.csv fue actualizado desde madecentro.db:\n"
            f"{MASTER_DATA_EXPORT_PATH}",
        )

    def export_master_data_xlsx(self) -> None:
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar maestros a Excel",
            str(MASTER_DATA_EXPORT_PATH.with_suffix(".xlsx")),
            "Excel (*.xlsx)",
        )
        if not file_path:
            return
        target = Path(file_path)
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        try:
            self._export_master_data_xlsx(target)
        except OSError as exc:
            self._show_error("No se pudo exportar el Excel", exc)
            return
        QMessageBox.information(
            self,
            "Excel exportado",
            f"Datos maestros exportados desde madecentro.db:\n{target}",
        )

    def choose_signature_file(self) -> None:
        SIGNATURE_DIR.mkdir(parents=True, exist_ok=True)
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Elegir firma",
            str(SIGNATURE_DIR),
            "Imagen de firma (*.png *.jpg *.jpeg *.webp)",
        )
        if not file_path:
            return
        source = Path(file_path)
        if source.suffix.lower() not in {".png", ".jpg", ".jpeg", ".webp"}:
            QMessageBox.warning(
                self,
                "Formato no soportado",
                "Seleccione una imagen PNG, JPG, JPEG o WEBP.",
            )
            return
        target = SIGNATURE_DIR / f"00_firma_activa{source.suffix.lower()}"
        try:
            for previous in SIGNATURE_DIR.glob("00_firma_activa.*"):
                if previous.resolve() != target.resolve():
                    previous.unlink(missing_ok=True)
            if source.resolve() != target.resolve():
                shutil.copy2(source, target)
        except OSError as exc:
            self._show_error("No se pudo seleccionar la firma", exc)
            return
        self._refresh_asset_status()
        QMessageBox.information(
            self,
            "Firma seleccionada",
            f"Firma activa: {target.name}",
        )

    def import_certificate_customers(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Importar clientes para certificados",
            str(CUSTOMER_CATALOG_PATH.parent),
            "Excel (*.xlsx)",
        )
        if not file_path:
            return
        source = Path(file_path)
        try:
            customers = self.customer_catalog.read_excel(source)
            CUSTOMER_CATALOG_PATH.parent.mkdir(parents=True, exist_ok=True)
            if source.resolve() != CUSTOMER_CATALOG_PATH.resolve():
                shutil.copy2(source, CUSTOMER_CATALOG_PATH)
            count = self.db.replace_certificate_customers(customers)
        except (OSError, ValueError, sqlite3.DatabaseError) as exc:
            self._show_error("No se pudo importar el catálogo de clientes", exc)
            return
        self.refresh_mapping_combos()
        self.refresh_certificate_status()
        QMessageBox.information(
            self,
            "Clientes importados",
            f"Clientes disponibles para certificados: {count}",
        )

    def _load_customer_catalog_if_available(self) -> None:
        if self.db.count_certificate_customers() or not CUSTOMER_CATALOG_PATH.is_file():
            return
        try:
            customers = self.customer_catalog.read_excel(CUSTOMER_CATALOG_PATH)
            self.db.replace_certificate_customers(customers)
        except (OSError, ValueError, sqlite3.DatabaseError):
            return

    def import_completed_form(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Actualizar maestros desde formulario diligenciado",
            "data/entrada",
            "Formularios (*.pdf *.xlsx *.docx)",
        )
        if not file_path:
            return
        source = Path(file_path)
        try:
            extracted = self.form_import.extract(source)
            resolved = self.form_import.last_resolved_payload or {}
            template_id_value = resolved.get("template_id")
            self.db.record_form_submission(
                extracted,
                source,
                int(template_id_value) if template_id_value is not None else None,
            )
            changes = self.form_import.compare(
                extracted,
                self.db.list_master_data(),
            )
        except (OSError, ValueError, RuntimeError, sqlite3.DatabaseError) as exc:
            self._show_error("No se pudo leer el formulario", exc)
            return

        if not changes:
            QMessageBox.information(
                self,
                "Sin cambios",
                "Los datos reconocidos ya coinciden con los datos maestros.",
            )
            return

        dialog = ImportPreviewDialog(changes, source.name, self)
        if dialog.exec() != QDialog.Accepted:
            return
        selected = dialog.selected_changes()
        if not selected:
            QMessageBox.information(
                self,
                "Sin seleccion",
                "No se seleccionaron datos para actualizar.",
            )
            return
        try:
            count = self.db.apply_form_import(selected, source)
            self._sync_master_data_file()
        except (OSError, sqlite3.DatabaseError) as exc:
            self._show_error("No se pudieron actualizar los datos maestros", exc)
            return
        self.refresh_all()
        QMessageBox.information(
            self,
            "Maestros actualizados",
            f"Datos aprobados: {count}\n"
            f"Base actualizada: madecentro.db\n"
            f"CSV sincronizado: {MASTER_DATA_EXPORT_PATH}",
        )

    def show_import_history(self) -> None:
        rows = self.db.list_import_history()
        dialog = QDialog(self)
        dialog.setWindowTitle("Historial de importaciones")
        dialog.resize(850, 420)
        layout = QVBoxLayout(dialog)
        table = QTableWidget(len(rows), 5)
        table.setHorizontalHeaderLabels(
            ["ID", "Fecha", "Formato", "Cambios", "Archivo"]
        )
        table.horizontalHeader().setStretchLastSection(True)
        for row_index, row in enumerate(rows):
            values = [
                row["id"],
                row["created_at"],
                row["formato"],
                row["cantidad"],
                row["archivo"],
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                table.setItem(row_index, column, item)
        layout.addWidget(table)
        close_btn = QPushButton("Cerrar")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        dialog.exec()

    def _read_master_file(self, path: Path) -> list[MasterData]:
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as file:
                reader = csv.DictReader(file)
                headers = {
                    header: header.strip().lower()
                    for header in (reader.fieldnames or [])
                    if header is not None
                }
                if {"clave", "valor"}.issubset(set(headers.values())):
                    items: list[MasterData] = []
                    for row in reader:
                        normalized = {
                            normalized_header: row.get(original_header, "")
                            for original_header, normalized_header in headers.items()
                        }
                        clave = str(normalized.get("clave") or "").strip()
                        if not clave:
                            continue
                        items.append(
                            MasterData(
                                clave=clave,
                                valor=str(normalized.get("valor") or ""),
                                categoria=str(normalized.get("categoria") or "general"),
                            )
                        )
                    return items
            with path.open("r", encoding="utf-8-sig", newline="") as file:
                rows = list(csv.reader(file))
                return [MasterData(clave=row[0], valor=row[1] if len(row) > 1 else "", categoria=row[2] if len(row) > 2 else "general") for row in rows if row]
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
            has_headers = "clave" in headers and "valor" in headers
            data_rows = rows[1:] if has_headers else rows
            items: list[MasterData] = []
            for row in data_rows:
                if has_headers:
                    values = dict(zip(headers, row))
                    clave = values.get("clave")
                    valor = values.get("valor") or ""
                    categoria = values.get("categoria") or "general"
                else:
                    clave = row[0] if len(row) > 0 else None
                    valor = row[1] if len(row) > 1 else ""
                    categoria = row[2] if len(row) > 2 else "general"
                if clave:
                    items.append(MasterData(clave=str(clave), valor=str(valor), categoria=str(categoria)))
            return items
        finally:
            workbook.close()

    def save_master_row(self) -> None:
        try:
            item = MasterData(
                clave=self.master_key_input.text(),
                valor=self.master_value_input.text().strip(),
                categoria=self.master_category_input.text() or "general",
            )
            if self.selected_master_id is None:
                self.db.upsert_master_data(item)
            else:
                self.db.update_master_data_by_id(self.selected_master_id, item)
        except (sqlite3.IntegrityError, ValidationError) as exc:
            self._show_error("No se pudo guardar el dato maestro", exc)
            return
        self.clear_master_selection()
        self.refresh_all()
        self._sync_master_data_file()

    def load_selected_master_row(self) -> None:
        row = self._selected_master_row()
        if row is None:
            return
        self.load_master_row_from_selection(row, 0)

    def load_master_row_from_selection(self, row: int, _col: int) -> None:
        id_item = self.master_table.item(row, 0)
        key_item = self.master_table.item(row, 1)
        value_item = self.master_table.item(row, 2)
        category_item = self.master_table.item(row, 3)
        if not id_item or not key_item or not value_item or not category_item:
            return
        self.selected_master_id = int(id_item.text())
        self.master_key_input.setText(key_item.text())
        self.master_value_input.setText(value_item.text())
        self.master_category_input.setText(category_item.text())

    def edit_selected_master_row(self) -> None:
        row = self._selected_master_row()
        if row is None:
            QMessageBox.warning(self, "Sin seleccion", "Seleccione una fila de datos maestros primero.")
            return
        self.load_master_row_from_selection(row, 0)

    def delete_selected_master_row(self) -> None:
        row = self._selected_master_row()
        if row is None:
            QMessageBox.warning(self, "Sin seleccion", "Seleccione una fila de datos maestros primero.")
            return
        self.load_master_row_from_selection(row, 0)
        key = self.master_key_input.text()
        usages = self._master_key_usage(key)
        usage_text = ""
        if usages:
            preview = "\n".join(f"- {name}" for name in usages[:8])
            suffix = "\n..." if len(usages) > 8 else ""
            usage_text = (
                "\n\nEsta clave se usa en estos mapeos:\n"
                f"{preview}{suffix}\n\n"
                "Si la eliminas, esos formularios pueden quedar incompletos."
            )
        response = QMessageBox.question(
            self,
            "Confirmar eliminacion",
            f"¿Eliminar la clave maestra '{key}'?{usage_text}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if response != QMessageBox.Yes or self.selected_master_id is None:
            return
        self.db.delete_master_data_by_id(self.selected_master_id)
        self.clear_master_selection()
        self.refresh_all()
        self._sync_master_data_file()

    def _master_key_usage(self, master_key: str) -> list[str]:
        if not master_key:
            return []
        templates = {
            int(template["id"]): self._template_display_name(template)
            for template in self.db.list_templates()
        }
        usages: list[str] = []
        for mapping_record in self.db.list_mappings():
            path = Path(str(mapping_record["ruta_json"]))
            if not path.is_file():
                continue
            try:
                payload = self.mapping_service.load_payload(path)
                mapping = self.form_templates.mapping(payload)
            except (OSError, ValueError):
                continue
            if master_key not in set(mapping.values()):
                continue
            template_id = mapping_record.get("plantilla_id")
            if template_id is not None and int(template_id) in templates:
                usages.append(templates[int(template_id)])
            else:
                usages.append(str(mapping_record["nombre"]))
        return sorted(set(usages))

    def clear_master_selection(self) -> None:
        self.selected_master_id = None
        self.master_key_input.clear()
        self.master_value_input.clear()
        self.master_category_input.setText("general")
        self.master_table.clearSelection()

    def _selected_master_row(self) -> int | None:
        selected_rows = self.master_table.selectionModel().selectedRows()
        if not selected_rows:
            return None
        return selected_rows[0].row()

    def delete_selected_template(self) -> None:
        row = self._selected_template_row()
        if row is None:
            QMessageBox.warning(self, "Sin seleccion", "Seleccione una plantilla primero.")
            return
        id_item = self.template_table.item(row, 0)
        name_item = self.template_table.item(row, 1)
        if not id_item or not name_item:
            return
        template_id = int(id_item.text())
        name = name_item.text()
        response = QMessageBox.question(
            self,
            "Confirmar eliminacion",
            f"¿Eliminar la plantilla '{name}' y sus mapeos asociados?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if response != QMessageBox.Yes:
            return
        try:
            self.db.delete_template(template_id)
        except sqlite3.DatabaseError as exc:
            self._show_error("No se pudo eliminar la plantilla", exc)
            return
        if self.current_template_id == template_id:
            self._clear_template_state()
        self.refresh_all()

    def _selected_template_row(self) -> int | None:
        selected_rows = self.template_table.selectionModel().selectedRows()
        if not selected_rows:
            return None
        return selected_rows[0].row()

    def load_template(self) -> None:
        empty_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar formulario VACIO",
            "data/entrada",
            "Formularios (*.pdf *.xlsx *.docx)",
        )
        if not empty_path:
            return
        empty_source = Path(empty_path)
        completed_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar el MISMO formulario DILIGENCIADO",
            str(empty_source.parent),
            f"Formulario (*{empty_source.suffix.lower()})",
        )
        if not completed_path:
            return
        completed_source = Path(completed_path)
        name = self.template_name_input.text().strip() or empty_source.stem
        template_id: int | None = None
        target_dir = {
            ".pdf": PDF_TEMPLATES_DIR,
            ".xlsx": EXCEL_TEMPLATES_DIR,
            ".docx": WORD_TEMPLATES_DIR,
        }.get(empty_source.suffix.lower())
        if target_dir is None:
            QMessageBox.warning(
                self,
                "Formato no soportado",
                "Use un formulario PDF, XLSX o DOCX.",
            )
            return
        version_dir = self._template_storage_dir(target_dir, name)
        target = version_dir / empty_source.name
        reference = version_dir / (
            f"{completed_source.stem}_referencia{completed_source.suffix.lower()}"
        )
        try:
            payload = self.form_templates.learn(
                empty_source,
                completed_source,
                self.db.get_master_data(),
            )
            version_dir.mkdir(parents=True, exist_ok=True)
            if empty_source.resolve() != target.resolve():
                shutil.copy2(empty_source, target)
            if completed_source.resolve() != reference.resolve():
                shutil.copy2(completed_source, reference)
            template_id = self.db.add_template(
                TemplateRecord(
                    nombre=name,
                    ruta_pdf=str(target.resolve()),
                    descripcion=f"Formulario {empty_source.suffix.upper()}",
                    formato=empty_source.suffix.lower().lstrip("."),
                    ruta_referencia=str(reference.resolve()),
                )
            )
            mapping_name = f"mapeo_{name}"
            payload["template_id"] = template_id
            mapping_path = self.mapping_service.save_payload(
                payload,
                mapping_name,
                f"plantilla_{template_id}_{mapping_name}",
            )
            self.db.add_mapping_record(
                mapping_name,
                str(mapping_path.resolve()),
                template_id,
            )
            version = self.db.add_template_version(
                template_id,
                str(target.resolve()),
                str(reference.resolve()),
                str(mapping_path.resolve()),
                str(payload.get("template_fingerprint") or ""),
            )
            payload["template_version"] = version
            mapping_path = self.mapping_service.save_payload(
                payload,
                mapping_name,
                f"plantilla_{template_id}_{mapping_name}",
            )
        except (OSError, ValueError, ValidationError, sqlite3.DatabaseError) as exc:
            if template_id is not None:
                try:
                    self.db.delete_template(template_id)
                except sqlite3.DatabaseError:
                    pass
            self._show_error("No se pudo cargar la plantilla", exc)
            return
        self._set_current_template(
            {
                "id": template_id,
                "ruta_pdf": str(target.resolve()),
                "formato": empty_source.suffix.lower().lstrip("."),
                "ruta_referencia": str(reference.resolve()),
            }
        )
        self.current_mapping_path = mapping_path
        self.current_mapping_payload = payload
        self._show_current_fields()
        self._populate_mapping_table(self.form_templates.mapping(payload))
        self.refresh_template_table()
        self._offer_sample_import(payload, completed_source)
        QMessageBox.information(
            self,
            "Plantilla registrada",
            f"Formato: {empty_source.suffix.upper()}\n"
            f"Campos detectados: {len(self.form_templates.fields(payload))}",
        )

    def learn_pdf_template(self) -> None:
        pdf_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar PDF AcroForm",
            "data/entrada",
            "PDF AcroForm (*.pdf)",
        )
        if not pdf_path:
            return
        source = Path(pdf_path)
        name = self.template_name_input.text().strip() or source.stem
        try:
            payload = self.form_templates.learn_pdf_acroform(
                source,
                self.db.get_master_data(),
            )
        except (OSError, ValueError, RuntimeError) as exc:
            self._show_error("No se pudo aprender la plantilla PDF", exc)
            return

        dialog = PdfLearningDialog(
            self.form_templates.mapping(payload),
            self._available_mapping_keys(),
            self,
        )
        if dialog.exec() != QDialog.Accepted:
            return
        mapping = dialog.mapping()
        self.form_templates.apply_field_mapping(payload, mapping)
        if set(mapping.values()) & CUSTOMER_KEYS:
            payload["purpose"] = "certificate"

        version_dir = self._template_storage_dir(PDF_TEMPLATES_DIR, name)
        target = version_dir / source.name
        mapping_name = f"mapeo_{name}"
        try:
            version_dir.mkdir(parents=True, exist_ok=True)
            if source.resolve() != target.resolve():
                shutil.copy2(source, target)
            template_id = self.db.add_template(
                TemplateRecord(
                    nombre=name,
                    ruta_pdf=str(target.resolve()),
                    descripcion="Plantilla PDF AcroForm aprendida",
                    formato="pdf",
                    ruta_referencia="",
                )
            )
            payload["template_id"] = template_id
            mapping_path = self.mapping_service.save_payload(
                payload,
                mapping_name,
                f"plantilla_{template_id}_{mapping_name}",
            )
            self.db.add_mapping_record(
                mapping_name,
                str(mapping_path.resolve()),
                template_id,
            )
            version = self.db.add_template_version(
                template_id,
                str(target.resolve()),
                "",
                str(mapping_path.resolve()),
                str(payload.get("template_fingerprint") or ""),
            )
            payload["template_version"] = version
            mapping_path = self.mapping_service.save_payload(
                payload,
                mapping_name,
                f"plantilla_{template_id}_{mapping_name}",
            )
        except (OSError, ValueError, ValidationError, sqlite3.DatabaseError) as exc:
            self._show_error("No se pudo guardar la plantilla PDF", exc)
            return

        self._set_current_template(
            {
                "id": template_id,
                "ruta_pdf": str(target.resolve()),
                "formato": "pdf",
                "ruta_referencia": "",
            }
        )
        self.current_mapping_path = mapping_path
        self.current_mapping_payload = payload
        self.mapping_name_input.setText(mapping_name)
        self._show_current_fields()
        self._populate_mapping_table(mapping)
        self.refresh_all()
        if payload.get("purpose") == "certificate":
            self.tabs.setCurrentWidget(self.certificate_tab)
        QMessageBox.information(
            self,
            "Plantilla aprendida",
            f"Campos AcroForm detectados: {len(mapping)}\n"
            f"Mapeo guardado en:\n{mapping_path}",
        )

    def select_template_from_table(self, row: int, _col: int) -> None:
        path_item = self.template_table.item(row, 3)
        if not path_item:
            return
        path = Path(path_item.text())
        id_item = self.template_table.item(row, 0)
        if not id_item:
            return
        format_item = self.template_table.item(row, 2)
        reference_item = self.template_table.item(row, 4)
        self._select_template_record(
            {
                "id": int(id_item.text()),
                "ruta_pdf": str(path),
                "formato": format_item.text().lower() if format_item else path.suffix,
                "ruta_referencia": reference_item.text() if reference_item else "",
            }
        )

    def select_template_from_list(self, item: QListWidgetItem) -> None:
        template_id = int(item.data(Qt.UserRole))
        template = next(
            (
                row
                for row in self._registered_template_rows()
                if int(row["id"]) == template_id
            ),
            None,
        )
        if template is None:
            QMessageBox.warning(
                self,
                "Plantilla no encontrada",
                "La plantilla seleccionada ya no está registrada.",
            )
            self.refresh_template_table()
            return
        self._select_template_record(template)

    def _select_template_record(self, template: dict[str, object]) -> None:
        self._set_current_template(template)
        try:
            self._show_current_fields()
        except (OSError, ValueError, RuntimeError) as exc:
            self._show_error("No se pudieron leer los campos", exc)

    def list_pdf_fields(self) -> None:
        if not self.current_pdf:
            QMessageBox.warning(
                self,
                "Sin plantilla",
                "Seleccione o registre una plantilla primero.",
            )
            return
        try:
            self._show_current_fields()
            fields = self.current_fields
        except (OSError, ValueError, RuntimeError) as exc:
            self._show_error("No se pudieron leer los campos", exc)
            return
        QMessageBox.information(
            self,
            "Campos detectados",
            f"Campos: {len(fields)}",
        )

    def create_empty_mapping(self) -> None:
        if not self._ensure_current_mapping_fields(
            "Seleccione una plantilla antes de crear un mapeo vacio."
        ):
            return
        self._populate_mapping_table({field: "" for field in self.current_fields})

    def suggest_mapping(self) -> None:
        if not self._ensure_current_mapping_fields(
            "Seleccione una plantilla antes de sugerir el mapeo."
        ):
            return
        existing_mapping = self._mapping_from_table()
        try:
            master_keys = list(self.db.get_master_data().keys())
            suggestions = self.ai_mapping.suggest_mapping(
                self.current_fields,
                master_keys,
                use_openai=False,
            )
            mapping = {
                field: existing_mapping.get(field) or suggestions.get(field, "")
                for field in self.current_fields
            }
            self._populate_mapping_table(mapping)
            self._save_suggested_mapping(mapping)
        except Exception as exc:
            self._show_error("No se pudo sugerir el mapeo automatico", exc)

    def complete_mapping_with_ai(self) -> None:
        if not self._ensure_current_mapping_fields(
            "Seleccione una plantilla antes de completar el mapeo con IA."
        ):
            return
        if not self.ai_mapping.api_key:
            QMessageBox.warning(
                self,
                "IA no configurada",
                "No hay OPENAI_API_KEY configurada. Use Sugerir con diccionario o configure la clave en .env.",
            )
            return
        existing_mapping = self._mapping_from_table()
        pending_fields = [
            field for field in self.current_fields if not existing_mapping.get(field)
        ]
        if not pending_fields:
            QMessageBox.information(
                self,
                "Mapeo completo",
                "No hay campos pendientes por completar.",
            )
            return
        try:
            master_keys = list(self.db.get_master_data().keys())
            suggestions = self.ai_mapping.suggest_mapping(
                pending_fields,
                master_keys,
                use_openai=True,
            )
            mapping = {
                field: existing_mapping.get(field) or suggestions.get(field, "")
                for field in self.current_fields
            }
            self._populate_mapping_table(mapping)
            self._save_suggested_mapping(mapping)
        except Exception as exc:
            QMessageBox.warning(
                self,
                "IA no respondio",
                (
                    "Se conserva el mapeo actual. La IA no respondio para los "
                    f"campos faltantes:\n{exc}"
                ),
            )

    def _ensure_current_mapping_fields(self, message: str) -> bool:
        if not self.current_pdf:
            QMessageBox.warning(self, "Sin plantilla", message)
            return False
        if not self.current_fields:
            try:
                self._show_current_fields()
            except (OSError, ValueError, RuntimeError) as exc:
                self._show_error("No se pudieron leer los campos", exc)
                return False
        if not self.current_fields:
            QMessageBox.warning(
                self,
                "Sin campos",
                "La plantilla seleccionada no tiene campos detectados para mapear.",
            )
            return False
        return True

    def _save_suggested_mapping(self, mapping: dict[str, str]) -> None:
        name = self._default_mapping_name()
        file_name = self._default_mapping_file_name(name)
        payload = self.current_mapping_payload or {
            "format": self.current_pdf.suffix.lower().lstrip(".")
            if self.current_pdf
            else "pdf",
            "mapping": mapping,
        }
        self.form_templates.apply_field_mapping(payload, mapping)
        self.current_mapping_path = self.mapping_service.save_payload(
            payload,
            name,
            file_name,
        )
        self.current_mapping_payload = payload
        self.mapping_name_input.setText(name)
        if self.current_template_id is not None:
            self.db.add_mapping_record(
                name,
                str(self.current_mapping_path.resolve()),
                self.current_template_id,
            )
        self.selected_mapping_label.setText(f"Mapeo: {self.current_mapping_path}")

    def _default_mapping_name(self) -> str:
        template_name = ""
        if self.current_template_id is not None:
            template = next(
                (
                    row
                    for row in self.db.list_templates()
                    if int(row["id"]) == self.current_template_id
                ),
                None,
            )
            if template is not None:
                template_name = self._template_display_name(template)
        if not template_name and self.current_pdf is not None:
            template_name = self._clean_template_display_name(self.current_pdf.stem)
        return f"mapeo_{template_name}" if template_name else "mapeo_madecentro"

    def _default_mapping_file_name(self, mapping_name: str) -> str:
        if self.current_template_id is None:
            return mapping_name
        return f"plantilla_{self.current_template_id}_{mapping_name}"

    def _populate_mapping_table(self, mapping: dict[str, str]) -> None:
        master_keys = sorted(
            {
                *self._available_mapping_keys(),
                *(key for key in mapping.values() if key),
            }
        )
        self.mapping_table.setRowCount(len(mapping))
        for row, (field_name, master_key) in enumerate(mapping.items()):
            field_item = QTableWidgetItem(field_name)
            field_item.setFlags(field_item.flags() & ~Qt.ItemIsEditable)
            combo = QComboBox()
            combo.setEditable(True)
            combo.addItem("")
            combo.addItems(master_keys)
            combo.setCurrentText(master_key)
            combo.currentTextChanged.connect(
                lambda value, row_index=row: self._set_mapping_value_preview(
                    row_index,
                    value,
                )
            )
            self.mapping_table.setItem(row, 0, field_item)
            self.mapping_table.setCellWidget(row, 1, combo)
            self._set_mapping_value_preview(row, master_key)

    def _set_mapping_value_preview(self, row: int, master_key: str) -> None:
        value = self._mapping_value_preview(master_key)
        item = QTableWidgetItem(value)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        item.setToolTip(value)
        self.mapping_table.setItem(row, 2, item)

    def _mapping_value_preview(self, master_key: str) -> str:
        key = master_key.strip()
        if not key:
            return ""
        if key in CUSTOMER_KEYS:
            return "(se elige cliente al generar)"
        data = self.db.get_master_data()
        data.update(self._date_generation_data())
        return str(data.get(key, ""))

    def _available_mapping_keys(self) -> list[str]:
        return sorted(
            {
                *self.db.get_master_data().keys(),
                *CUSTOMER_KEYS,
                "dia_expedicion",
                "mes_expedicion",
                "ano_expedicion",
            }
        )

    def _mapping_from_table(self) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for row in range(self.mapping_table.rowCount()):
            field_item = self.mapping_table.item(row, 0)
            combo = self.mapping_table.cellWidget(row, 1)
            if field_item and isinstance(combo, QComboBox):
                mapping[field_item.text()] = combo.currentText()
        return mapping

    def _effective_mapping(self, mapping: dict[str, str]) -> dict[str, str]:
        return {field_name: master_key for field_name, master_key in mapping.items() if master_key}

    def save_mapping(self) -> None:
        mapping = self._mapping_from_table()
        if not mapping:
            QMessageBox.warning(self, "Mapeo vacio", "Cree o sugiera un mapeo primero.")
            return
        if not self._effective_mapping(mapping):
            QMessageBox.warning(self, "Mapeo sin campos", "Seleccione al menos un campo maestro antes de guardar.")
            return
        name = self.mapping_name_input.text().strip()
        if not name or name == "mapeo_madecentro":
            name = self._default_mapping_name()
        file_name = self._default_mapping_file_name(name)
        try:
            payload = self.current_mapping_payload or {
                "format": self.current_pdf.suffix.lower().lstrip(".")
                if self.current_pdf
                else "pdf",
                "mapping": mapping,
            }
            self.form_templates.apply_field_mapping(payload, mapping)
            self.current_mapping_path = self.mapping_service.save_payload(
                payload,
                name,
                file_name,
            )
            self.current_mapping_payload = payload
            self.db.add_mapping_record(
                name,
                str(self.current_mapping_path.resolve()),
                self.current_template_id,
            )
        except (OSError, ValueError) as exc:
            self._show_error("No se pudo guardar el mapeo", exc)
            return
        self.selected_mapping_label.setText(f"Mapeo: {self.current_mapping_path}")
        if self.current_reference_path and self.current_reference_path.exists():
            self._offer_sample_import(payload, self.current_reference_path)
        QMessageBox.information(self, "Mapeo guardado", str(self.current_mapping_path))

    def fill_pdf(self) -> None:
        if not self.current_pdf:
            QMessageBox.warning(
                self,
                "Sin plantilla",
                "Seleccione una plantilla primero.",
            )
            return
        mapping = self._mapping_from_table()
        payload = self.current_mapping_payload
        if (
            (not mapping or not self._effective_mapping(mapping))
            and self.current_mapping_path
        ):
            try:
                payload = self.mapping_service.load_payload(
                    self.current_mapping_path
                )
                mapping = self.form_templates.mapping(payload)
            except (OSError, ValueError) as exc:
                self._show_error("No se pudo cargar el mapeo", exc)
                return
        mapping = self._complete_dynamic_mapping(mapping)
        effective_mapping = self._effective_mapping(mapping)
        if not effective_mapping:
            QMessageBox.warning(self, "Sin mapeo", "Cree o cargue un mapeo primero.")
            return
        if payload is None:
            payload = {
                "format": self.current_pdf.suffix.lower().lstrip("."),
                "mapping": mapping,
            }
        mapping = self._complete_dynamic_mapping(mapping)
        self.form_templates.apply_field_mapping(payload, mapping)
        output_path = self._next_output_path(
            self._template_output_name(self.current_pdf),
            self.current_pdf.suffix.lower(),
        )
        try:
            master_data = self._generation_data(mapping)
            if master_data is None:
                return
            result = self.form_templates.fill(
                self.current_pdf,
                output_path,
                payload,
                master_data,
            )
        except (OSError, ValueError, RuntimeError) as exc:
            self._show_error("No se pudo diligenciar el formulario", exc)
            return
        QMessageBox.information(self, "Formulario diligenciado", str(result))

    def _complete_dynamic_mapping(
        self,
        mapping: dict[str, str],
    ) -> dict[str, str]:
        completed = dict(mapping)
        dynamic_keys = {
            *CUSTOMER_KEYS,
            "dia_expedicion",
            "mes_expedicion",
            "ano_expedicion",
        }
        for field_name in completed:
            if not completed[field_name] and field_name in dynamic_keys:
                completed[field_name] = field_name
        return completed

    def _generation_data(
        self,
        mapping: dict[str, str],
    ) -> dict[str, str] | None:
        data = self.db.get_master_data()
        required_keys = {key for key in mapping.values() if key}
        if required_keys & CUSTOMER_KEYS:
            customer = self._select_certificate_customer()
            if customer is None:
                return None
            year = str(customer["anio_vinculacion"])
            data.update(
                {
                    "cliente_razon_social": str(customer["razon_social"]),
                    "cliente_nit": str(customer["nit"]),
                    "cliente_anio_vinculacion": year,
                    "cliente_ano_vinculacion": year,
                }
            )
        data.update(self._date_generation_data())
        return data

    def _select_certificate_customer(self) -> dict[str, object] | None:
        if self.db.count_certificate_customers() == 0:
            QMessageBox.warning(
                self,
                "Sin clientes",
                "Importe primero el Excel de clientes certificados desde "
                "Datos maestros.",
            )
            return None
        query, accepted = QInputDialog.getText(
            self,
            "Buscar cliente",
            "Razón social o NIT:",
        )
        if not accepted or not query.strip():
            return None
        matches = self.db.search_certificate_customers(query)
        if not matches:
            QMessageBox.warning(
                self,
                "Cliente no encontrado",
                "No existe un cliente que coincida con la búsqueda.",
            )
            return None
        if len(matches) == 1:
            return matches[0]
        labels = [
            f"{item['razon_social']} | {item['nit']}"
            for item in matches
        ]
        selected, accepted = QInputDialog.getItem(
            self,
            "Seleccionar cliente",
            "Coincidencias:",
            labels,
            0,
            False,
        )
        if not accepted:
            return None
        return matches[labels.index(selected)]

    def create_certificate_customer(self) -> None:
        dialog = QDialog(self)
        dialog.setWindowTitle("Nuevo cliente certificado")
        dialog.resize(520, 220)
        layout = QVBoxLayout(dialog)
        form = QGridLayout()
        form.setHorizontalSpacing(12)
        form.setVerticalSpacing(8)

        reason_input = QLineEdit()
        nit_input = QLineEdit()
        year_input = QLineEdit()
        reason_input.setPlaceholderText("Razón social")
        nit_input.setPlaceholderText("Ej. 900.456.789-1")
        year_input.setPlaceholderText("Ej. 2010")

        form.addWidget(QLabel("Razón social"), 0, 0)
        form.addWidget(reason_input, 0, 1)
        form.addWidget(QLabel("NIT"), 1, 0)
        form.addWidget(nit_input, 1, 1)
        form.addWidget(QLabel("Año de vinculación"), 2, 0)
        form.addWidget(year_input, 2, 1)
        form.setColumnStretch(1, 1)
        layout.addLayout(form)

        actions = QHBoxLayout()
        cancel_btn = QPushButton("Cancelar")
        save_btn = QPushButton("Guardar cliente")
        save_btn.setProperty("role", "primary")
        actions.addStretch()
        actions.addWidget(cancel_btn)
        actions.addWidget(save_btn)
        layout.addLayout(actions)

        saved_customer: dict[str, object] | None = None
        updated_existing_excel = False

        def save_customer() -> None:
            nonlocal saved_customer, updated_existing_excel
            reason = reason_input.text().strip()
            nit = nit_input.text().strip()
            year = year_input.text().strip()
            normalized_nit = self.customer_catalog.normalize_nit(nit)
            if not reason or not nit or not year:
                QMessageBox.warning(
                    self,
                    "Datos incompletos",
                    "Ingrese razón social, NIT y año de vinculación.",
                )
                return
            if not normalized_nit:
                QMessageBox.warning(
                    self,
                    "NIT inválido",
                    "Ingrese un NIT con al menos un número.",
                )
                return
            if not year.isdigit() or len(year) != 4:
                QMessageBox.warning(
                    self,
                    "Año inválido",
                    "Ingrese el año de vinculación con cuatro dígitos.",
                )
                return

            customer = {
                "cliente_razon_social": reason,
                "cliente_nit": nit,
                "cliente_anio_vinculacion": year,
                "cliente_ano_vinculacion": year,
                "nit_normalizado": normalized_nit,
            }
            try:
                updated_existing_excel = self.customer_catalog.upsert_excel(
                    CUSTOMER_CATALOG_PATH,
                    customer,
                )
                saved_customer = self.db.upsert_certificate_customer(customer)
            except (OSError, ValueError, sqlite3.DatabaseError) as exc:
                self._show_error("No se pudo guardar el cliente certificado", exc)
                return
            dialog.accept()

        cancel_btn.clicked.connect(dialog.reject)
        save_btn.clicked.connect(save_customer)
        if dialog.exec() != QDialog.Accepted or saved_customer is None:
            return

        self.refresh_certificate_status()
        self.certificate_search_input.setText(str(saved_customer["nit"]))
        self.search_certificate_customers()
        self._select_certificate_customer_by_id(int(saved_customer["id"]))
        action = "actualizado" if updated_existing_excel else "creado"
        QMessageBox.information(
            self,
            "Cliente guardado",
            f"Cliente certificado {action}: {saved_customer['razon_social']}",
        )

    def search_certificate_customers(self) -> None:
        query = self.certificate_search_input.text().strip()
        if not query:
            QMessageBox.warning(
                self,
                "Búsqueda vacía",
                "Escriba una razón social o un NIT.",
            )
            return
        matches = self.db.search_certificate_customers(query, limit=100)
        self.certificate_results_table.setRowCount(len(matches))
        for row_index, customer in enumerate(matches):
            values = [
                customer["id"],
                customer["razon_social"],
                customer["nit"],
                customer["anio_vinculacion"],
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.certificate_results_table.setItem(
                    row_index,
                    column,
                    item,
                )
        self.selected_certificate_customer = None
        self._set_certificate_selection_label(None)
        if not matches:
            QMessageBox.information(
                self,
                "Sin resultados",
                "No se encontraron clientes con ese nombre o NIT.",
            )

    def select_certificate_customer(self, row: int, _column: int) -> None:
        id_item = self.certificate_results_table.item(row, 0)
        name_item = self.certificate_results_table.item(row, 1)
        nit_item = self.certificate_results_table.item(row, 2)
        year_item = self.certificate_results_table.item(row, 3)
        if not all((id_item, name_item, nit_item, year_item)):
            return
        self.selected_certificate_customer = {
            "id": int(id_item.text()),
            "razon_social": name_item.text(),
            "nit": nit_item.text(),
            "anio_vinculacion": year_item.text(),
        }
        self._set_certificate_selection_label(
            self.selected_certificate_customer
        )

    def _select_certificate_customer_by_id(self, customer_id: int) -> None:
        for row in range(self.certificate_results_table.rowCount()):
            id_item = self.certificate_results_table.item(row, 0)
            if id_item and int(id_item.text()) == customer_id:
                self.certificate_results_table.selectRow(row)
                self.select_certificate_customer(row, 0)
                return

    def _set_certificate_selection_label(
        self,
        customer: dict[str, object] | None,
    ) -> None:
        if customer:
            self.selected_certificate_label.setText(
                f"Cliente: {customer['razon_social']} | "
                f"NIT {customer['nit']} | "
                f"Vinculado desde {customer['anio_vinculacion']}"
            )
            object_name = "statusReady"
        else:
            self.selected_certificate_label.setText(
                "Cliente: sin seleccionar"
            )
            object_name = "statusMissing"
        self.selected_certificate_label.setObjectName(object_name)
        self.selected_certificate_label.style().unpolish(
            self.selected_certificate_label
        )
        self.selected_certificate_label.style().polish(
            self.selected_certificate_label
        )

    def create_certificate(self) -> None:
        customer = self.selected_certificate_customer
        if customer is None:
            QMessageBox.warning(
                self,
                "Sin cliente",
                "Busque y seleccione un cliente antes de crear el certificado.",
            )
            return
        context = self._certificate_template_context()
        if context is None:
            QMessageBox.warning(
                self,
                "Sin plantilla",
                "Aprenda primero una plantilla PDF que utilice claves cliente_*.",
            )
            return
        template, payload, _mapping_path = context
        mapping = self._complete_dynamic_mapping(
            self.form_templates.mapping(payload)
        )
        self.form_templates.apply_field_mapping(payload, mapping)
        master_data = self._certificate_generation_data(customer)
        source = Path(str(template["ruta_pdf"]))
        output_path = self._next_output_path(
            f"Certificado_{self._safe_output_name(str(customer['razon_social']))}",
            source.suffix.lower(),
        )
        try:
            result = self.form_templates.fill(
                source,
                output_path,
                payload,
                master_data,
            )
        except (OSError, ValueError, RuntimeError) as exc:
            self._show_error("No se pudo crear el certificado", exc)
            return
        QMessageBox.information(
            self,
            "Certificado creado",
            str(result),
        )

    def _certificate_generation_data(
        self,
        customer: dict[str, object],
    ) -> dict[str, str]:
        year = str(customer["anio_vinculacion"])
        data = self._date_generation_data()
        data.update(
            {
                "cliente_razon_social": str(customer["razon_social"]),
                "cliente_nit": str(customer["nit"]),
                "cliente_anio_vinculacion": year,
                "cliente_ano_vinculacion": year,
            }
        )
        return data

    def _date_generation_data(self) -> dict[str, str]:
        now = datetime.now()
        months = (
            "",
            "enero",
            "febrero",
            "marzo",
            "abril",
            "mayo",
            "junio",
            "julio",
            "agosto",
            "septiembre",
            "octubre",
            "noviembre",
            "diciembre",
        )
        return {
            "dia_expedicion": str(now.day),
            "mes_expedicion": months[now.month],
            "ano_expedicion": str(now.year),
        }

    def _template_output_name(self, path: Path) -> str:
        stem = path.stem
        stem = re.sub(
            r"(?i)(?:[\s_-]+)(vacio|vacío|plantilla|formato)$",
            "",
            stem,
        )
        return self._safe_output_name(stem)

    def _safe_output_name(self, value: str) -> str:
        cleaned = re.sub(r"[^\w]+", "_", value, flags=re.UNICODE)
        cleaned = re.sub(r"_+", "_", cleaned).strip("_")
        return cleaned or "Formulario"

    def _next_output_path(self, base_name: str, suffix: str) -> Path:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        candidate = OUTPUT_DIR / f"{base_name}{suffix}"
        index = 2
        while candidate.exists():
            candidate = OUTPUT_DIR / f"{base_name}-{index}{suffix}"
            index += 1
        return candidate

    def _certificate_template_ids(self) -> set[int]:
        ids: set[int] = set()
        for template in self.db.list_templates():
            template_id = int(template["id"])
            for mapping in self.db.list_mappings(template_id):
                path = Path(str(mapping["ruta_json"]))
                if not path.is_file():
                    continue
                try:
                    payload = self.mapping_service.load_payload(path)
                except (OSError, ValueError):
                    continue
                mapped_keys = set(
                    self.form_templates.mapping(payload).values()
                )
                if (
                    payload.get("purpose") == "certificate"
                    or bool(mapped_keys & CUSTOMER_KEYS)
                ):
                    ids.add(template_id)
                    break
        return ids

    def _certificate_template_context(
        self,
    ) -> tuple[dict[str, object], dict[str, object], Path] | None:
        for template in self.db.list_templates():
            template_id = int(template["id"])
            for mapping in self.db.list_mappings(template_id):
                path = Path(str(mapping["ruta_json"]))
                if not path.is_file():
                    continue
                try:
                    payload = self.mapping_service.load_payload(path)
                except (OSError, ValueError):
                    continue
                mapped_keys = set(
                    self.form_templates.mapping(payload).values()
                )
                if (
                    payload.get("purpose") == "certificate"
                    or bool(mapped_keys & CUSTOMER_KEYS)
                ):
                    return template, payload, path
        return None

    def load_mapping(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Cargar mapeo",
            str(self.current_mapping_path.parent if self.current_mapping_path else Path("plantillas/mapeos")),
            "Mapeos JSON (*.json)",
        )
        if not file_path:
            return
        try:
            payload = self.mapping_service.load_payload(file_path)
            mapping = self.form_templates.mapping(payload)
        except (OSError, ValueError) as exc:
            self._show_error("No se pudo cargar el mapeo", exc)
            return
        if self.current_pdf:
            if not self.current_fields:
                try:
                    self._show_current_fields()
                except (OSError, ValueError, RuntimeError):
                    self.current_fields = []
            unknown_fields = sorted(set(mapping) - set(self.current_fields))
            if self.current_fields and unknown_fields:
                preview = ", ".join(unknown_fields[:5])
                suffix = "..." if len(unknown_fields) > 5 else ""
                response = QMessageBox.question(
                    self,
                    "Mapeo posiblemente distinto",
                    (
                        "Este mapeo contiene campos que no aparecen en la plantilla "
                        f"seleccionada:\n{preview}{suffix}\n\n"
                        "¿Deseas cargarlo de todos modos?"
                    ),
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No,
                )
                if response != QMessageBox.Yes:
                    return
        self.current_mapping_path = Path(file_path)
        self.current_mapping_payload = payload
        self._populate_mapping_table(mapping)
        self.selected_mapping_label.setText(f"Mapeo: {self.current_mapping_path}")

    def _set_current_template(self, template: dict[str, object]) -> None:
        path = Path(str(template["ruta_pdf"]))
        if path.suffix.lower() not in {".pdf", ".xlsx", ".docx"} or not path.exists():
            self._clear_template_state()
            return
        self.current_template_id = int(template["id"])
        self.current_pdf = path
        reference = str(template.get("ruta_referencia") or "")
        self.current_reference_path = Path(reference) if reference else None
        self.current_fields = []
        self.current_mapping_path = None
        self.current_mapping_payload = None
        self.field_table.setRowCount(0)
        self.mapping_table.setRowCount(0)
        display_name = self._template_display_name(template)
        self.selected_pdf_label.setText(
            f"Plantilla {path.suffix.upper()}: {display_name}"
        )
        self._set_mapping_actions_enabled(True)
        self.refresh_selected_mapping()
        if self.current_mapping_payload:
            mapping = self.form_templates.mapping(self.current_mapping_payload)
            self._populate_mapping_table(mapping)
        self._sync_template_selection()

    def _show_current_fields(self) -> None:
        if not self.current_pdf:
            return
        payload = self.current_mapping_payload
        if payload is None and self.current_mapping_path:
            payload = self.mapping_service.load_payload(self.current_mapping_path)
            self.current_mapping_payload = payload

        rows: list[list[str]] = []
        if payload is not None:
            fields = self.form_templates.fields(payload)
            self.current_fields = [
                str(field.get("field_id") or "")
                for field in fields
            ]
            for field in fields:
                location = str(
                    field.get("location")
                    or (
                        f"{field.get('sheet')}!{field.get('cell')}"
                        if field.get("sheet") and field.get("cell")
                        else ""
                    )
                )
                rows.append(
                    [
                        str(field.get("field_id") or ""),
                        str(
                            field.get("kind")
                            or field.get("value_type")
                            or self.current_pdf.suffix.lower().lstrip(".")
                        ),
                        location,
                        str(field.get("sample_value") or ""),
                    ]
                )
        elif self.current_pdf.suffix.lower() == ".pdf":
            pdf_fields = self.pdf_fields.list_fields(self.current_pdf)
            self.current_fields = [field.field_name for field in pdf_fields]
            rows = [
                [
                    field.field_name,
                    field.field_type,
                    f"Pagina {field.page}",
                    field.value or "",
                ]
                for field in pdf_fields
            ]
        else:
            raise ValueError("La plantilla no tiene un mapeo asociado.")

        self.field_table.setRowCount(len(rows))
        for row_index, values in enumerate(rows):
            for column, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.field_table.setItem(row_index, column, item)

    def _offer_sample_import(
        self,
        payload: dict[str, object],
        source: Path,
    ) -> None:
        extracted = self.form_templates.sample_values(payload)
        if not extracted:
            return
        changes = self.form_import.compare(
            extracted,
            self.db.list_master_data(),
        )
        if not changes:
            return
        dialog = ImportPreviewDialog(changes, source.name, self)
        dialog.setWindowTitle("Importar datos del formulario de referencia")
        if dialog.exec() != QDialog.Accepted:
            return
        selected = dialog.selected_changes()
        if selected:
            count = self.db.apply_form_import(selected, source)
            self._sync_master_data_file()
            self.refresh_master_table()
            self.refresh_mapping_combos()
            QMessageBox.information(
                self,
                "Maestros actualizados",
                f"Datos aprobados desde la referencia: {count}\n"
                f"CSV sincronizado: {MASTER_DATA_EXPORT_PATH}",
            )

    def _sync_master_data_file(self) -> None:
        rows = self.db.list_master_data()
        MASTER_DATA_EXPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
        with MASTER_DATA_EXPORT_PATH.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=["clave", "valor", "categoria"])
            writer.writeheader()
            for row in rows:
                writer.writerow(
                    {
                        "clave": row["clave"],
                        "valor": row["valor"],
                        "categoria": row["categoria"],
                    }
                )

    def _export_master_data_xlsx(self, path: Path) -> None:
        rows = self.db.list_master_data()
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "datos_maestros"
        sheet.append(["clave", "valor", "categoria"])
        for row in rows:
            sheet.append([row["clave"], row["valor"], row["categoria"]])
        workbook.save(path)
        workbook.close()

    def _template_storage_dir(self, base_dir: Path, template_name: str) -> Path:
        folder_name = re.sub(r'[<>:"/\\|?*]+', "-", template_name).strip(" .")
        folder_name = re.sub(r"\s+", " ", folder_name)
        if not folder_name:
            folder_name = "Plantilla"
        candidate = base_dir / folder_name
        index = 2
        while candidate.exists():
            candidate = base_dir / f"{folder_name}-v{index}"
            index += 1
        return candidate

    def _clear_template_state(self) -> None:
        self.current_template_id = None
        self.current_pdf = None
        self.current_reference_path = None
        self.current_fields = []
        self.current_mapping_path = None
        self.current_mapping_payload = None
        self.selected_pdf_label.setText("Plantilla: sin seleccionar")
        self.selected_mapping_label.setText("Mapeo: sin guardar")
        self._set_mapping_actions_enabled(False)

    def _set_mapping_actions_enabled(self, enabled: bool) -> None:
        for name in (
            "create_mapping_btn",
            "suggest_mapping_btn",
            "complete_ai_mapping_btn",
        ):
            button = getattr(self, name, None)
            if button is not None:
                button.setEnabled(enabled)

    def _show_error(self, title: str, exc: Exception) -> None:
        QMessageBox.critical(self, title, str(exc))

    def _validate_mapping_for_current_pdf(self, mapping: dict[str, str]) -> None:
        if not self.current_pdf:
            raise ValueError("Seleccione una plantilla antes de cargar el mapeo.")
        valid_fields = {field.field_name for field in self.pdf_fields.list_fields(self.current_pdf)}
        unknown_fields = sorted(set(mapping) - valid_fields)
        if unknown_fields:
            preview = ", ".join(unknown_fields[:5])
            suffix = "..." if len(unknown_fields) > 5 else ""
            raise ValueError(
                f"El mapeo contiene campos que no existen en la plantilla: {preview}{suffix}"
            )

    def open_output_folder(self) -> None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        self._open_directory(OUTPUT_DIR)

    def _open_directory(self, directory: Path) -> None:
        directory.mkdir(parents=True, exist_ok=True)
        if sys.platform.startswith("win"):
            os.startfile(directory)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(directory)])
        else:
            subprocess.Popen(["xdg-open", str(directory)])

