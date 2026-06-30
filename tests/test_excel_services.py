import tempfile
import unittest
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from app.services.excel_fill_service import ExcelFillService
from app.services.excel_signature_service import ExcelSignatureService
from app.services.excel_template_service import ExcelTemplateService


def create_workbook(path: Path, value: object = None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FORMULARIO"
    sheet["A1"] = "NIT"
    sheet["B1"] = value
    workbook.save(path)
    workbook.close()


class ExcelServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        root = Path(self.temp_dir.name)
        self.empty_path = root / "empty.xlsx"
        self.completed_path = root / "completed.xlsx"
        self.output_path = root / "output.xlsx"
        create_workbook(self.empty_path)
        create_workbook(self.completed_path, "900123")

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_learns_and_fills_changed_cells(self) -> None:
        mapping = ExcelTemplateService().compare_workbooks(
            self.empty_path,
            self.completed_path,
        )
        self.assertEqual(len(mapping["cells"]), 1)
        self.assertEqual(mapping["cells"][0]["cell"], "B1")
        self.assertEqual(mapping["cells"][0]["sample_value"], "900123")

        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {},
            use_sample_values=True,
        )
        workbook = load_workbook(self.output_path, data_only=False)
        try:
            self.assertEqual(workbook["FORMULARIO"]["B1"].value, "900123")
        finally:
            workbook.close()

    def test_fills_from_master_data(self) -> None:
        mapping = {
            "cells": [
                {
                    "sheet": "FORMULARIO",
                    "cell": "B1",
                    "value_type": "string",
                    "master_key": "nit",
                }
            ],
            "controls": [],
        }
        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {"nit": "811.028.650-1"},
        )
        workbook = load_workbook(self.output_path, data_only=False)
        try:
            self.assertEqual(
                workbook["FORMULARIO"]["B1"].value,
                "811.028.650-1",
            )
        finally:
            workbook.close()

    def test_splits_tax_id_number_and_check_digit(self) -> None:
        mapping = {
            "cells": [
                {
                    "sheet": "FORMULARIO",
                    "cell": "B1",
                    "value_type": "string",
                    "master_key": "nit",
                    "value_transform": "tax_id_number",
                },
                {
                    "sheet": "FORMULARIO",
                    "cell": "C1",
                    "value_type": "string",
                    "master_key": "nit",
                    "value_transform": "tax_id_check_digit",
                },
            ],
            "controls": [],
        }

        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {"nit": "811.028.650-1"},
        )

        workbook = load_workbook(self.output_path, data_only=False)
        try:
            sheet = workbook["FORMULARIO"]
            self.assertEqual(sheet["B1"].value, "811.028.650")
            self.assertEqual(sheet["C1"].value, "1")
        finally:
            workbook.close()

    def test_preserves_numeric_type_for_string_master_data(self) -> None:
        mapping = {
            "cells": [
                {
                    "sheet": "FORMULARIO",
                    "cell": "B1",
                    "value_type": "number",
                    "master_key": "dv",
                }
            ],
            "controls": [],
        }

        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {"dv": "1"},
        )

        workbook = load_workbook(self.output_path, data_only=False)
        try:
            self.assertEqual(workbook["FORMULARIO"]["B1"].value, 1)
        finally:
            workbook.close()

    def test_uses_reference_checkbox_state_without_master_key(self) -> None:
        service = ExcelFillService()

        self.assertTrue(
            service._control_value(
                {"master_key": "", "sample_checked": True},
                {},
                False,
            )
        )
        self.assertFalse(
            service._control_value(
                {"master_key": "", "sample_checked": False},
                {},
                False,
            )
        )

    def test_fills_current_generation_date(self) -> None:
        mapping = {
            "cells": [
                {
                    "field_id": "FORMULARIO!B1",
                    "sheet": "FORMULARIO",
                    "cell": "B1",
                    "value_type": "number",
                    "master_key": "",
                    "auto_value": "current_day",
                },
                {
                    "field_id": "FORMULARIO!C1",
                    "sheet": "FORMULARIO",
                    "cell": "C1",
                    "value_type": "string",
                    "master_key": "",
                    "auto_value": "current_month_name",
                },
                {
                    "field_id": "FORMULARIO!D1",
                    "sheet": "FORMULARIO",
                    "cell": "D1",
                    "value_type": "number",
                    "master_key": "",
                    "auto_value": "current_year",
                },
            ],
            "controls": [],
        }

        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {},
        )

        now = datetime.now()
        workbook = load_workbook(self.output_path, data_only=False)
        try:
            sheet = workbook["FORMULARIO"]
            self.assertEqual(sheet["B1"].value, now.day)
            self.assertEqual(sheet["D1"].value, now.year)
            self.assertTrue(sheet["C1"].value)
        finally:
            workbook.close()

    def test_preserves_reference_selection_mark_without_master_key(self) -> None:
        mapping = {
            "cells": [
                {
                    "field_id": "FORMULARIO!B1",
                    "sheet": "FORMULARIO",
                    "cell": "B1",
                    "value_type": "string",
                    "empty_value": None,
                    "sample_value": "X",
                    "master_key": "",
                }
            ],
            "controls": [],
        }

        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {},
        )

        workbook = load_workbook(self.output_path, data_only=False)
        try:
            self.assertEqual(workbook["FORMULARIO"]["B1"].value, "X")
        finally:
            workbook.close()

    def test_uses_sample_value_when_master_key_is_empty(self) -> None:
        mapping = {
            "cells": [
                {
                    "field_id": "FORMULARIO!B1",
                    "sheet": "FORMULARIO",
                    "cell": "B1",
                    "value_type": "string",
                    "sample_value": "Referencia",
                    "master_key": "clave_sin_valor",
                }
            ],
            "controls": [],
        }

        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {},
        )

        workbook = load_workbook(self.output_path, data_only=False)
        try:
            self.assertEqual(workbook["FORMULARIO"]["B1"].value, "Referencia")
        finally:
            workbook.close()

    def test_fills_current_date_in_single_cell(self) -> None:
        mapping = {
            "cells": [
                {
                    "field_id": "FORMULARIO!B1",
                    "sheet": "FORMULARIO",
                    "cell": "B1",
                    "value_type": "string",
                    "master_key": "",
                    "auto_value": "current_date",
                }
            ],
            "controls": [],
        }

        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {},
        )

        now = datetime.now()
        workbook = load_workbook(self.output_path, data_only=False)
        try:
            self.assertEqual(
                workbook["FORMULARIO"]["B1"].value,
                now.strftime("%d/%m/%Y"),
            )
        finally:
            workbook.close()

    def test_value_format_can_include_current_date(self) -> None:
        mapping = {
            "cells": [
                {
                    "field_id": "FORMULARIO!B1",
                    "sheet": "FORMULARIO",
                    "cell": "B1",
                    "value_type": "string",
                    "master_key": "representante_legal",
                    "value_format": "NOMBRE: {value} FECHA: {current_date}",
                }
            ],
            "controls": [],
        }

        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {"representante_legal": "ADRIAN"},
        )

        now = datetime.now()
        workbook = load_workbook(self.output_path, data_only=False)
        try:
            self.assertEqual(
                workbook["FORMULARIO"]["B1"].value,
                f"NOMBRE: ADRIAN FECHA: {now.strftime('%d/%m/%Y')}",
            )
        finally:
            workbook.close()

    def test_embeds_transparent_signature_without_replacing_images(self) -> None:
        logo_path = Path(self.temp_dir.name) / "logo.png"
        signature_path = Path(self.temp_dir.name) / "signature.png"
        Image.new("RGB", (20, 20), "orange").save(logo_path)
        signature = Image.new("RGB", (100, 40), "white")
        for x in range(10, 90):
            signature.putpixel((x, 20), (0, 0, 0))
        signature.save(signature_path)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "FORMULARIO"
        logo_buffer = BytesIO(logo_path.read_bytes())
        sheet.add_image(ExcelImage(logo_buffer), "B2")
        workbook.save(self.output_path)
        workbook.close()
        logo_buffer.close()

        ExcelSignatureService().apply_signature(
            self.output_path,
            signature_path,
        )

        with ZipFile(self.output_path) as archive:
            drawing = archive.read("xl/drawings/drawing1.xml")
            relationships = archive.read(
                "xl/drawings/_rels/drawing1.xml.rels"
            )
            embedded = archive.read("xl/media/madecentro_signature.png")
        self.assertIn(b"Firma Madecentro", drawing)
        self.assertIn(b"madecentro_signature.png", relationships)
        with Image.open(BytesIO(embedded)) as image:
            self.assertEqual(image.mode, "RGBA")
            self.assertEqual(image.getpixel((0, 0))[3], 255)

    def test_signature_layout_preserves_aspect_ratio_inside_range(self) -> None:
        signature_path = Path(self.temp_dir.name) / "signature-wide.png"
        Image.new("RGB", (400, 100), "black").save(signature_path)
        service = ExcelSignatureService()
        prepared = service._prepare_signature(signature_path)

        layout = service._resolve_layout(
            self.empty_path,
            prepared,
            {
                "sheet": "FORMULARIO",
                "cell": "A2",
                "end_cell": "D4",
                "padding_pixels": 3,
            },
        )

        self.assertAlmostEqual(
            layout["width_pixels"] / layout["height_pixels"],
            4.0,
            delta=0.15,
        )

    def test_signature_layout_can_align_left_inside_range(self) -> None:
        signature_path = Path(self.temp_dir.name) / "signature-wide.png"
        Image.new("RGB", (400, 100), "black").save(signature_path)
        service = ExcelSignatureService()
        prepared = service._prepare_signature(signature_path)

        layout = service._resolve_layout(
            self.empty_path,
            prepared,
            {
                "sheet": "FORMULARIO",
                "cell": "A2",
                "end_cell": "D4",
                "padding_pixels": 3,
                "horizontal_align": "left",
                "horizontal_offset_pixels": 10,
            },
        )

        self.assertEqual(layout["cell"], "A2")
        self.assertEqual(layout["col_offset_emu"], 13 * 9525)

    def test_signature_service_detects_firma_label_without_config(self) -> None:
        signature_path = Path(self.temp_dir.name) / "signature.png"
        Image.new("RGB", (120, 40), "black").save(signature_path)
        workbook_path = Path(self.temp_dir.name) / "firma_label.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "FORMULARIO"
        sheet.merge_cells("D20:H20")
        sheet["D20"] = "Firma representante legal"
        workbook.save(workbook_path)
        workbook.close()

        ExcelSignatureService().apply_signature(workbook_path, signature_path)

        with ZipFile(workbook_path) as archive:
            names = archive.namelist()
            self.assertTrue(any(name.startswith("xl/media/") for name in names))
            drawing_names = [
                name
                for name in names
                if name.startswith("xl/drawings/drawing")
                and name.endswith(".xml")
            ]
            drawing = b"".join(archive.read(name) for name in drawing_names)
        self.assertIn(b"<col>4</col>", drawing)
        self.assertIn(b"<row>16</row>", drawing)

    def test_signature_service_uses_large_merged_firma_block(self) -> None:
        signature_path = Path(self.temp_dir.name) / "signature.png"
        Image.new("RGB", (120, 40), "black").save(signature_path)
        workbook_path = Path(self.temp_dir.name) / "firma_merged.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "UNILAQUILERES"
        sheet.merge_cells("A103:K117")
        sheet["A103"] = (
            "Espacio para diligenciar el cliente\n\n\n\n\n\n\n\n"
            "Firma                                      Huella Digital\n"
            "Nombre:\nCedula:\nFecha:"
        )
        workbook.save(workbook_path)
        workbook.close()

        ExcelSignatureService().apply_signature(workbook_path, signature_path)

        with ZipFile(workbook_path) as archive:
            drawing = archive.read("xl/drawings/drawing1.xml")
        self.assertIn(b"<col>0</col>", drawing)
        self.assertIn(b"<row>105</row>", drawing)

    def test_signature_service_uses_merged_area_above_firma_label(self) -> None:
        signature_path = Path(self.temp_dir.name) / "signature.png"
        Image.new("RGB", (120, 40), "black").save(signature_path)
        workbook_path = Path(self.temp_dir.name) / "firma_above_label.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Conocimiento contraparte "
        sheet.merge_cells("C84:M84")
        sheet["C84"] = "IX. FIRMA Y HUELLA"
        sheet.merge_cells("D85:J89")
        sheet["D85"] = "DECLARO QUE LA INFORMACION ES EXACTA Y FIRMO"
        sheet.merge_cells("L85:M90")
        sheet["L85"] = "Huella Indice Derecho"
        sheet.merge_cells("D90:J90")
        sheet["D90"] = "Firma Cliente / Representante Legal"
        workbook.save(workbook_path)
        workbook.close()

        service = ExcelSignatureService()
        detected = service._detect_signature_area(workbook_path)

        self.assertEqual(detected["cell"], "D85")
        self.assertEqual(detected["end_cell"], "J89")

        ExcelSignatureService().apply_signature(workbook_path, signature_path)

        with ZipFile(workbook_path) as archive:
            drawing = archive.read("xl/drawings/drawing1.xml")
        self.assertIn(b"<col>5</col>", drawing)
        self.assertIn(b"<row>84</row>", drawing)

    def test_injects_drawing_text_marks(self) -> None:
        workbook_path = Path(self.temp_dir.name) / "drawing.xlsx"
        drawing_path = "xl/drawings/drawing1.xml"
        drawing = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor>
    <xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
    <xdr:to><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
    <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="1" name="Existing"/></xdr:nvSpPr></xdr:sp>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
</xdr:wsDr>"""
        mark = """<xdr:twoCellAnchor xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:from><xdr:col>2</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>2</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>3</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
  <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="1" name="X"/></xdr:nvSpPr><xdr:txBody><a:p><a:r><a:t>X</a:t></a:r></a:p></xdr:txBody></xdr:sp>
  <xdr:clientData/>
</xdr:twoCellAnchor>"""
        with ZipFile(workbook_path, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr(drawing_path, drawing)

        replacements: dict[str, bytes] = {}
        with ZipFile(workbook_path) as archive:
            ExcelFillService()._apply_drawing_marks(
                archive,
                {"drawing_marks": [{"path": drawing_path, "xml": mark}]},
                replacements,
            )

        output = replacements[drawing_path].decode("utf-8")
        self.assertIn("Madecentro Mark 1", output)
        self.assertIn(">X<", output)

    def test_preserves_office_compatibility_namespace_prefixes(self) -> None:
        self._add_office_compatibility_namespaces(self.empty_path)
        mapping = {
            "cells": [
                {
                    "sheet": "FORMULARIO",
                    "cell": "B1",
                    "value_type": "string",
                    "master_key": "nit",
                }
            ],
            "controls": [],
        }

        ExcelFillService().fill_workbook(
            self.empty_path,
            self.output_path,
            mapping,
            {"nit": "900123"},
        )

        with ZipFile(self.output_path) as archive:
            sheet_xml = archive.read("xl/worksheets/sheet1.xml")
        self.assertIn(b'xmlns:mc=', sheet_xml)
        self.assertIn(b'xmlns:x14ac=', sheet_xml)
        self.assertIn(b'mc:Ignorable="x14ac"', sheet_xml)
        self.assertNotIn(b'ns0:Ignorable="x14ac"', sheet_xml)

    def _add_office_compatibility_namespaces(self, path: Path) -> None:
        replacement = path.with_suffix(".replacement.xlsx")
        with ZipFile(path) as source, ZipFile(
            replacement,
            "w",
            compression=ZIP_DEFLATED,
        ) as target:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    data = data.replace(
                        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
                        (
                            b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                            b'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
                            b'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/ac" '
                            b'mc:Ignorable="x14ac"'
                        ),
                        1,
                    )
                target.writestr(item, data)
        replacement.replace(path)


if __name__ == "__main__":
    unittest.main()
