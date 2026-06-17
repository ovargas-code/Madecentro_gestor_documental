import tempfile
import unittest
from datetime import date
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import fitz
from lxml import etree
from openpyxl import Workbook, load_workbook

from app.services.form_template_service import FormTemplateService
from app.services.word_signature_service import WordSignatureService
from PIL import Image


WORD_XML = """<?xml version="1.0" encoding="UTF-8"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    <w:tbl>
      <w:tr>
        <w:tc><w:p><w:r><w:t>NIT</w:t></w:r></w:p></w:tc>
        <w:tc><w:p><w:r><w:t>{value}</w:t></w:r></w:p></w:tc>
      </w:tr>
    </w:tbl>
  </w:body>
</w:document>
"""

WORD_INLINE_XML = """<?xml version="1.0" encoding="UTF-8"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    <w:tbl>
      <w:tr>
        <w:tc><w:p><w:r><w:t>{value}</w:t></w:r></w:p></w:tc>
      </w:tr>
    </w:tbl>
  </w:body>
</w:document>
"""

WORD_CHECKBOX_XML = """<?xml version="1.0" encoding="UTF-8"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
 xmlns:w14="http://schemas.microsoft.com/office/word/2010/wordml">
  <w:body>
    <w:sdt>
      <w:sdtPr>
        <w14:checkbox><w14:checked w14:val="0"/></w14:checkbox>
      </w:sdtPr>
      <w:sdtContent><w:r><w:t>☐</w:t></w:r></w:sdtContent>
    </w:sdt>
  </w:body>
</w:document>
"""


class FormTemplateServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.service = FormTemplateService()

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_learns_and_fills_excel_pair(self) -> None:
        empty = self.root / "empty.xlsx"
        completed = self.root / "completed.xlsx"
        output = self.root / "output.xlsx"
        self._create_excel(empty, None)
        self._create_excel(completed, "811.028.650-1")

        payload = self.service.learn(
            empty,
            completed,
            {"nit": "811.028.650-1"},
        )
        self.service.fill(
            empty,
            output,
            payload,
            {"nit": "900.123.456-7"},
        )

        self.assertEqual(
            self.service.mapping(payload)["FORMULARIO!B1"],
            "nit",
        )
        workbook = load_workbook(output, data_only=False)
        try:
            self.assertEqual(
                workbook["FORMULARIO"]["B1"].value,
                "900.123.456-7",
            )
        finally:
            workbook.close()

    def test_excel_learning_gives_distinct_keys_to_ambiguous_labels(self) -> None:
        empty = self.root / "empty.xlsx"
        completed = self.root / "completed.xlsx"
        empty_book = Workbook()
        empty_sheet = empty_book.active
        empty_sheet.title = "FORMULARIO"
        empty_sheet["A1"] = "Dirección"
        empty_sheet["A2"] = "Dirección"
        empty_book.save(empty)
        empty_book.close()

        completed_book = load_workbook(empty)
        completed_sheet = completed_book["FORMULARIO"]
        completed_sheet["B1"] = "Calle principal"
        completed_sheet["B2"] = "Dirección referencia"
        completed_book.save(completed)
        completed_book.close()

        payload = self.service.learn(
            empty,
            completed,
            {"direccion": "Otra dirección"},
        )

        self.assertEqual(
            [field["master_key"] for field in payload["cells"]],
            ["direccion_b1", "direccion_b2"],
        )

    def test_excel_learning_does_not_match_single_digit_samples(self) -> None:
        empty = self.root / "empty.xlsx"
        completed = self.root / "completed.xlsx"
        empty_book = Workbook()
        empty_sheet = empty_book.active
        empty_sheet.title = "FORMULARIO"
        empty_sheet["A1"] = "Código interno"
        empty_book.save(empty)
        empty_book.close()
        completed_book = load_workbook(empty)
        completed_book["FORMULARIO"]["B1"] = 4
        completed_book.save(completed)
        completed_book.close()

        payload = self.service.learn(
            empty,
            completed,
            {"unrelated_code": "4"},
        )

        self.assertEqual(
            payload["cells"][0]["master_key"],
            "codigo_interno_b1",
        )

    def test_learns_and_fills_word_table_pair(self) -> None:
        empty = self.root / "empty.docx"
        completed = self.root / "completed.docx"
        output = self.root / "output.docx"
        self._create_word(empty, "")
        self._create_word(completed, "811.028.650-1")

        payload = self.service.learn(
            empty,
            completed,
            {"nit": "811.028.650-1"},
        )
        self.service.fill(
            empty,
            output,
            payload,
            {"nit": "900.123.456-7"},
        )

        with ZipFile(output) as archive:
            document = archive.read("word/document.xml")
        self.assertIn(b"900.123.456-7", document)

    def test_word_learning_preserves_label_inside_value_cell(self) -> None:
        empty = self.root / "empty.docx"
        completed = self.root / "completed.docx"
        output = self.root / "output.docx"
        self._create_inline_word(empty, "País :")
        self._create_inline_word(completed, "País : COLOMBIA")

        payload = self.service.learn(
            empty,
            completed,
            {"pais": "COLOMBIA"},
        )
        self.service.fill(empty, output, payload, {"pais": "PERÚ"})

        with ZipFile(output) as archive:
            document = archive.read("word/document.xml")
        self.assertIn("País : PERÚ".encode("utf-8"), document)

    def test_word_signature_uses_learned_table_paragraph(self) -> None:
        document = self.root / "signed.docx"
        signature = self.root / "signature.png"
        self._create_inline_word(document, "Firma:")
        Image.new("RGB", (100, 30), "black").save(signature)

        WordSignatureService().apply_signature(
            document,
            signature,
            {"table": 1, "row": 1, "cell": 1, "paragraph": 1},
        )

        with ZipFile(document) as archive:
            root = etree.fromstring(archive.read("word/document.xml"))
        drawing_count = root.xpath(
            "count(.//w:tbl[1]/w:tr[1]/w:tc[1]/w:p[1]//w:drawing)",
            namespaces={
                "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
            },
        )
        self.assertEqual(drawing_count, 1.0)

    def test_word_fill_supports_current_date_transform(self) -> None:
        document = self.root / "date.docx"
        output = self.root / "date-output.docx"
        self._create_inline_word(document, "Fecha: ")
        payload = {
            "fields": [
                {
                    "field_id": "table:1:1:1",
                    "value_transform": "current_date",
                    "date_format": "%d/%m/%Y",
                    "value_format": "Fecha: {value}",
                }
            ],
            "signature": {"enabled": False},
        }

        self.service.fill(document, output, payload, {})

        with ZipFile(output) as archive:
            document_xml = archive.read("word/document.xml").decode("utf-8")
        self.assertIn(f"Fecha: {date.today():%d/%m/%Y}", document_xml)

    def test_word_fill_updates_checkbox_property(self) -> None:
        document = self.root / "checkbox.docx"
        output = self.root / "checkbox-output.docx"
        with ZipFile(document, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr("word/document.xml", WORD_CHECKBOX_XML.encode("utf-8"))
        payload = {
            "fields": [
                {
                    "field_id": "sdt-index:1",
                    "kind": "checkbox",
                    "preserve_reference": True,
                    "reference_value": "☒",
                }
            ],
            "signature": {"enabled": False},
        }

        self.service.fill(document, output, payload, {})

        with ZipFile(output) as archive:
            document_xml = archive.read("word/document.xml").decode("utf-8")
        self.assertIn('checked w14:val="1"', document_xml)
        self.assertIn("☒", document_xml)

    def test_learns_pdf_mapping_from_reference_values(self) -> None:
        empty = self.root / "empty.pdf"
        completed = self.root / "completed.pdf"
        self._create_pdf(empty, "")
        self._create_pdf(completed, "811.028.650-1")

        payload = self.service.learn(
            empty,
            completed,
            {"nit": "811.028.650-1"},
        )

        self.assertEqual(payload["mapping"]["campo_empresa"], "nit")

    def test_learns_pdf_acroform_without_completed_reference(self) -> None:
        template = self.root / "template.pdf"
        self._create_pdf(template, "")

        payload = self.service.learn_pdf_acroform(
            template,
            {"campo_empresa": "MADECENTRO"},
        )

        self.assertEqual(payload["format"], "pdf")
        self.assertEqual(payload["schema_version"], 2)
        self.assertEqual(payload["mapping"]["campo_empresa"], "campo_empresa")
        self.assertTrue(payload["template_fingerprint"])
        self.assertEqual(payload["field_metadata"][0]["field_type"], "text")

    def test_adds_signature_relationship_to_word(self) -> None:
        document = self.root / "signed.docx"
        signature = self.root / "signature.png"
        self._create_word(document, "")
        Image.new("RGB", (100, 30), "black").save(signature)

        WordSignatureService().apply_signature(document, signature)

        with ZipFile(document) as archive:
            self.assertIn(
                "word/media/madecentro_signature.png",
                archive.namelist(),
            )
            self.assertIn(
                b"Firma Madecentro",
                archive.read("word/document.xml"),
            )

    def test_detects_blank_signature_area_above_excel_line(self) -> None:
        workbook_path = self.root / "signature_area.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "FORMULARIO"
        sheet.merge_cells("H18:N19")
        sheet["H18"] = "________________________"
        sheet.merge_cells("G20:O20")
        sheet["G20"] = "Firma"
        workbook.save(workbook_path)
        workbook.close()

        config = self.service._excel_signature_cell(workbook_path)

        self.assertEqual(config["sheet"], "FORMULARIO")
        self.assertEqual(config["cell"], "H14")
        self.assertEqual(config["end_cell"], "N17")

    def _create_excel(self, path: Path, value: object) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "FORMULARIO"
        sheet["A1"] = "NIT"
        sheet["B1"] = value
        workbook.save(path)
        workbook.close()

    def _create_word(self, path: Path, value: str) -> None:
        with ZipFile(path, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr(
                "word/document.xml",
                WORD_XML.format(value=value).encode("utf-8"),
            )

    def _create_inline_word(self, path: Path, value: str) -> None:
        with ZipFile(path, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr(
                "word/document.xml",
                WORD_INLINE_XML.format(value=value).encode("utf-8"),
            )

    def _create_pdf(self, path: Path, value: str) -> None:
        document = fitz.open()
        page = document.new_page()
        widget = fitz.Widget()
        widget.field_name = "campo_empresa"
        widget.field_type = fitz.PDF_WIDGET_TYPE_TEXT
        widget.field_value = value
        widget.rect = fitz.Rect(20, 20, 200, 45)
        page.add_widget(widget)
        document.save(path)
        document.close()


if __name__ == "__main__":
    unittest.main()
