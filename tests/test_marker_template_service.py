from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.marker_template_service import MarkerTemplateService


class MarkerTemplateServiceTests(unittest.TestCase):
    def test_builds_excel_payload_and_removes_markers_from_template_copy(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "generada.xlsx"
            target = root / "plantilla.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "FORMATO PJ"
            sheet["A1"] = "DIRECCIÓN <<DIRECCION>>"
            sheet["B1"] = "<<NIT>>"
            workbook.save(source)
            workbook.close()

            payload = MarkerTemplateService().build_excel_payload(
                source,
                target,
                ["direccion", "nit"],
            )

            result = load_workbook(target)
            try:
                self.assertEqual(result["FORMATO PJ"]["A1"].value, "DIRECCIÓN")
                self.assertIsNone(result["FORMATO PJ"]["B1"].value)
            finally:
                result.close()

        mapping = {
            field["field_id"]: field
            for field in payload["cells"]
        }
        self.assertEqual(mapping["FORMATO PJ!A1"]["master_key"], "direccion")
        self.assertEqual(mapping["FORMATO PJ!A1"]["value_format"], "DIRECCIÓN {value}")
        self.assertEqual(mapping["FORMATO PJ!B1"]["master_key"], "nit")
        self.assertNotIn("value_format", mapping["FORMATO PJ!B1"])


if __name__ == "__main__":
    unittest.main()
