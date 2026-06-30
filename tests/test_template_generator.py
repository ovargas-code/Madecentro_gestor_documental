from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.plantilla_generator.dictionary_loader import DictionaryLoader
from app.plantilla_generator.generator import TemplateGenerator


class TemplateGeneratorTests(unittest.TestCase):
    def test_loads_json_dictionary_with_default_marker(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "diccionario.json"
            path.write_text(
                json.dumps(
                    [
                        {
                            "categoria": "nit",
                            "valor": "811.028.650-1",
                            "reemplazo": "",
                            "confianza": "alto",
                        }
                    ]
                ),
                encoding="utf-8",
            )

            rules = DictionaryLoader().load(path, mode="markers")

        self.assertEqual(len(rules), 1)
        self.assertEqual(rules[0].replacement, "<<NIT>>")

    def test_loads_xlsx_dictionary_with_custom_replacement(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "diccionario.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["categoria", "valor", "reemplazo", "confianza"])
            sheet.append(["correo", "notificaciones@madecentro.co", "<<EMAIL>>", "alto"])
            workbook.save(path)
            workbook.close()

            rules = DictionaryLoader().load(path, mode="markers")

        self.assertEqual(rules[0].replacement, "<<EMAIL>>")

    def test_generates_clean_excel_copy_and_report(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "entrada"
            output_dir = root / "salida"
            input_dir.mkdir()
            dictionary_path = root / "diccionario.json"
            source_path = input_dir / "formulario.xlsx"

            dictionary_path.write_text(
                json.dumps(
                    [
                        {
                            "categoria": "razon_social",
                            "valor": "MADECENTRO COLOMBIA SAS",
                            "reemplazo": "",
                            "confianza": "alto",
                        }
                    ]
                ),
                encoding="utf-8",
            )
            workbook = Workbook()
            workbook.active["A1"] = "Cliente: MADECENTRO COLOMBIA SAS"
            workbook.save(source_path)
            workbook.close()

            rows = TemplateGenerator().run(
                input_dir,
                dictionary_path,
                output_dir,
                mode="markers",
            )

            result = load_workbook(output_dir / "formulario.xlsx")
            try:
                self.assertEqual(
                    result.active["A1"].value,
                    "Cliente: <<RAZON_SOCIAL>>",
                )
            finally:
                result.close()
            self.assertEqual(rows[0].estado, "ok")
            self.assertEqual(rows[0].cantidad_reemplazos, 1)
            self.assertTrue((output_dir / "reporte_generacion_plantillas.xlsx").is_file())


if __name__ == "__main__":
    unittest.main()
